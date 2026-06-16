"""Render the consolidated workbook's Executive Summary KPI strip to a PNG.

A quick visual-QA helper for the Excel Exec Summary: it reads the actual cell
values from ``consolidated_risk_production.xlsx`` and draws the Main + MR KPI
cards (Production / Risk / Rejection Rate / System Rejection Rate / Production
Delta) so the layout can be eyeballed without opening Excel or LibreOffice.

Usage:
    uv run python scripts/preview_exec_summary_kpis.py
    uv run python scripts/preview_exec_summary_kpis.py -i path/to.xlsx -o preview.png
"""

from __future__ import annotations

import argparse
from pathlib import Path

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt  # noqa: E402
import openpyxl  # noqa: E402
from matplotlib.patches import FancyBboxPatch, Rectangle  # noqa: E402

# KPI cards occupy 2-column blocks; value on the first row, label on the next.
CARD_COLS = (1, 3, 5, 7, 9)
MAIN_VALUE_ROW, MAIN_LABEL_ROW = 4, 5
MR_VALUE_ROW, MR_LABEL_ROW = 6, 7

_KPI_BLUE = "#D6EAF8"
_MR_BLUE = "#EAF2F8"
_ACCENT = "#2E86C1"
_EDGE = "#1F618D"


def _cards(ws, value_row: int, label_row: int) -> list[tuple[str, str]]:
    out = []
    for col in CARD_COLS:
        val = ws.cell(row=value_row, column=col).value
        lab = ws.cell(row=label_row, column=col).value
        out.append(("" if val is None else str(val), "" if lab is None else str(lab)))
    return out


def _draw_row(ax, cards: list[tuple[str, str]], y: float, period: str, tint: str) -> None:
    ax.text(-0.02, y + 0.42, period, fontsize=9, fontweight="bold", color=_EDGE, ha="right", va="center")
    for i, (val, lab) in enumerate(cards):
        ax.add_patch(
            FancyBboxPatch(
                (i + 0.04, y), 0.92, 0.82, boxstyle="round,pad=0.01", facecolor=tint, edgecolor=_EDGE, linewidth=1.2
            )
        )
        ax.add_patch(Rectangle((i + 0.04, y), 0.025, 0.82, color=_ACCENT))
        ax.text(i + 0.5, y + 0.55, val, fontsize=12, fontweight="bold", ha="center", va="center", color="#0f172a")
        ax.text(i + 0.5, y + 0.18, lab, fontsize=7.5, ha="center", va="center", color="#475569")


def render(input_path: Path, output_path: Path) -> Path:
    wb = openpyxl.load_workbook(input_path, data_only=True)
    if "Executive Summary" not in wb.sheetnames:
        raise SystemExit(f"'Executive Summary' sheet not found in {input_path}")
    ws = wb["Executive Summary"]

    fig, ax = plt.subplots(figsize=(15, 4.2))
    ax.set_xlim(0, len(CARD_COLS))
    ax.set_ylim(0, 2.4)
    ax.axis("off")
    _draw_row(ax, _cards(ws, MAIN_VALUE_ROW, MAIN_LABEL_ROW), 1.3, "MAIN", _KPI_BLUE)
    _draw_row(ax, _cards(ws, MR_VALUE_ROW, MR_LABEL_ROW), 0.3, "MR", _MR_BLUE)
    ax.set_title(
        "Executive Summary — KPI cards (rendered from actual cell values)",
        fontsize=11,
        fontweight="bold",
        loc="left",
    )
    fig.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=110, bbox_inches="tight")
    plt.close(fig)
    return output_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Render the Exec Summary KPI strip to a PNG for visual QA.")
    parser.add_argument(
        "-i",
        "--input",
        default="output/consolidated_risk_production.xlsx",
        help="Consolidated workbook (default: output/consolidated_risk_production.xlsx).",
    )
    parser.add_argument(
        "-o",
        "--output",
        default="output/exec_kpi_preview.png",
        help="Output PNG path (default: output/exec_kpi_preview.png).",
    )
    args = parser.parse_args()
    out = render(Path(args.input), Path(args.output))
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
