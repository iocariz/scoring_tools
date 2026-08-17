"""
Tests for consolidation module and scenario generation.

Tests cover:
- Scenario naming (pessimistic, base, optimistic)
- Scenario suffix detection from filenames
- Metrics extraction from summary tables
- Metrics aggregation across segments
- Risk calculation (sum(todu_30ever_h6) / sum(todu_amt_pile_h6) * 7)
- Percentage formatting
- Full consolidation workflow
"""

import os
import sys

sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import tempfile
from pathlib import Path

import numpy as np
import pandas as pd
import pytest

from src.consolidation import (
    ConsolidatedMetrics,
    _get_total_row,
    aggregate_metrics,
    consolidate_segments,
    create_consolidation_dashboard,
    export_consolidated_excel,
    extract_metrics_from_table,
    find_scenario_suffix,
    map_scenario_names,
    patch_consolidated_production_from_segment_audits,
)

# =============================================================================
# Fixtures
# =============================================================================


@pytest.fixture
def sample_summary_table():
    """Create a sample risk_production_summary_table."""
    return pd.DataFrame(
        {
            "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
            "Risk (%)": [1.5, 1.2, 2.0, 1.4],
            "Production (€)": [1000000, 200000, 150000, 1050000],
            "Production (%)": [1.0, 0.2, 0.15, 1.05],
            "todu_30ever_h6": [1000, 200, 300, 900],
            "todu_amt_pile_h6": [50000, 10000, 10000, 45000],
        }
    )


@pytest.fixture
def temp_output_dir():
    """Create a temporary directory with mock segment data."""
    with tempfile.TemporaryDirectory() as tmpdir:
        yield Path(tmpdir)


# =============================================================================
# Scenario Suffix Detection Tests
# =============================================================================


class TestFindScenarioSuffix:
    """Tests for find_scenario_suffix function."""

    def test_named_pessimistic(self):
        """Test detection of pessimistic scenario."""
        result = find_scenario_suffix("risk_production_summary_table_pessimistic.csv")
        assert result == "_pessimistic"

    def test_named_base(self):
        """Test detection of base scenario."""
        result = find_scenario_suffix("risk_production_summary_table_base.csv")
        assert result == "_base"

    def test_named_optimistic(self):
        """Test detection of optimistic scenario."""
        result = find_scenario_suffix("risk_production_summary_table_optimistic.csv")
        assert result == "_optimistic"

    def test_legacy_numeric(self):
        """Test detection of legacy numeric scenarios."""
        assert find_scenario_suffix("risk_production_summary_table_1.0.csv") == "_1.0"
        assert find_scenario_suffix("risk_production_summary_table_0.9.csv") == "_0.9"
        assert find_scenario_suffix("risk_production_summary_table_1.1.csv") == "_1.1"

    def test_no_suffix(self):
        """Test file without scenario suffix."""
        result = find_scenario_suffix("risk_production_summary_table.csv")
        assert result == ""

    def test_mr_file_with_scenario(self):
        """Test MR file with scenario suffix."""
        result = find_scenario_suffix("risk_production_summary_table_mr_pessimistic.csv")
        assert result == "_pessimistic"

    def test_case_insensitive(self):
        """Test case insensitivity for named scenarios."""
        result = find_scenario_suffix("risk_production_summary_table_PESSIMISTIC.csv")
        assert result == "_pessimistic"


# =============================================================================
# Scenario Name Mapping Tests
# =============================================================================


class TestMapScenarioNames:
    """Tests for map_scenario_names function."""

    def test_named_scenarios(self):
        """Test mapping of named scenario suffixes."""
        suffixes = ["_pessimistic", "_base", "_optimistic"]
        mapping = map_scenario_names(suffixes)

        assert mapping["_pessimistic"] == "pessimistic"
        assert mapping["_base"] == "base"
        assert mapping["_optimistic"] == "optimistic"

    def test_legacy_numeric_three_scenarios(self):
        """Test mapping of three legacy numeric scenarios."""
        suffixes = ["_0.9", "_1.0", "_1.1"]
        mapping = map_scenario_names(suffixes)

        assert mapping["_0.9"] == "pessimistic"
        assert mapping["_1.0"] == "base"
        assert mapping["_1.1"] == "optimistic"

    def test_legacy_numeric_two_scenarios(self):
        """Test mapping of two legacy numeric scenarios."""
        suffixes = ["_0.9", "_1.1"]
        mapping = map_scenario_names(suffixes)

        assert mapping["_0.9"] == "pessimistic"
        assert mapping["_1.1"] == "optimistic"

    def test_empty_suffix_as_base(self):
        """Test that empty suffix maps to base."""
        suffixes = ["", "_pessimistic", "_optimistic"]
        mapping = map_scenario_names(suffixes)

        assert mapping[""] == "base"
        assert mapping["_pessimistic"] == "pessimistic"
        assert mapping["_optimistic"] == "optimistic"

    def test_single_scenario(self):
        """Test mapping of single scenario."""
        suffixes = ["_1.0"]
        mapping = map_scenario_names(suffixes)

        assert mapping["_1.0"] == "base"

    def test_four_or_more_middles_warn_on_collapse(self, caplog):
        """>3 legacy numeric scenarios: extremes map to pessimistic/optimistic and every
        middle collapses to 'base' (they overwrite each other). This must be WARNED, not silent."""
        import logging

        from src import consolidation

        suffixes = ["_0.8", "_1.0", "_1.2", "_1.4"]  # two middles both -> base
        # loguru -> caplog bridge for this call
        handler_id = consolidation.logger.add(caplog.handler, format="{message}", level="WARNING")
        try:
            with caplog.at_level(logging.WARNING):
                mapping = map_scenario_names(suffixes)
        finally:
            consolidation.logger.remove(handler_id)

        assert mapping["_0.8"] == "pessimistic"
        assert mapping["_1.4"] == "optimistic"
        assert mapping["_1.0"] == "base"
        assert mapping["_1.2"] == "base"  # both middles collapse
        assert any("collapse" in rec.message for rec in caplog.records), "expected a collapse warning"


# =============================================================================
# Metrics Extraction Tests
# =============================================================================


class TestExtractMetricsFromTable:
    """Tests for extract_metrics_from_table function."""

    def test_basic_extraction(self, sample_summary_table):
        """Test basic metrics extraction."""
        metrics = extract_metrics_from_table(sample_summary_table)

        assert "actual" in metrics
        assert "optimum" in metrics
        assert "swap_in" in metrics
        assert "swap_out" in metrics

    def test_production_values(self, sample_summary_table):
        """Test production value extraction."""
        metrics = extract_metrics_from_table(sample_summary_table)

        assert metrics["actual"]["production"] == 1000000
        assert metrics["optimum"]["production"] == 1050000
        assert metrics["swap_in"]["production"] == 200000
        assert metrics["swap_out"]["production"] == 150000

    def test_todu_values(self, sample_summary_table):
        """Test todu value extraction."""
        metrics = extract_metrics_from_table(sample_summary_table)

        assert metrics["actual"]["todu_30ever_h6"] == 1000
        assert metrics["actual"]["todu_amt_pile_h6"] == 50000
        assert metrics["optimum"]["todu_30ever_h6"] == 900
        assert metrics["optimum"]["todu_amt_pile_h6"] == 45000

    def test_empty_dataframe(self):
        """Test handling of empty DataFrame."""
        metrics = extract_metrics_from_table(pd.DataFrame())

        assert metrics["actual"]["production"] == 0
        assert metrics["actual"]["todu_30ever_h6"] == 0

    def test_none_input(self):
        """Test handling of None input."""
        metrics = extract_metrics_from_table(None)

        assert metrics["actual"]["production"] == 0


# =============================================================================
# Metrics Aggregation Tests
# =============================================================================


class TestAggregateMetrics:
    """Tests for aggregate_metrics function."""

    def test_single_segment(self):
        """Test aggregation of single segment."""
        metrics = [
            {
                "actual": {"production": 1000, "todu_30ever_h6": 100, "todu_amt_pile_h6": 5000},
                "optimum": {"production": 1100, "todu_30ever_h6": 90, "todu_amt_pile_h6": 5000},
                "swap_in": {"production": 200, "todu_30ever_h6": 10, "todu_amt_pile_h6": 500},
                "swap_out": {"production": 100, "todu_30ever_h6": 20, "todu_amt_pile_h6": 500},
            }
        ]

        result = aggregate_metrics(metrics)

        assert result["actual"]["production"] == 1000
        assert result["actual"]["todu_30ever_h6"] == 100

    def test_multiple_segments(self):
        """Test aggregation of multiple segments."""
        metrics = [
            {
                "actual": {"production": 1000, "todu_30ever_h6": 100, "todu_amt_pile_h6": 5000},
                "optimum": {"production": 1100, "todu_30ever_h6": 90, "todu_amt_pile_h6": 5000},
                "swap_in": {"production": 200, "todu_30ever_h6": 10, "todu_amt_pile_h6": 500},
                "swap_out": {"production": 100, "todu_30ever_h6": 20, "todu_amt_pile_h6": 500},
            },
            {
                "actual": {"production": 2000, "todu_30ever_h6": 200, "todu_amt_pile_h6": 10000},
                "optimum": {"production": 2200, "todu_30ever_h6": 180, "todu_amt_pile_h6": 10000},
                "swap_in": {"production": 400, "todu_30ever_h6": 20, "todu_amt_pile_h6": 1000},
                "swap_out": {"production": 200, "todu_30ever_h6": 40, "todu_amt_pile_h6": 1000},
            },
        ]

        result = aggregate_metrics(metrics)

        # Check sums
        assert result["actual"]["production"] == 3000
        assert result["actual"]["todu_30ever_h6"] == 300
        assert result["actual"]["todu_amt_pile_h6"] == 15000
        assert result["optimum"]["production"] == 3300

    def test_empty_list(self):
        """Test aggregation of empty list."""
        result = aggregate_metrics([])

        assert result["actual"]["production"] == 0
        assert result["actual"]["todu_30ever_h6"] == 0

    def test_production_ci_aggregates_using_variance_addition(self):
        """Test that aggregate production CI uses variance addition (independence assumption)."""
        metrics = [
            {
                "actual": {"production": 1000, "todu_30ever_h6": 10, "todu_amt_pile_h6": 100},
                "optimum": {
                    "production": 1100,
                    "todu_30ever_h6": 9,
                    "todu_amt_pile_h6": 100,
                    "production_ci_lower": 1000,
                    "production_ci_upper": 1200,
                    "risk_ci_lower": 8,
                    "risk_ci_upper": 10,
                },
                "swap_in": {"production": 200, "todu_30ever_h6": 1, "todu_amt_pile_h6": 20},
                "swap_out": {"production": 100, "todu_30ever_h6": 2, "todu_amt_pile_h6": 20},
            },
            {
                "actual": {"production": 2000, "todu_30ever_h6": 20, "todu_amt_pile_h6": 200},
                "optimum": {
                    "production": 2200,
                    "todu_30ever_h6": 18,
                    "todu_amt_pile_h6": 200,
                    "production_ci_lower": 2100,
                    "production_ci_upper": 2300,
                    "risk_ci_lower": 8,
                    "risk_ci_upper": 10,
                },
                "swap_in": {"production": 400, "todu_30ever_h6": 2, "todu_amt_pile_h6": 40},
                "swap_out": {"production": 200, "todu_30ever_h6": 4, "todu_amt_pile_h6": 40},
            },
        ]

        result = aggregate_metrics(metrics, multiplier=1)

        # Variance addition: combined SE = sqrt(SE1² + SE2²), not sum of bounds
        # Segment 1: CI width=200, SE=200/(2*1.96)≈51.02
        # Segment 2: CI width=200, SE=200/(2*1.96)≈51.02
        # Combined SE = sqrt(51.02² + 51.02²) ≈ 72.15
        # Aggregate production = 3300
        # Lower = 3300 - 1.96*72.15 ≈ 3158.6
        # Upper = 3300 + 1.96*72.15 ≈ 3441.4
        import math

        z = 1.96
        se1 = 200 / (2 * z)
        se2 = 200 / (2 * z)
        combined_se = math.sqrt(se1**2 + se2**2)
        expected_lower = 3300 - z * combined_se
        expected_upper = 3300 + z * combined_se
        assert abs(result["optimum"]["production_ci_lower"] - expected_lower) < 0.5
        assert abs(result["optimum"]["production_ci_upper"] - expected_upper) < 0.5

    def test_risk_ci_uses_exposure_weighted_delta_method(self):
        """Test that aggregate risk CI is weighted by segment exposure, not added directly."""
        metrics = [
            {
                "actual": {"production": 1000, "todu_30ever_h6": 20, "todu_amt_pile_h6": 100},
                "optimum": {
                    "production": 1100,
                    "todu_30ever_h6": 20,
                    "todu_amt_pile_h6": 100,
                    "production_ci_lower": 1050,
                    "production_ci_upper": 1150,
                    "risk_ci_lower": 10.0,
                    "risk_ci_upper": 30.0,
                },
                "swap_in": {"production": 200, "todu_30ever_h6": 1, "todu_amt_pile_h6": 20},
                "swap_out": {"production": 100, "todu_30ever_h6": 2, "todu_amt_pile_h6": 20},
            },
            {
                "actual": {"production": 2000, "todu_30ever_h6": 90, "todu_amt_pile_h6": 900},
                "optimum": {
                    "production": 2200,
                    "todu_30ever_h6": 90,
                    "todu_amt_pile_h6": 900,
                    "production_ci_lower": 2150,
                    "production_ci_upper": 2250,
                    "risk_ci_lower": 8.0,
                    "risk_ci_upper": 12.0,
                },
                "swap_in": {"production": 400, "todu_30ever_h6": 2, "todu_amt_pile_h6": 40},
                "swap_out": {"production": 200, "todu_30ever_h6": 4, "todu_amt_pile_h6": 40},
            },
        ]

        result = aggregate_metrics(metrics, multiplier=1)

        z_95 = 1.96
        seg1_se = (30.0 - 10.0) / (2 * z_95)
        seg2_se = (12.0 - 8.0) / (2 * z_95)
        expected_point = (20 + 90) / (100 + 900) * 100
        expected_se = np.sqrt((0.1 * seg1_se) ** 2 + (0.9 * seg2_se) ** 2)

        assert np.isclose(result["optimum"]["risk_ci_lower"], expected_point - z_95 * expected_se)
        assert np.isclose(result["optimum"]["risk_ci_upper"], expected_point + z_95 * expected_se)


# =============================================================================
# ConsolidatedMetrics Tests
# =============================================================================


class TestConsolidatedMetrics:
    """Tests for ConsolidatedMetrics dataclass."""

    def test_risk_calculation(self):
        """Test that risk is calculated correctly from todu values."""
        metrics = ConsolidatedMetrics(
            group_name="test",
            period="main",
            scenario="base",
            segments=["seg1"],
            actual_production=1000000,
            actual_todu_30ever_h6=100,
            actual_todu_amt_pile_h6=10000,
        )

        # Risk = todu_30ever_h6 / todu_amt_pile_h6 * 7 * 100 (percentage)
        # = 100 / 10000 * 7 * 100 = 7.0
        assert np.isclose(metrics.actual_risk, 7.0)

    def test_risk_percentage_in_dict(self):
        """Test that risk is converted to percentage in to_dict()."""
        metrics = ConsolidatedMetrics(
            group_name="test",
            period="main",
            scenario="base",
            segments=["seg1"],
            actual_production=1000000,
            actual_todu_30ever_h6=100,
            actual_todu_amt_pile_h6=10000,
        )

        result = metrics.to_dict()

        # actual_risk property returns percentage directly: 7.0%
        assert np.isclose(result["actual_risk_pct"], 7.0)

    def test_zero_denominator(self):
        """Test handling of zero denominator in risk calculation."""
        metrics = ConsolidatedMetrics(
            group_name="test",
            period="main",
            scenario="base",
            segments=["seg1"],
            actual_production=1000000,
            actual_todu_30ever_h6=100,
            actual_todu_amt_pile_h6=0,  # Zero denominator
        )

        assert metrics.actual_risk == 0.0

    def test_production_delta(self):
        """Test production delta calculation."""
        metrics = ConsolidatedMetrics(
            group_name="test",
            period="main",
            scenario="base",
            segments=["seg1"],
            actual_production=1000000,
            optimum_production=1100000,
        )

        assert metrics.production_delta == 100000
        assert np.isclose(metrics.production_delta_pct, 0.1)

    def test_production_delta_pct_in_dict(self):
        """Test production delta percentage is in correct format."""
        metrics = ConsolidatedMetrics(
            group_name="test",
            period="main",
            scenario="base",
            segments=["seg1"],
            actual_production=1000000,
            optimum_production=1100000,
        )

        result = metrics.to_dict()

        # 0.1 * 100 = 10%
        assert np.isclose(result["production_delta_pct"], 10.0)

    def test_rejection_rate_in_dict(self):
        """Test rejection rate is computed correctly in to_dict()."""
        metrics = ConsolidatedMetrics(
            group_name="test",
            period="main",
            scenario="base",
            segments=["seg1"],
            actual_production=800000,
            optimum_production=900000,
            total_demand=1000000,
        )

        result = metrics.to_dict()

        # actual rejection = (1 - 800k/1M) * 100 = 20%
        assert np.isclose(result["actual_rejection_rate_pct"], 20.0)
        # optimum rejection = (1 - 900k/1M) * 100 = 10%
        assert np.isclose(result["optimum_rejection_rate_pct"], 10.0)

    def test_rejection_rate_zero_demand(self):
        """Test rejection rate is 0 when total_demand is 0."""
        metrics = ConsolidatedMetrics(
            group_name="test",
            period="main",
            scenario="base",
            segments=["seg1"],
            actual_production=0,
            optimum_production=0,
            total_demand=0,
        )

        assert metrics.actual_rejection_rate == 0.0
        assert metrics.optimum_rejection_rate == 0.0


class TestExtractMetricsRejectionRate:
    """Tests for rejection rate extraction from summary tables."""

    def test_total_demand_derived_from_rejection_rate(self):
        """Test that total_demand is derived from Actual row's rejection rate."""
        df = pd.DataFrame(
            {
                "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
                "Risk (%)": [1.5, 1.2, 2.0, 1.4],
                "Production (€)": [800000, 100000, 50000, 850000],
                "Production (%)": [1.0, 0.125, 0.0625, 1.0625],
                "todu_30ever_h6": [1000, 200, 300, 900],
                "todu_amt_pile_h6": [50000, 10000, 10000, 45000],
                "Rejection Rate (%)": [20.0, None, None, 15.0],
            }
        )

        metrics = extract_metrics_from_table(df)

        # total_demand = 800000 / (1 - 20/100) = 800000 / 0.8 = 1000000
        assert np.isclose(metrics["_total_demand"], 1000000)

    def test_total_demand_from_explicit_column(self):
        """Test that total_demand is read from 'Total Demand (€)' column when present."""
        df = pd.DataFrame(
            {
                "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
                "Risk (%)": [1.5, 1.2, 2.0, 1.4],
                "Production (€)": [800000, 100000, 50000, 850000],
                "Production (%)": [1.0, 0.125, 0.0625, 1.0625],
                "todu_30ever_h6": [1000, 200, 300, 900],
                "todu_amt_pile_h6": [50000, 10000, 10000, 45000],
                "Rejection Rate (%)": [20.0, None, None, 15.0],
                "Total Demand (€)": [1200000, None, None, None],
            }
        )

        metrics = extract_metrics_from_table(df)

        # Should use explicit column, not derive from rejection rate
        assert np.isclose(metrics["_total_demand"], 1200000)

    def test_total_demand_fallback_without_rejection_rate(self, sample_summary_table):
        """Test that total_demand defaults to actual_production without rejection rate column."""
        metrics = extract_metrics_from_table(sample_summary_table)

        # No rejection_rate column → total_demand = actual_production
        assert np.isclose(metrics["_total_demand"], 1000000)


class TestAggregateMetricsTotalDemand:
    """Tests for total_demand aggregation."""

    def test_total_demand_summed(self):
        """Test total_demand is summed across segments."""
        metrics_list = [
            {
                "actual": {"production": 500, "todu_30ever_h6": 50, "todu_amt_pile_h6": 5000},
                "optimum": {"production": 600, "todu_30ever_h6": 45, "todu_amt_pile_h6": 5000},
                "swap_in": {"production": 150, "todu_30ever_h6": 5, "todu_amt_pile_h6": 500},
                "swap_out": {"production": 50, "todu_30ever_h6": 10, "todu_amt_pile_h6": 500},
                "_total_demand": 1000,
            },
            {
                "actual": {"production": 300, "todu_30ever_h6": 30, "todu_amt_pile_h6": 3000},
                "optimum": {"production": 350, "todu_30ever_h6": 25, "todu_amt_pile_h6": 3000},
                "swap_in": {"production": 80, "todu_30ever_h6": 3, "todu_amt_pile_h6": 300},
                "swap_out": {"production": 30, "todu_30ever_h6": 8, "todu_amt_pile_h6": 300},
                "_total_demand": 600,
            },
        ]

        result = aggregate_metrics(metrics_list)

        assert np.isclose(result["_total_demand"], 1600)


# =============================================================================
# Full Consolidation Workflow Tests
# =============================================================================


class TestConsolidateSegments:
    """Tests for consolidate_segments function."""

    def test_all_scenarios_included(self, temp_output_dir):
        """Test that all three scenarios are included in output."""
        # Create mock segment with all scenarios
        seg_dir = temp_output_dir / "test-segment" / "data"
        seg_dir.mkdir(parents=True)

        for scenario in ["pessimistic", "base", "optimistic"]:
            df = pd.DataFrame(
                {
                    "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
                    "Risk (%)": [1.5, 1.2, 2.0, 1.4],
                    "Production (€)": [1000000, 200000, 150000, 1050000],
                    "Production (%)": [1.0, 0.2, 0.15, 1.05],
                    "todu_30ever_h6": [1000, 200, 300, 900],
                    "todu_amt_pile_h6": [50000, 10000, 10000, 45000],
                }
            )
            df.to_csv(seg_dir / f"risk_production_summary_table_{scenario}.csv", index=False)
            df.to_csv(seg_dir / f"risk_production_summary_table_mr_{scenario}.csv", index=False)

        segments = {"test-segment": {"name": "test-segment"}}
        result = consolidate_segments(temp_output_dir, segments, {})

        scenarios_found = result["scenario"].unique().tolist()
        assert "pessimistic" in scenarios_found
        assert "base" in scenarios_found
        assert "optimistic" in scenarios_found

    def test_both_periods_included(self, temp_output_dir):
        """Test that both main and MR periods are included."""
        seg_dir = temp_output_dir / "test-segment" / "data"
        seg_dir.mkdir(parents=True)

        df = pd.DataFrame(
            {
                "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
                "Risk (%)": [1.5, 1.2, 2.0, 1.4],
                "Production (€)": [1000000, 200000, 150000, 1050000],
                "Production (%)": [1.0, 0.2, 0.15, 1.05],
                "todu_30ever_h6": [1000, 200, 300, 900],
                "todu_amt_pile_h6": [50000, 10000, 10000, 45000],
            }
        )
        df.to_csv(seg_dir / "risk_production_summary_table_base.csv", index=False)
        df.to_csv(seg_dir / "risk_production_summary_table_mr_base.csv", index=False)

        segments = {"test-segment": {"name": "test-segment"}}
        result = consolidate_segments(temp_output_dir, segments, {})

        periods_found = result["period"].unique().tolist()
        assert "main" in periods_found
        assert "mr" in periods_found

    def test_supersegment_aggregation(self, temp_output_dir):
        """Test that segments in a supersegment are aggregated."""
        # Create two segments in same supersegment
        for seg_name in ["seg-a", "seg-b"]:
            seg_dir = temp_output_dir / seg_name / "data"
            seg_dir.mkdir(parents=True)

            df = pd.DataFrame(
                {
                    "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
                    "Risk (%)": [1.5, 1.2, 2.0, 1.4],
                    "Production (€)": [1000000, 200000, 150000, 1050000],
                    "Production (%)": [1.0, 0.2, 0.15, 1.05],
                    "todu_30ever_h6": [1000, 200, 300, 900],
                    "todu_amt_pile_h6": [50000, 10000, 10000, 45000],
                }
            )
            df.to_csv(seg_dir / "risk_production_summary_table_base.csv", index=False)

        segments = {
            "seg-a": {"name": "seg-a", "supersegment": "super1"},
            "seg-b": {"name": "seg-b", "supersegment": "super1"},
        }
        supersegments = {"super1": {}}

        result = consolidate_segments(temp_output_dir, segments, supersegments)

        # Should have supersegment total
        ss_rows = result[result["group"] == "supersegment_super1"]
        assert len(ss_rows) > 0

        # Supersegment production should be sum of both segments
        ss_main = ss_rows[ss_rows["period"] == "main"]
        assert ss_main["actual_production"].values[0] == 2000000

    def test_individual_segments_in_supersegment(self, temp_output_dir):
        """Test that individual segments are shown even when in supersegment."""
        for seg_name in ["seg-a", "seg-b"]:
            seg_dir = temp_output_dir / seg_name / "data"
            seg_dir.mkdir(parents=True)

            df = pd.DataFrame(
                {
                    "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
                    "Risk (%)": [1.5, 1.2, 2.0, 1.4],
                    "Production (€)": [1000000, 200000, 150000, 1050000],
                    "Production (%)": [1.0, 0.2, 0.15, 1.05],
                    "todu_30ever_h6": [1000, 200, 300, 900],
                    "todu_amt_pile_h6": [50000, 10000, 10000, 45000],
                }
            )
            df.to_csv(seg_dir / "risk_production_summary_table_base.csv", index=False)

        segments = {
            "seg-a": {"name": "seg-a", "supersegment": "super1"},
            "seg-b": {"name": "seg-b", "supersegment": "super1"},
        }
        supersegments = {"super1": {}}

        result = consolidate_segments(temp_output_dir, segments, supersegments)

        # Should have individual segment rows with supersegment prefix
        groups = result["group"].unique().tolist()
        assert "super1/seg-a" in groups
        assert "super1/seg-b" in groups

    def test_total_row_included(self, temp_output_dir):
        """Test that TOTAL row is included."""
        seg_dir = temp_output_dir / "test-segment" / "data"
        seg_dir.mkdir(parents=True)

        df = pd.DataFrame(
            {
                "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
                "Risk (%)": [1.5, 1.2, 2.0, 1.4],
                "Production (€)": [1000000, 200000, 150000, 1050000],
                "Production (%)": [1.0, 0.2, 0.15, 1.05],
                "todu_30ever_h6": [1000, 200, 300, 900],
                "todu_amt_pile_h6": [50000, 10000, 10000, 45000],
            }
        )
        df.to_csv(seg_dir / "risk_production_summary_table_base.csv", index=False)

        segments = {"test-segment": {"name": "test-segment"}}
        result = consolidate_segments(temp_output_dir, segments, {})

        assert "TOTAL" in result["group"].values

    def test_risk_aggregation_formula(self, temp_output_dir):
        """Test that aggregated risk uses correct formula: sum(num)/sum(den)*7."""
        # Create two segments with different risk profiles
        seg1_dir = temp_output_dir / "seg1" / "data"
        seg1_dir.mkdir(parents=True)
        seg2_dir = temp_output_dir / "seg2" / "data"
        seg2_dir.mkdir(parents=True)

        # Segment 1: todu_30ever_h6=100, todu_amt_pile_h6=5000 -> risk = 0.14
        df1 = pd.DataFrame(
            {
                "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
                "Risk (%)": [0.14, 0.1, 0.2, 0.12],
                "Production (€)": [1000000, 200000, 150000, 1050000],
                "Production (%)": [1.0, 0.2, 0.15, 1.05],
                "todu_30ever_h6": [100, 20, 30, 90],
                "todu_amt_pile_h6": [5000, 1000, 1000, 4500],
            }
        )
        df1.to_csv(seg1_dir / "risk_production_summary_table_base.csv", index=False)

        # Segment 2: todu_30ever_h6=200, todu_amt_pile_h6=10000 -> risk = 0.14
        df2 = pd.DataFrame(
            {
                "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
                "Risk (%)": [0.14, 0.1, 0.2, 0.12],
                "Production (€)": [2000000, 400000, 300000, 2100000],
                "Production (%)": [1.0, 0.2, 0.15, 1.05],
                "todu_30ever_h6": [200, 40, 60, 180],
                "todu_amt_pile_h6": [10000, 2000, 2000, 9000],
            }
        )
        df2.to_csv(seg2_dir / "risk_production_summary_table_base.csv", index=False)

        segments = {
            "seg1": {"name": "seg1"},
            "seg2": {"name": "seg2"},
        }
        result = consolidate_segments(temp_output_dir, segments, {})

        # Get TOTAL row for main period
        total_row = result[(result["group"] == "TOTAL") & (result["period"] == "main")]

        # Aggregated: todu_30ever_h6 = 100+200=300, todu_amt_pile_h6 = 5000+10000=15000
        # Risk = 300/15000*7 = 0.14 -> 14% in output
        expected_risk_pct = (300 / 15000 * 7) * 100  # 14.0
        assert np.isclose(total_row["actual_risk_pct"].values[0], expected_risk_pct, rtol=0.01)

    def test_auto_detect_scenarios_across_all_segments(self, temp_output_dir):
        """Test that scenario auto-detection scans all segments, not just the first one."""
        seg1_dir = temp_output_dir / "seg1" / "data"
        seg1_dir.mkdir(parents=True)
        seg2_dir = temp_output_dir / "seg2" / "data"
        seg2_dir.mkdir(parents=True)

        df = pd.DataFrame(
            {
                "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
                "Risk (%)": [1.5, 1.2, 2.0, 1.4],
                "Production (€)": [1000000, 200000, 150000, 1050000],
                "Production (%)": [1.0, 0.2, 0.15, 1.05],
                "todu_30ever_h6": [1000, 200, 300, 900],
                "todu_amt_pile_h6": [50000, 10000, 10000, 45000],
            }
        )

        df.to_csv(seg1_dir / "risk_production_summary_table_base.csv", index=False)
        df.to_csv(seg2_dir / "risk_production_summary_table_pessimistic.csv", index=False)
        df.to_csv(seg2_dir / "risk_production_summary_table_optimistic.csv", index=False)

        segments = {"seg1": {"name": "seg1"}, "seg2": {"name": "seg2"}}
        result = consolidate_segments(temp_output_dir, segments, {})

        scenarios_found = set(result["scenario"].unique().tolist())
        assert scenarios_found == {"pessimistic", "base", "optimistic"}


# =============================================================================
# Scenario Order Tests
# =============================================================================


class TestScenarioOrder:
    """Tests for scenario ordering (pessimistic -> base -> optimistic)."""

    def test_scenario_order_in_mapping(self):
        """Test that scenarios are mapped in correct order."""
        # The order should be: pessimistic (lowest), base (middle), optimistic (highest)
        suffixes = ["_1.1", "_0.9", "_1.0"]  # Out of order
        mapping = map_scenario_names(suffixes)

        assert mapping["_0.9"] == "pessimistic"
        assert mapping["_1.0"] == "base"
        assert mapping["_1.1"] == "optimistic"


class TestScenarioDeduplication:
    """Tests for scenario deduplication."""

    def test_no_duplicate_base_scenarios(self, temp_output_dir):
        """Test that '' and '_base' don't create duplicate entries."""
        # Create mock segment with both unsuffixed and _base files
        seg_dir = temp_output_dir / "test-segment" / "data"
        seg_dir.mkdir(parents=True)

        df = pd.DataFrame(
            {
                "Metric": ["Actual", "Swap-in", "Swap-out", "Optimum selected"],
                "Risk (%)": [1.5, 1.2, 2.0, 1.4],
                "Production (€)": [1000000, 200000, 150000, 1050000],
                "Production (%)": [1.0, 0.2, 0.15, 1.05],
                "todu_30ever_h6": [1000, 200, 300, 900],
                "todu_amt_pile_h6": [50000, 10000, 10000, 45000],
            }
        )

        # Create both unsuffixed (backward compat) and _base files
        df.to_csv(seg_dir / "risk_production_summary_table.csv", index=False)
        df.to_csv(seg_dir / "risk_production_summary_table_base.csv", index=False)
        df.to_csv(seg_dir / "risk_production_summary_table_pessimistic.csv", index=False)
        df.to_csv(seg_dir / "risk_production_summary_table_optimistic.csv", index=False)

        segments = {"test-segment": {"name": "test-segment"}}
        result = consolidate_segments(temp_output_dir, segments, {})

        # Count occurrences of 'base' scenario in main period
        main_base_count = ((result["scenario"] == "base") & (result["period"] == "main")).sum()

        # Should only have one 'base' entry per group/period combination, not duplicates
        # With one segment + TOTAL, we expect 2 rows for main period with base scenario
        assert main_base_count == 2  # segment + TOTAL

    def test_deduplication_prefers_named_suffix(self):
        """Test that deduplication prefers '_base' over '' suffix."""
        suffixes = ["", "_base", "_pessimistic", "_optimistic"]
        mapping = map_scenario_names(suffixes)

        # Both '' and '_base' map to 'base'
        assert mapping[""] == "base"
        assert mapping["_base"] == "base"

        # Deduplication logic
        seen_names = {}
        for suffix in suffixes:
            name = mapping.get(suffix, "base")
            if name not in seen_names or suffix:
                seen_names[name] = suffix

        # Should prefer '_base' over ''
        assert seen_names["base"] == "_base"
        assert len(seen_names) == 3  # base, pessimistic, optimistic


# =============================================================================
# Dashboard Tests
# =============================================================================


class TestCreateConsolidationDashboard:
    """Tests for create_consolidation_dashboard function."""

    def test_dashboard_has_executive_layout(self):
        """Verify the dashboard includes KPI, scenario, heatmap, and ranking views."""
        df = pd.DataFrame(
            [
                {
                    "group": "TOTAL",
                    "period": "main",
                    "scenario": "pessimistic",
                    "actual_production": 1000,
                    "optimum_production": 900,
                    "production_delta": -100,
                    "production_delta_pct": -10.0,
                    "actual_risk_pct": 1.6,
                    "optimum_risk_pct": 1.2,
                    "risk_delta_pct": -0.4,
                    "optimum_rejection_rate_pct": 18.0,
                    "production_ci_lower": 850,
                    "production_ci_upper": 950,
                },
                {
                    "group": "TOTAL",
                    "period": "main",
                    "scenario": "base",
                    "actual_production": 1000,
                    "optimum_production": 1050,
                    "production_delta": 50,
                    "production_delta_pct": 5.0,
                    "actual_risk_pct": 1.6,
                    "optimum_risk_pct": 1.3,
                    "risk_delta_pct": -0.3,
                    "optimum_rejection_rate_pct": 15.0,
                    "production_ci_lower": 1000,
                    "production_ci_upper": 1100,
                },
                {
                    "group": "TOTAL",
                    "period": "main",
                    "scenario": "optimistic",
                    "actual_production": 1000,
                    "optimum_production": 1120,
                    "production_delta": 120,
                    "production_delta_pct": 12.0,
                    "actual_risk_pct": 1.6,
                    "optimum_risk_pct": 1.45,
                    "risk_delta_pct": -0.15,
                    "optimum_rejection_rate_pct": 13.0,
                    "production_ci_lower": 1080,
                    "production_ci_upper": 1160,
                },
                {
                    "group": "TOTAL",
                    "period": "mr",
                    "scenario": "base",
                    "actual_production": 980,
                    "optimum_production": 1040,
                    "production_delta": 60,
                    "production_delta_pct": 6.1,
                    "actual_risk_pct": 1.5,
                    "optimum_risk_pct": 1.0,
                    "risk_delta_pct": -0.5,
                    "optimum_rejection_rate_pct": 14.0,
                    "production_ci_lower": 1000,
                    "production_ci_upper": 1080,
                },
                {
                    "group": "supersegment_ss1",
                    "period": "main",
                    "scenario": "base",
                    "actual_production": 1000,
                    "optimum_production": 1050,
                    "production_delta": 50,
                    "production_delta_pct": 5.0,
                    "actual_risk_pct": 1.6,
                    "optimum_risk_pct": 1.3,
                    "risk_delta_pct": -0.3,
                    "optimum_rejection_rate_pct": 15.0,
                },
                {
                    "group": "supersegment_ss1",
                    "period": "mr",
                    "scenario": "base",
                    "actual_production": 980,
                    "optimum_production": 1040,
                    "production_delta": 60,
                    "production_delta_pct": 6.1,
                    "actual_risk_pct": 1.5,
                    "optimum_risk_pct": 1.0,
                    "risk_delta_pct": -0.5,
                    "optimum_rejection_rate_pct": 14.0,
                },
                {
                    "group": "ss1/seg1",
                    "period": "main",
                    "scenario": "base",
                    "actual_production": 500,
                    "optimum_production": 560,
                    "production_delta": 60,
                    "production_delta_pct": 12.0,
                    "actual_risk_pct": 1.8,
                    "optimum_risk_pct": 1.4,
                    "risk_delta_pct": -0.4,
                    "optimum_rejection_rate_pct": 16.0,
                },
                {
                    "group": "ss1/seg1",
                    "period": "mr",
                    "scenario": "base",
                    "actual_production": 480,
                    "optimum_production": 530,
                    "production_delta": 50,
                    "production_delta_pct": 10.4,
                    "actual_risk_pct": 1.7,
                    "optimum_risk_pct": 1.1,
                    "risk_delta_pct": -0.6,
                    "optimum_rejection_rate_pct": 15.0,
                },
            ]
        )

        fig = create_consolidation_dashboard(df)

        trace_types = {trace.type for trace in fig.data}
        assert "indicator" in trace_types
        assert "heatmap" in trace_types
        assert "bar" in trace_types
        assert "scatter" in trace_types
        assert fig.layout.height == 1450
        assert "Executive portfolio view" in fig.layout.title.text


# =============================================================================
# Excel Export Tests
# =============================================================================


class TestExportConsolidatedExcel:
    """Tests for export_consolidated_excel function."""

    def _make_consolidated_df(self):
        """Create a minimal consolidated DataFrame with expected structure."""
        return pd.DataFrame(
            [
                {
                    "group": "TOTAL",
                    "period": "main",
                    "scenario": "base",
                    "n_segments": 1,
                    "segments": "seg1",
                    "actual_production": 1000000,
                    "actual_risk_pct": 1.5,
                    "optimum_production": 1050000,
                    "optimum_risk_pct": 1.4,
                    "production_delta": 50000,
                    "production_delta_pct": 5.0,
                },
                {
                    "group": "supersegment_ss1",
                    "period": "main",
                    "scenario": "base",
                    "n_segments": 1,
                    "segments": "seg1",
                    "actual_production": 1000000,
                    "actual_risk_pct": 1.5,
                    "optimum_production": 1050000,
                    "optimum_risk_pct": 1.4,
                    "production_delta": 50000,
                    "production_delta_pct": 5.0,
                },
                {
                    "group": "ss1/seg1",
                    "period": "main",
                    "scenario": "base",
                    "n_segments": 1,
                    "segments": "seg1",
                    "actual_production": 1000000,
                    "actual_risk_pct": 1.5,
                    "optimum_production": 1050000,
                    "optimum_risk_pct": 1.4,
                    "production_delta": 50000,
                    "production_delta_pct": 5.0,
                },
            ]
        )

    def test_creates_file_with_expected_sheets(self, temp_output_dir):
        """Verify .xlsx is created and has expected sheets including grid sheets."""
        df = self._make_consolidated_df()
        seg_dir = temp_output_dir / "seg1" / "data"
        seg_dir.mkdir(parents=True)
        # Cutoff data with accepted column for acceptance grid
        pd.DataFrame(
            {
                "octroi_bin": [1, 1, 2, 2],
                "efx_bin": [1, 2, 1, 2],
                "accepted": [1, 0, 1, 1],
                "segment": ["seg1"] * 4,
                "scenario": ["base"] * 4,
                "risk_pct": [1.0, 2.0, 1.5, 1.8],
                "production": [100, 200, 150, 180],
            }
        ).to_csv(seg_dir / "cutoff_summary_wide.csv", index=False)
        pd.DataFrame({"Metric": ["Actual"], "Risk (%)": [1.5], "Production (€)": [1000000]}).to_csv(
            seg_dir / "risk_production_summary_table_base.csv", index=False
        )

        xlsx_path = export_consolidated_excel(df, temp_output_dir, {"seg1": {"name": "seg1"}})

        assert xlsx_path.exists()
        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path)
        sheet_names = wb.sheetnames
        assert "Executive Summary" in sheet_names
        assert "RP seg1" in sheet_names
        # Removed as redundant: the scenario/total tables + acceptance grids live on the
        # Executive Summary, and per-segment detail on the RP sheets.
        for removed in ("Portfolio Summary", "Segment Detail", "Cutoff Comparison", "Grid seg1"):
            assert removed not in sheet_names
        wb.close()

    def test_executive_summary_has_kpi_cards(self, temp_output_dir):
        """Verify Executive Summary sheet has KPI values and table."""
        df = self._make_consolidated_df()
        # Add required fields for TOTAL row KPI extraction
        df.loc[df["group"] == "TOTAL", "risk_delta_pct"] = -0.1
        df.loc[df["group"] == "TOTAL", "optimum_rejection_rate_pct"] = 15.0
        xlsx_path = export_consolidated_excel(df, temp_output_dir, {})

        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Executive Summary"]
        # Title row
        assert "Dashboard" in str(ws["A1"].value)
        # KPI card row 4: optimum production value should be present
        kpi_values = [ws.cell(row=4, column=c).value for c in range(1, 9)]
        assert any(v is not None and "1,050,000" in str(v) for v in kpi_values)
        wb.close()

    def test_acceptance_grid_rendered(self, temp_output_dir):
        """Verify acceptance grid has A/R cells with correct coloring."""
        df = self._make_consolidated_df()
        seg_dir = temp_output_dir / "seg1" / "data"
        seg_dir.mkdir(parents=True)
        pd.DataFrame(
            {
                "octroi_bin": [1, 1, 2, 2],
                "efx_bin": [1, 2, 1, 2],
                "accepted": [1, 0, 1, 1],
                "segment": ["seg1"] * 4,
                "scenario": ["base"] * 4,
                "risk_pct": [1.0, 2.0, 1.5, 1.8],
                "production": [100, 200, 150, 180],
            }
        ).to_csv(seg_dir / "cutoff_summary_wide.csv", index=False)

        xlsx_path = export_consolidated_excel(df, temp_output_dir, {"seg1": {"name": "seg1"}})

        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Executive Summary"]  # acceptance grids are inlined here (Grid sheets removed)
        # Collect all cell values in the sheet
        all_values = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
            all_values.extend(v for v in row if v is not None)
        # Should contain A (accept) and R (reject) cells
        assert "A" in all_values
        assert "R" in all_values
        wb.close()

    def test_acceptance_grid_3d_slices_by_income_bin(self, temp_output_dir):
        """Verify 3-variable grids produce sub-grids sliced by income_bin."""
        df = self._make_consolidated_df()
        seg_dir = temp_output_dir / "seg1" / "data"
        seg_dir.mkdir(parents=True)
        # 3 variables: octroi_bin x efx_bin x income_bin
        rows = []
        for inc in [1, 2]:
            for o in [1, 2]:
                for e in [1, 2]:
                    rows.append(
                        {
                            "octroi_bin": o,
                            "efx_bin": e,
                            "income_bin": inc,
                            "accepted": 1 if (o + e + inc) <= 4 else 0,
                            "segment": "seg1",
                            "scenario": "base",
                            "risk_pct": 1.0,
                            "production": 100,
                        }
                    )
        pd.DataFrame(rows).to_csv(seg_dir / "cutoff_summary_wide.csv", index=False)

        xlsx_path = export_consolidated_excel(df, temp_output_dir, {"seg1": {"name": "seg1"}})

        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb["Executive Summary"]  # acceptance grids are inlined here (Grid sheets removed)
        # Collect all values — should have slice labels like "income_bin=1"
        all_values = [
            str(ws.cell(row=r, column=c).value or "")
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        ]
        assert any("income_bin=1" in v for v in all_values)
        assert any("income_bin=2" in v for v in all_values)
        # Should still have A and R cells
        assert "A" in all_values
        assert "R" in all_values
        wb.close()

    def test_empty_dataframe(self, temp_output_dir):
        """Verify graceful handling of empty DataFrame."""
        df = pd.DataFrame(columns=["group", "period", "scenario", "actual_production"])
        xlsx_path = export_consolidated_excel(df, temp_output_dir, {})

        assert xlsx_path.exists()
        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path)
        assert "Executive Summary" in wb.sheetnames
        wb.close()


# =============================================================================
# Audit patching (audit #25)
# =============================================================================


class TestPatchFromSegmentAudits:
    """Regression tests for patch_consolidated_production_from_segment_audits."""

    @staticmethod
    def _write_audit(output_base: Path, seg: str) -> None:
        """Audit CSV: actual=120 (booked), swap_out=20, swap_in=30 → optimum=130."""
        d = output_base / seg / "data"
        d.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(
            {
                "status_name": ["booked", "booked", "rejected"],
                "classification": ["keep", "swap_out", "swap_in"],
                "oa_amt": [100.0, 20.0, 30.0],
            }
        ).to_csv(d / "audit_base.csv", index=False)

    @staticmethod
    def _row(group: str, segments: str) -> dict:
        return {
            "group": group,
            "segments": segments,
            "period": "main",
            "scenario": "base",
            "actual_production": 1.0,  # stale RP-table number the audit must overwrite
            "optimum_production": 1.0,
            "swap_in_production": 0.0,
            "swap_out_production": 0.0,
            "production_delta": 0.0,
            "production_delta_pct": 0.0,
        }

    def test_single_member_supersegment_patch_lands_on_member_row(self, temp_output_dir):
        """Audit #25: a one-member supersegment row carries the member's name in
        `segments` and (listed first) used to swallow the audit patch, which the
        re-aggregation then overwrote from the unpatched member — correction lost."""
        self._write_audit(temp_output_dir, "new_ob")
        df = pd.DataFrame(
            [
                self._row("supersegment_pl_new", "new_ob"),  # aggregate FIRST: the failure ordering
                self._row("pl_new/new_ob", "new_ob"),  # member rows are named "<ss>/<seg>"
                self._row("TOTAL", "new_ob"),
            ]
        )

        out = patch_consolidated_production_from_segment_audits(df, temp_output_dir, {"new_ob": {}})

        member = out[out["group"] == "pl_new/new_ob"].iloc[0]
        ss = out[out["group"] == "supersegment_pl_new"].iloc[0]
        total = out[out["group"] == "TOTAL"].iloc[0]
        # member row patched from the audit
        assert member["actual_production"] == 120.0
        assert member["optimum_production"] == 130.0
        assert member["swap_in_production"] == 30.0
        assert member["swap_out_production"] == 20.0
        # supersegment + TOTAL re-aggregated from the PATCHED member
        assert ss["actual_production"] == 120.0
        assert ss["optimum_production"] == 130.0
        assert total["actual_production"] == 120.0
        assert total["optimum_production"] == 130.0

    def test_multi_member_supersegment_unaffected(self, temp_output_dir):
        """Multi-member supersegments (segments = comma list) never matched the
        member mask; verify they still re-aggregate from patched members."""
        self._write_audit(temp_output_dir, "seg_a")
        self._write_audit(temp_output_dir, "seg_b")
        df = pd.DataFrame(
            [
                self._row("supersegment_pl_known", "seg_a, seg_b"),
                self._row("segment_seg_a", "seg_a"),
                self._row("segment_seg_b", "seg_b"),
                self._row("TOTAL", "seg_a, seg_b"),
            ]
        )

        out = patch_consolidated_production_from_segment_audits(df, temp_output_dir, {"seg_a": {}, "seg_b": {}})

        ss = out[out["group"] == "supersegment_pl_known"].iloc[0]
        total = out[out["group"] == "TOTAL"].iloc[0]
        assert ss["actual_production"] == 240.0  # 120 + 120
        assert ss["optimum_production"] == 260.0
        assert total["actual_production"] == 240.0
        assert total["optimum_production"] == 260.0

    @classmethod
    def _rate_row(cls, group: str, segments: str, sys_rate: float) -> dict:
        r = cls._row(group, segments)
        r["total_demand"] = 1000.0
        r["actual_system_rejection_rate_pct"] = sys_rate
        return r

    def test_pooled_system_rate_none_poisoned_when_a_member_lacks_it(self, temp_output_dir):
        """#62: a supersegment/TOTAL system-rejection rate must be n/a (NaN) when
        any member lacks it — not a blended number from treating the missing
        member's € rejections as 0 while its demand still counts in the pool."""
        df = pd.DataFrame(
            [
                self._rate_row("supersegment_pl_known", "seg_a, seg_b", float("nan")),
                self._rate_row("segment_seg_a", "seg_a", 10.0),  # has a rate
                self._rate_row("segment_seg_b", "seg_b", float("nan")),  # no se_decision_id → NaN
                self._rate_row("TOTAL", "seg_a, seg_b", float("nan")),
            ]
        )

        out = patch_consolidated_production_from_segment_audits(df, temp_output_dir, {"seg_a": {}, "seg_b": {}})

        ss = out[out["group"] == "supersegment_pl_known"].iloc[0]
        total = out[out["group"] == "TOTAL"].iloc[0]
        assert pd.isna(ss["actual_system_rejection_rate_pct"])  # n/a, not a fabricated blend
        assert pd.isna(total["actual_system_rejection_rate_pct"])

    def test_pooled_system_rate_blended_when_all_members_have_it(self, temp_output_dir):
        """Control: with every member's rate present, the pooled rate is the
        demand-weighted blend (10%·1000 + 20%·1000) / 2000 = 15%."""
        df = pd.DataFrame(
            [
                self._rate_row("supersegment_pl_known", "seg_a, seg_b", float("nan")),
                self._rate_row("segment_seg_a", "seg_a", 10.0),
                self._rate_row("segment_seg_b", "seg_b", 20.0),
                self._rate_row("TOTAL", "seg_a, seg_b", float("nan")),
            ]
        )

        out = patch_consolidated_production_from_segment_audits(df, temp_output_dir, {"seg_a": {}, "seg_b": {}})

        ss = out[out["group"] == "supersegment_pl_known"].iloc[0]
        assert ss["actual_system_rejection_rate_pct"] == pytest.approx(15.0)


# =============================================================================
# Classification grid Total risk (audit #43)
# =============================================================================


def _grid_row(cat, ib, vol, risk, den=None, cnt=1):
    row = {
        "category": cat,
        "income_bin": ib,
        "income_label": f"ib{ib}",
        "volume": vol,
        "risk": risk,
        "count": cnt,
    }
    if den is not None:
        row["risk_den"] = den
    return row


class TestClassificationGridTotalRisk:
    """The Total risk column must be an exposure-pooled rate (Σ(risk·den)/Σden ==
    mult·Σnum/Σden), not a production-volume-weighted mean with None→0 bias."""

    @staticmethod
    def _keep_total_risk(grid):
        import openpyxl

        from src.consolidation import _write_classification_grid

        wb = openpyxl.Workbook()
        _write_classification_grid(wb.active, grid, start_row=1)
        # layout: data rows start at start_row+3 (Keep first); cols = 1 category +
        # 3 per income bin (Volume/Risk/Count) + Total group → with 2 bins the
        # Total risk cell is column 9.
        return wb.active.cell(row=4, column=9).value

    def test_no_risk_bins_excluded_from_total(self):
        """A bin without risk data used to contribute 0% to the numerator while its
        full volume stayed in the denominator — biasing the Total downward."""
        grid = [
            _grid_row("Keep", 1, vol=1000.0, risk=1.0, den=100.0),
            _grid_row("Keep", 2, vol=1000.0, risk=None, den=0.0),
        ]
        assert self._keep_total_risk(grid) == 1.0  # old behavior: 0.5

    def test_total_pooled_by_risk_denominator_not_volume(self):
        """risk_i = mult·num_i/den_i, so Σ(risk·den)/Σden is the exact pooled rate;
        oa_amt volume is the wrong weight."""
        grid = [
            _grid_row("Keep", 1, vol=10000.0, risk=1.0, den=100.0),
            _grid_row("Keep", 2, vol=10.0, risk=3.0, den=300.0),
        ]
        # (1·100 + 3·300) / 400 = 2.5; old volume weighting gave ≈ 1.002
        assert self._keep_total_risk(grid) == 2.5

    def test_legacy_rows_without_risk_den_fall_back_to_volume_over_risk_bins(self):
        grid = [
            _grid_row("Keep", 1, vol=100.0, risk=1.0),
            _grid_row("Keep", 2, vol=900.0, risk=None),
        ]
        assert self._keep_total_risk(grid) == 1.0  # old behavior: 0.1


# =============================================================================
# Run Tests
# =============================================================================

if __name__ == "__main__":
    pytest.main([__file__, "-v"])


class TestGetTotalRowNoCrossScenarioFallback:
    """#62: the exec-summary TOTAL lookup must not substitute a different
    scenario's row under the requested scenario's label."""

    @staticmethod
    def _df():
        return pd.DataFrame(
            [
                {"group": "TOTAL", "period": "main", "scenario": "pessimistic", "risk": 1.0},
                {"group": "TOTAL", "period": "main", "scenario": "optimistic", "risk": 2.0},
            ]
        )

    def test_missing_scenario_returns_none(self):
        # 'base' absent → None (NOT pessimistic/optimistic mislabeled as base)
        assert _get_total_row(self._df(), "main", "base") is None

    def test_present_scenario_returned(self):
        assert _get_total_row(self._df(), "main", "pessimistic")["risk"] == 1.0

    def test_missing_period_returns_none(self):
        assert _get_total_row(self._df(), "mr", "pessimistic") is None
