"""Tests for the consolidated-report trust layer (CIs, Validation & Governance, Out-of-time sheets)."""

import json
import os
import sys

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import pandas as pd
from openpyxl import Workbook

from src.consolidation import (
    _ci_eur,
    _ci_pct,
    _read_backtest_consolidated,
    _segment_status,
    _write_sheet_oot_validation,
    _write_sheet_validation_governance,
)

# ------------------------------- CI formatters -------------------------------


def test_ci_pct_and_eur():
    row = {"risk_ci_lower": 0.42, "risk_ci_upper": 1.24, "production_ci_lower": 39.4e6, "production_ci_upper": 42.0e6}
    assert _ci_pct(row, "risk_ci_lower", "risk_ci_upper") == "95% CI [0.42–1.24]"
    assert _ci_eur(row, "production_ci_lower", "production_ci_upper") == "95% CI [€39.4M–€42.0M]"


def test_ci_none_when_missing():
    assert _ci_pct({"risk_ci_lower": None, "risk_ci_upper": 1.0}, "risk_ci_lower", "risk_ci_upper") is None
    assert _ci_eur({}, "production_ci_lower", "production_ci_upper") is None


# ------------------------------- status mapping ------------------------------


def test_segment_status_traffic_light():
    assert _segment_status("OK", {})[0] == "OK"
    assert _segment_status("INCONCLUSIVE", {})[0] == "WATCH"
    assert _segment_status("DRIFT", {})[0] == "REVIEW"
    assert _segment_status("OK", {"critical": 1})[0] == "REVIEW"  # stability overrides up
    assert _segment_status("OK", {"warning": 2})[0] == "WATCH"
    assert _segment_status(None, {"worst_status": "moderate"})[0] == "WATCH"


# --------------------------- synthetic output tree ---------------------------


def _build_tree(tmp_path, seg="no_premium_cd"):
    data = tmp_path / seg / "data"
    data.mkdir(parents=True)
    (data / "run_lineage.json").write_text(
        json.dumps(
            {
                "run_id": "20260606T120000Z_abc123",
                "data": {
                    "path": "data/x.sas7bdat",
                    "sha256": "deadbeefcafe0000",
                    "mtime": "2026-03-11",
                    "rows_loaded": 1444802,
                },
                "git": {"short": "abc123", "dirty": False},
                "config": {"hash": "cfg123"},
                "assumptions": {"multiplier": 7.0, "stress_mode": "disabled", "use_mr_outcomes": True},
            }
        )
    )
    (data / "drift_alerts_base.json").write_text(json.dumps({"summary": {"info": 4, "warning": 0, "critical": 0}}))
    pd.DataFrame({"variable": ["oa_amt"], "psi": [0.08], "status": ["stable"], "n_bins": [9]}).to_csv(
        data / "stability_psi_base.csv", index=False
    )
    bt = tmp_path / "backtest"
    bt.mkdir()
    pd.DataFrame(
        [
            {
                "segment": seg,
                "drift_flag": "OK",
                "n_defaults_oot": 12,
                "predicted_risk_pct": 1.389,
                "in_sample_realized_risk_pct": 0.97,
                "in_sample_risk_ci": "[0.62, 1.37]",
                "oot_realized_risk_pct": 2.11,
                "oot_risk_ci": "[0.73, 3.93]",
                "in_sample_acceptance_rate": 0.46,
                "oot_acceptance_rate": 0.45,
                "holdout_start": "2025-05-01",
                "holdout_end": "2025-07-01",
                "pct_mature": 0.26,
            }
        ]
    ).to_csv(bt / "backtest_consolidated_base.csv", index=False)
    return seg


def test_read_backtest_consolidated(tmp_path):
    seg = _build_tree(tmp_path)
    df = _read_backtest_consolidated(tmp_path, "_base")
    assert df is not None and df.iloc[0]["segment"] == seg and df.iloc[0]["drift_flag"] == "OK"
    assert _read_backtest_consolidated(tmp_path / "nope", "_base") is None


def test_validation_governance_sheet(tmp_path):
    seg = _build_tree(tmp_path)
    wb = Workbook()
    _write_sheet_validation_governance(wb, tmp_path, {seg: {}}, pd.DataFrame(), "_base")
    assert "Validation & Governance" in wb.sheetnames
    text = "\n".join(
        str(c.value) for rowcells in wb["Validation & Governance"].iter_rows() for c in rowcells if c.value
    )
    assert "deadbeefcafe0000" in text  # data SHA pinned
    assert "multiplier" in text and "FIXED" in text  # assumptions + tier
    assert seg in text  # per-segment reproducibility/stability row


def test_oot_validation_sheet(tmp_path):
    seg = _build_tree(tmp_path)
    wb = Workbook()
    _write_sheet_oot_validation(wb, tmp_path, "_base")
    assert "Out-of-time Validation" in wb.sheetnames
    text = "\n".join(str(c.value) for rowcells in wb["Out-of-time Validation"].iter_rows() for c in rowcells if c.value)
    assert seg in text and "OK" in text and "noise-aware" in text


def test_oot_sheet_graceful_when_missing(tmp_path):
    (tmp_path / "empty").mkdir()
    wb = Workbook()
    _write_sheet_oot_validation(wb, tmp_path / "empty", "_base")
    text = "\n".join(str(c.value) for rowcells in wb["Out-of-time Validation"].iter_rows() for c in rowcells if c.value)
    assert "not available" in text and "run_backtest" in text
