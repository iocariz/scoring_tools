"""
Consolidation Module for Risk Production Tables

Aggregates risk_production_summary_table data across:
- Segments within a supersegment
- Total across all segments
- Main period and MR period
- Multiple scenarios

Produces portfolio-level views for executive reporting.
"""

from dataclasses import dataclass
from functools import cached_property
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import plotly.graph_objects as go
from loguru import logger
from plotly.subplots import make_subplots

from .audit import audit_production_kpis, reconcile_risk_production_summary_with_audit
from .constants import DEFAULT_RISK_MULTIPLIER, DEFAULT_RISK_MULTIPLIER_H3
from .utils import calculate_b2_ever_h6, resolve_reporting_supersegment

# ---------------------------------------------------------------------------
# Last-resort default variable list (todo #65)
# ---------------------------------------------------------------------------
# When BOTH the segment's config_segment.toml AND the global config.toml are
# unreadable (corrupted, missing, parse error), fall back to the canonical
# 3-variable setup this project was originally built around. This is a
# survivor-bias default: it exists so that a malformed config still produces
# *something* rather than crashing the consolidated report. New deployments
# that use different variable names MUST supply a valid config — this list
# is not the starting point for new projects.
#
# Kept as a module-level constant so it is grep-able and callers opt into
# the default explicitly rather than burying it in a function body.
_FALLBACK_REPORTING_VARIABLES: tuple[str, ...] = (
    "new_efx_clus",
    "sc_octroi_new_clus",
    "income_bin",
)


@dataclass
class ConsolidatedMetrics:
    """Consolidated metrics for a group of segments."""

    group_name: str
    period: str  # 'main' or 'mr'
    scenario: str
    segments: list[str]

    # Risk multipliers (from config; default to constants for backwards compatibility)
    multiplier: float = float(DEFAULT_RISK_MULTIPLIER)
    multiplier_h3: float = float(DEFAULT_RISK_MULTIPLIER_H3)

    # Aggregated metrics - Production (€)
    actual_production: float = 0.0
    optimum_production: float = 0.0
    swap_in_production: float = 0.0
    swap_out_production: float = 0.0

    # Raw risk components for proper aggregation
    # Risk = todu_30ever_h6 / todu_amt_pile_h6 * 7
    actual_todu_30ever_h6: float = 0.0
    actual_todu_amt_pile_h6: float = 0.0
    optimum_todu_30ever_h6: float = 0.0
    optimum_todu_amt_pile_h6: float = 0.0
    swap_in_todu_30ever_h6: float = 0.0
    swap_in_todu_amt_pile_h6: float = 0.0
    swap_out_todu_30ever_h6: float = 0.0
    swap_out_todu_amt_pile_h6: float = 0.0

    # Raw risk components for H3 (complementary 3-month horizon metric)
    actual_todu_30ever_h3: float = 0.0
    actual_todu_amt_pile_h3: float = 0.0
    optimum_todu_30ever_h3: float = 0.0
    optimum_todu_amt_pile_h3: float = 0.0
    swap_in_todu_30ever_h3: float = 0.0
    swap_in_todu_amt_pile_h3: float = 0.0
    swap_out_todu_30ever_h3: float = 0.0
    swap_out_todu_amt_pile_h3: float = 0.0

    # Total demand (booked + rejected, excluding canceled)
    total_demand: float = 0.0

    # Confidence Intervals (for optimum solution)
    optimum_production_ci_lower: float = 0.0
    optimum_production_ci_upper: float = 0.0
    optimum_risk_ci_lower: float = 0.0
    optimum_risk_ci_upper: float = 0.0

    # Calculated risk properties (as percentage, e.g. 1.5 means 1.5%)
    @cached_property
    def actual_risk(self) -> float:
        return float(
            np.nan_to_num(
                calculate_b2_ever_h6(
                    self.actual_todu_30ever_h6,
                    self.actual_todu_amt_pile_h6,
                    multiplier=self.multiplier,
                    as_percentage=True,
                    decimals=6,
                )
            )
        )

    @cached_property
    def optimum_risk(self) -> float:
        return float(
            np.nan_to_num(
                calculate_b2_ever_h6(
                    self.optimum_todu_30ever_h6,
                    self.optimum_todu_amt_pile_h6,
                    multiplier=self.multiplier,
                    as_percentage=True,
                    decimals=6,
                )
            )
        )

    @cached_property
    def swap_in_risk(self) -> float:
        return float(
            np.nan_to_num(
                calculate_b2_ever_h6(
                    self.swap_in_todu_30ever_h6,
                    self.swap_in_todu_amt_pile_h6,
                    multiplier=self.multiplier,
                    as_percentage=True,
                    decimals=6,
                )
            )
        )

    @cached_property
    def swap_out_risk(self) -> float:
        return float(
            np.nan_to_num(
                calculate_b2_ever_h6(
                    self.swap_out_todu_30ever_h6,
                    self.swap_out_todu_amt_pile_h6,
                    multiplier=self.multiplier,
                    as_percentage=True,
                    decimals=6,
                )
            )
        )

    # H3 risk properties (complementary 3-month horizon metric)
    @cached_property
    def actual_risk_h3(self) -> float:
        return float(
            np.nan_to_num(
                calculate_b2_ever_h6(
                    self.actual_todu_30ever_h3,
                    self.actual_todu_amt_pile_h3,
                    multiplier=self.multiplier_h3,
                    as_percentage=True,
                    decimals=6,
                )
            )
        )

    @cached_property
    def optimum_risk_h3(self) -> float:
        return float(
            np.nan_to_num(
                calculate_b2_ever_h6(
                    self.optimum_todu_30ever_h3,
                    self.optimum_todu_amt_pile_h3,
                    multiplier=self.multiplier_h3,
                    as_percentage=True,
                    decimals=6,
                )
            )
        )

    @cached_property
    def swap_in_risk_h3(self) -> float:
        return float(
            np.nan_to_num(
                calculate_b2_ever_h6(
                    self.swap_in_todu_30ever_h3,
                    self.swap_in_todu_amt_pile_h3,
                    multiplier=self.multiplier_h3,
                    as_percentage=True,
                    decimals=6,
                )
            )
        )

    @cached_property
    def swap_out_risk_h3(self) -> float:
        return float(
            np.nan_to_num(
                calculate_b2_ever_h6(
                    self.swap_out_todu_30ever_h3,
                    self.swap_out_todu_amt_pile_h3,
                    multiplier=self.multiplier_h3,
                    as_percentage=True,
                    decimals=6,
                )
            )
        )

    @property
    def actual_rejection_rate(self) -> float:
        """Rejection rate under the actual (current) policy: rejected / total_demand."""
        if self.total_demand <= 0:
            return 0.0
        return (1 - self.actual_production / self.total_demand) * 100

    @property
    def optimum_rejection_rate(self) -> float:
        """Rejection rate under the optimum policy: rejected / total_demand."""
        if self.total_demand <= 0:
            return 0.0
        return (1 - self.optimum_production / self.total_demand) * 100

    @property
    def production_delta(self) -> float:
        return self.optimum_production - self.actual_production

    @property
    def production_delta_pct(self) -> float:
        if self.actual_production == 0:
            return 0.0
        return (self.optimum_production - self.actual_production) / self.actual_production

    @property
    def risk_delta(self) -> float:
        return self.optimum_risk - self.actual_risk

    def to_dict(self) -> dict[str, Any]:
        # Risk properties already return percentage (e.g. 7.0 means 7%)
        d = {
            "group": self.group_name,
            "period": self.period,
            "scenario": self.scenario,
            "n_segments": len(self.segments),
            "segments": ", ".join(self.segments),
            "actual_production": self.actual_production,
            "actual_risk_pct": self.actual_risk,
            "actual_todu_30ever_h6": self.actual_todu_30ever_h6,
            "actual_todu_amt_pile_h6": self.actual_todu_amt_pile_h6,
            "optimum_production": self.optimum_production,
            "optimum_risk_pct": self.optimum_risk,
            "optimum_todu_30ever_h6": self.optimum_todu_30ever_h6,
            "optimum_todu_amt_pile_h6": self.optimum_todu_amt_pile_h6,
            "swap_in_production": self.swap_in_production,
            "swap_in_risk_pct": self.swap_in_risk,
            "swap_in_todu_30ever_h6": self.swap_in_todu_30ever_h6,
            "swap_in_todu_amt_pile_h6": self.swap_in_todu_amt_pile_h6,
            "swap_out_production": self.swap_out_production,
            "swap_out_risk_pct": self.swap_out_risk,
            "swap_out_todu_30ever_h6": self.swap_out_todu_30ever_h6,
            "swap_out_todu_amt_pile_h6": self.swap_out_todu_amt_pile_h6,
            "production_delta": self.production_delta,
            "production_delta_pct": self.production_delta_pct * 100,
            "risk_delta_pct": self.risk_delta,
            "production_ci_lower": self.optimum_production_ci_lower,
            "production_ci_upper": self.optimum_production_ci_upper,
            "risk_ci_lower": self.optimum_risk_ci_lower,
            "risk_ci_upper": self.optimum_risk_ci_upper,
            "total_demand": self.total_demand,
            "actual_rejection_rate_pct": self.actual_rejection_rate,
            "optimum_rejection_rate_pct": self.optimum_rejection_rate,
        }
        # H3 complementary risk metrics
        has_h3 = (
            self.actual_todu_amt_pile_h3 > 0
            or self.optimum_todu_amt_pile_h3 > 0
            or self.swap_in_todu_amt_pile_h3 > 0
            or self.swap_out_todu_amt_pile_h3 > 0
        )
        if has_h3:
            d.update(
                {
                    "actual_risk_h3_pct": self.actual_risk_h3,
                    "actual_todu_30ever_h3": self.actual_todu_30ever_h3,
                    "actual_todu_amt_pile_h3": self.actual_todu_amt_pile_h3,
                    "optimum_risk_h3_pct": self.optimum_risk_h3,
                    "optimum_todu_30ever_h3": self.optimum_todu_30ever_h3,
                    "optimum_todu_amt_pile_h3": self.optimum_todu_amt_pile_h3,
                    "swap_in_risk_h3_pct": self.swap_in_risk_h3,
                    "swap_in_todu_30ever_h3": self.swap_in_todu_30ever_h3,
                    "swap_in_todu_amt_pile_h3": self.swap_in_todu_amt_pile_h3,
                    "swap_out_risk_h3_pct": self.swap_out_risk_h3,
                    "swap_out_todu_30ever_h3": self.swap_out_todu_30ever_h3,
                    "swap_out_todu_amt_pile_h3": self.swap_out_todu_amt_pile_h3,
                }
            )
        return d


def find_scenario_suffix(filename: str) -> str:
    """Extract scenario suffix from filename.

    Handles both named scenarios (e.g., 'risk_production_summary_table_pessimistic.csv')
    and legacy numeric scenarios (e.g., 'risk_production_summary_table_1.1.csv').
    """
    # Remove extension
    name = Path(filename).stem

    # Named scenarios to detect
    named_scenarios = ["pessimistic", "base", "optimistic"]

    # Check for named scenario pattern first
    parts = name.split("_")
    for part in reversed(parts):
        if part.lower() in named_scenarios:
            return f"_{part.lower()}"

    # Check for legacy numeric pattern (e.g., _1.1, _0.9)
    for part in reversed(parts):
        try:
            float(part)
            return f"_{part}"
        except ValueError:
            continue

    return ""


def map_scenario_names(scenario_suffixes: list[str]) -> dict[str, str]:
    """
    Map scenario suffixes to meaningful names.

    Scenarios:
    - base: optimum risk threshold (middle value or no suffix)
    - pessimistic: optimum - step (lower value, more conservative)
    - optimistic: optimum + step (higher value, more aggressive)

    Args:
        scenario_suffixes: List of suffixes like ['_pessimistic', '_base', '_optimistic']
                          or legacy format ['', '_0.9', '_1.0', '_1.1']

    Returns:
        Dict mapping suffix to name, e.g., {'_base': 'base', '_pessimistic': 'pessimistic'}
    """
    mapping = {}

    # Check if we have named scenarios (new format)
    named_scenarios = {"pessimistic", "base", "optimistic"}

    for suffix in scenario_suffixes:
        clean_suffix = suffix.strip("_").lower()

        if clean_suffix in named_scenarios:
            # New format: _pessimistic, _base, _optimistic
            mapping[suffix] = clean_suffix
        elif suffix == "":
            # Empty suffix defaults to base
            mapping[suffix] = "base"
        else:
            # Legacy format: try to parse as numeric
            try:
                val = float(clean_suffix)
                # Will be mapped later based on relative values
                mapping[suffix] = None  # Placeholder
            except ValueError:
                # Unknown format, use as-is
                mapping[suffix] = clean_suffix

    # Handle legacy numeric format if any placeholders exist
    placeholders = [s for s, v in mapping.items() if v is None]
    if placeholders:
        # Extract numeric values and sort
        numeric_suffixes = []
        for suffix in placeholders:
            try:
                val = float(suffix.strip("_"))
                numeric_suffixes.append((suffix, val))
            except ValueError:
                mapping[suffix] = suffix.strip("_")

        if numeric_suffixes:
            numeric_suffixes.sort(key=lambda x: x[1])

            if len(numeric_suffixes) == 1:
                mapping[numeric_suffixes[0][0]] = "base"
            elif len(numeric_suffixes) == 2:
                mapping[numeric_suffixes[0][0]] = "pessimistic"
                mapping[numeric_suffixes[1][0]] = "optimistic"
            else:
                # Three or more: lowest is pessimistic, highest is optimistic, middle is base
                mapping[numeric_suffixes[0][0]] = "pessimistic"
                mapping[numeric_suffixes[-1][0]] = "optimistic"
                for suffix, _val in numeric_suffixes[1:-1]:
                    mapping[suffix] = "base"

    return mapping


def load_risk_production_table(
    segment_dir: Path, period: str = "main", scenario_suffix: str = ""
) -> pd.DataFrame | None:
    """
    Load risk_production_summary_table for a segment.

    Args:
        segment_dir: Path to segment output directory
        period: 'main' or 'mr'
        scenario_suffix: Scenario suffix like '_1.1' or ''

    Returns:
        DataFrame or None if file not found
    """
    if period == "mr":
        filename = f"risk_production_summary_table_mr{scenario_suffix}.csv"
    else:
        filename = f"risk_production_summary_table{scenario_suffix}.csv"

    filepath = segment_dir / "data" / filename

    if not filepath.exists():
        logger.debug(f"File not found: {filepath}")
        return None

    try:
        df = pd.read_csv(filepath)
        return df
    except (pd.errors.ParserError, OSError, ValueError) as e:
        logger.warning(f"Error loading {filepath}: {e}")
        return None


def extract_metrics_from_table(df: pd.DataFrame) -> dict[str, dict[str, float]]:
    """
    Extract key metrics from risk_production_summary_table.

    Expected columns:
        - Metric: 'Actual', 'Swap-in', 'Swap-out', 'Optimum selected'
        - Production (€): oa_amt_h0
        - Risk (%): b2_ever_h6 (calculated)
        - todu_30ever_h6: raw numerator for risk
        - todu_amt_pile_h6: raw denominator for risk

    Returns dict with production and raw todu values for proper aggregation.
    """
    metrics = {
        "actual": {
            "production": 0,
            "todu_30ever_h6": 0,
            "todu_amt_pile_h6": 0,
            "todu_30ever_h3": 0,
            "todu_amt_pile_h3": 0,
        },
        "optimum": {
            "production": 0,
            "todu_30ever_h6": 0,
            "todu_amt_pile_h6": 0,
            "todu_30ever_h3": 0,
            "todu_amt_pile_h3": 0,
        },
        "swap_in": {
            "production": 0,
            "todu_30ever_h6": 0,
            "todu_amt_pile_h6": 0,
            "todu_30ever_h3": 0,
            "todu_amt_pile_h3": 0,
        },
        "swap_out": {
            "production": 0,
            "todu_30ever_h6": 0,
            "todu_amt_pile_h6": 0,
            "todu_30ever_h3": 0,
            "todu_amt_pile_h3": 0,
        },
    }

    if df is None or df.empty:
        return metrics

    # Normalize column names
    df = df.copy()
    df.columns = df.columns.str.lower().str.strip()

    # Find metric column (could be 'metric', 'category', first column, etc.)
    metric_col = None
    for col in ["metric", "category", "type", "row"]:
        if col in df.columns:
            metric_col = col
            break
    if metric_col is None:
        metric_col = df.columns[0]

    # Find value columns
    prod_col = None
    todu_30_col = None
    todu_amt_col = None
    todu_30_h3_col = None
    todu_amt_h3_col = None
    rejection_rate_col = None

    for col in df.columns:
        col_lower = col.lower()
        if "production" in col_lower and "€" in col_lower:
            prod_col = col
        elif col_lower == "todu_30ever_h6":
            todu_30_col = col
        elif col_lower == "todu_amt_pile_h6":
            todu_amt_col = col
        elif col_lower == "todu_30ever_h3":
            todu_30_h3_col = col
        elif col_lower == "todu_amt_pile_h3":
            todu_amt_h3_col = col
        elif "rejection" in col_lower and "rate" in col_lower:
            rejection_rate_col = col

    # Fallback for production column
    if prod_col is None:
        for col in df.columns:
            if "prod" in col.lower() or "oa_amt" in col.lower():
                prod_col = col
                break

    logger.debug(f"Columns found - prod: {prod_col}, todu_30: {todu_30_col}, todu_amt: {todu_amt_col}")

    # Extract values for each row type
    for _, row in df.iterrows():
        row_type = str(row[metric_col]).lower().strip()

        # Map row type to our standard names
        if "actual" in row_type or "current" in row_type or "baseline" in row_type:
            key = "actual"
        elif "optim" in row_type or "target" in row_type:
            key = "optimum"
        elif "swap-in" in row_type or "swapin" in row_type or "swap_in" in row_type:
            key = "swap_in"
        elif "swap-out" in row_type or "swapout" in row_type or "swap_out" in row_type:
            key = "swap_out"
        else:
            continue

        # Extract production
        if prod_col and prod_col in row.index:
            try:
                metrics[key]["production"] = float(row[prod_col]) if pd.notna(row[prod_col]) else 0
            except (ValueError, TypeError):
                pass

        # Extract todu_30ever_h6
        if todu_30_col and todu_30_col in row.index:
            try:
                metrics[key]["todu_30ever_h6"] = float(row[todu_30_col]) if pd.notna(row[todu_30_col]) else 0
            except (ValueError, TypeError):
                pass

        # Extract todu_amt_pile_h6
        if todu_amt_col and todu_amt_col in row.index:
            try:
                metrics[key]["todu_amt_pile_h6"] = float(row[todu_amt_col]) if pd.notna(row[todu_amt_col]) else 0
            except (ValueError, TypeError):
                pass

        # Extract H3 columns (complementary metric)
        if todu_30_h3_col and todu_30_h3_col in row.index:
            try:
                metrics[key]["todu_30ever_h3"] = float(row[todu_30_h3_col]) if pd.notna(row[todu_30_h3_col]) else 0
            except (ValueError, TypeError):
                pass
        if todu_amt_h3_col and todu_amt_h3_col in row.index:
            try:
                metrics[key]["todu_amt_pile_h3"] = float(row[todu_amt_h3_col]) if pd.notna(row[todu_amt_h3_col]) else 0
            except (ValueError, TypeError):
                pass

        # Extract CI values (only for optimum)
        if key == "optimum":
            for col in df.columns:
                col_lower = col.lower()
                if "production_ci_lower" in col_lower:
                    metrics["optimum"]["production_ci_lower"] = float(row[col]) if pd.notna(row[col]) else 0
                elif "production_ci_upper" in col_lower:
                    metrics["optimum"]["production_ci_upper"] = float(row[col]) if pd.notna(row[col]) else 0
                elif "risk_ci_lower" in col_lower:
                    metrics["optimum"]["risk_ci_lower"] = float(row[col]) if pd.notna(row[col]) else 0
                elif "risk_ci_upper" in col_lower:
                    metrics["optimum"]["risk_ci_upper"] = float(row[col]) if pd.notna(row[col]) else 0

    # Read total_demand from explicit column, else derive from rejection rate
    actual_prod = metrics["actual"]["production"]
    total_demand_col = None
    for col in df.columns:
        if "total demand" in col.lower():
            total_demand_col = col
            break

    actual_row = df[df[metric_col].str.lower().str.strip().str.contains("actual|current|baseline", na=False)]
    if total_demand_col is not None and not actual_row.empty:
        td_val = actual_row.iloc[0].get(total_demand_col)
        if pd.notna(td_val) and float(td_val) > 0:
            metrics["_total_demand"] = float(td_val)
        else:
            metrics["_total_demand"] = actual_prod
    else:
        # Backward compat: derive from rejection rate
        actual_rej = 0.0
        if rejection_rate_col is not None and not actual_row.empty:
            rej_val = actual_row.iloc[0].get(rejection_rate_col)
            if pd.notna(rej_val):
                actual_rej = float(rej_val)
        if actual_rej < 100 and actual_prod > 0:
            metrics["_total_demand"] = actual_prod / (1 - actual_rej / 100)
        else:
            metrics["_total_demand"] = actual_prod

    return metrics


def _segment_audit_csv_path(output_base: Path, seg_name: str, period: str, scenario: str) -> Path:
    """Resolved audit file for a segment run (matches ``save_audit_tables`` naming)."""
    data_dir = output_base / seg_name / "data"
    if period == "mr":
        return data_dir / f"audit_{scenario}_mr.csv"
    return data_dir / f"audit_{scenario}.csv"


def patch_consolidated_production_from_segment_audits(
    df: pd.DataFrame,
    output_base: Path,
    segments: dict[str, dict[str, Any]],
) -> pd.DataFrame:
    """
    Overwrite production € (and derived deltas / rejection rates / total demand) from loan-level audit CSVs,
    then re-aggregate supersegment and TOTAL rows so consolidated outputs match ``audit_*.csv`` drill-down.
    """
    if df.empty or not segments:
        return df

    work = df.copy()

    def _apply_kpis(idx: int, kpis: dict[str, float]) -> None:
        a = kpis["actual"]
        o = kpis["optimum"]
        si = kpis["swap_in"]
        so = kpis["swap_out"]
        td = kpis["total_demand"]
        work.loc[idx, "actual_production"] = a
        work.loc[idx, "optimum_production"] = o
        work.loc[idx, "swap_in_production"] = si
        work.loc[idx, "swap_out_production"] = so
        work.loc[idx, "production_delta"] = o - a
        work.loc[idx, "production_delta_pct"] = (100.0 * (o - a) / a) if a else 0.0
        if "total_demand" in work.columns:
            work.loc[idx, "total_demand"] = td
        if td > 0:
            if "actual_rejection_rate_pct" in work.columns:
                work.loc[idx, "actual_rejection_rate_pct"] = (1.0 - a / td) * 100.0
            if "optimum_rejection_rate_pct" in work.columns:
                work.loc[idx, "optimum_rejection_rate_pct"] = (1.0 - o / td) * 100.0

    # Detect which segments ran in baseline mode (accept-all mask makes audit
    # swap-in classifications meaningless — skip audit patching for those).
    def _is_baseline_segment(seg_name_local: str) -> bool:
        try:
            import tomllib

            cfg_path = output_base / seg_name_local / "config_segment.toml"
            if cfg_path.exists():
                cfg = tomllib.loads(cfg_path.read_text(encoding="utf-8"))
                return bool(cfg.get("preprocessing", cfg).get("baseline_mode", False))
        except Exception:
            logger.warning(
                f"_is_baseline_segment: failed to read config for '{seg_name_local}'; treating as non-baseline",
                exc_info=True,
            )
        return False

    for seg_name in segments:
        if _is_baseline_segment(seg_name):
            continue
        for scenario in work["scenario"].dropna().unique():
            sc = str(scenario)
            for period in ("main", "mr"):
                path = _segment_audit_csv_path(output_base, seg_name, period, sc)
                if not path.exists():
                    continue
                try:
                    audit = pd.read_csv(path, skipinitialspace=True)
                    audit.columns = audit.columns.str.strip()
                except (pd.errors.ParserError, OSError, ValueError):
                    continue
                kpis = audit_production_kpis(audit)
                mask = (
                    (work["segments"].astype(str).str.strip() == seg_name)
                    & (work["period"] == period)
                    & (work["scenario"] == sc)
                )
                hit = work.loc[mask]
                if hit.empty:
                    continue
                _apply_kpis(int(hit.index[0]), kpis)

    def _detail_segment_mask(s: pd.Series) -> pd.Series:
        g = s.astype(str)
        return (g != "TOTAL") & ~g.str.startswith("supersegment_")

    # Supersegment rows = sum of member segment rows (disjoint segments)
    ss_idx = work["group"].astype(str).str.startswith("supersegment_")
    for idx, row in work.loc[ss_idx].iterrows():
        members = [x.strip() for x in str(row["segments"]).split(",") if x.strip()]
        if not members:
            continue
        period, scenario = row["period"], str(row["scenario"])
        sub = work.loc[
            work["segments"].astype(str).str.strip().isin(members)
            & (work["period"] == period)
            & (work["scenario"] == scenario)
            & _detail_segment_mask(work["group"]),
        ]
        if sub.empty:
            continue
        td = float(sub["total_demand"].sum()) if "total_demand" in sub.columns else 0.0
        kpis = {
            "actual": float(sub["actual_production"].sum()),
            "optimum": float(sub["optimum_production"].sum()),
            "swap_in": float(sub["swap_in_production"].sum()),
            "swap_out": float(sub["swap_out_production"].sum()),
            "total_demand": td,
        }
        _apply_kpis(int(idx), kpis)

    # TOTAL = sum of individual segment rows (same basis as consolidate_segments aggregation)
    tot_idx = work["group"].astype(str) == "TOTAL"
    for idx, row in work.loc[tot_idx].iterrows():
        period, scenario = row["period"], str(row["scenario"])
        sub = work.loc[
            _detail_segment_mask(work["group"]) & (work["period"] == period) & (work["scenario"] == scenario)
        ]
        if sub.empty:
            continue
        td = float(sub["total_demand"].sum()) if "total_demand" in sub.columns else 0.0
        kpis = {
            "actual": float(sub["actual_production"].sum()),
            "optimum": float(sub["optimum_production"].sum()),
            "swap_in": float(sub["swap_in_production"].sum()),
            "swap_out": float(sub["swap_out_production"].sum()),
            "total_demand": td,
        }
        _apply_kpis(int(idx), kpis)

    logger.info("Consolidated production metrics patched from segment audit CSVs (where present).")
    return work


def aggregate_metrics(
    metrics_list: list[dict[str, dict[str, float]]],
    multiplier: float = float(DEFAULT_RISK_MULTIPLIER),
) -> dict[str, dict[str, float]]:
    """
    Aggregate metrics from multiple segments.

    Sums production and raw todu values. Risk is calculated from aggregated
    todu values: risk = sum(todu_30ever_h6) / sum(todu_amt_pile_h6) * multiplier
    """
    aggregated = {
        "actual": {
            "production": 0,
            "todu_30ever_h6": 0,
            "todu_amt_pile_h6": 0,
            "todu_30ever_h3": 0,
            "todu_amt_pile_h3": 0,
        },
        "optimum": {
            "production": 0,
            "todu_30ever_h6": 0,
            "todu_amt_pile_h6": 0,
            "todu_30ever_h3": 0,
            "todu_amt_pile_h3": 0,
        },
        "swap_in": {
            "production": 0,
            "todu_30ever_h6": 0,
            "todu_amt_pile_h6": 0,
            "todu_30ever_h3": 0,
            "todu_amt_pile_h3": 0,
        },
        "swap_out": {
            "production": 0,
            "todu_30ever_h6": 0,
            "todu_amt_pile_h6": 0,
            "todu_30ever_h3": 0,
            "todu_amt_pile_h3": 0,
        },
    }

    total_demand_sum = 0.0
    for metrics in metrics_list:
        total_demand_sum += metrics.get("_total_demand", metrics["actual"]["production"])
        for key in aggregated:
            aggregated[key]["production"] += metrics[key]["production"]
            aggregated[key]["todu_30ever_h6"] += metrics[key]["todu_30ever_h6"]
            aggregated[key]["todu_amt_pile_h6"] += metrics[key]["todu_amt_pile_h6"]
            aggregated[key]["todu_30ever_h3"] += metrics[key].get("todu_30ever_h3", 0)
            aggregated[key]["todu_amt_pile_h3"] += metrics[key].get("todu_amt_pile_h3", 0)

            # Aggregate CIs for optimum using variance addition (assumes independence)
            if key == "optimum":
                if "_ci_segments" not in aggregated[key]:
                    aggregated[key]["_ci_segments"] = []

                prod_point = float(metrics[key].get("production", 0) or 0)
                risk_den = float(metrics[key].get("todu_amt_pile_h6", 0) or 0)
                risk_num = float(metrics[key].get("todu_30ever_h6", 0) or 0)
                risk_point_raw = calculate_b2_ever_h6(
                    risk_num,
                    risk_den,
                    multiplier=multiplier,
                    as_percentage=True,
                    decimals=6,
                )
                risk_point = float(risk_point_raw) if pd.notna(risk_point_raw) else 0.0
                prod_lower = metrics[key].get("production_ci_lower")
                prod_upper = metrics[key].get("production_ci_upper")
                risk_lower = metrics[key].get("risk_ci_lower")
                risk_upper = metrics[key].get("risk_ci_upper")

                aggregated[key]["_ci_segments"].append(
                    {
                        "prod_lower": float(prod_lower) if pd.notna(prod_lower) else prod_point,
                        "prod_upper": float(prod_upper) if pd.notna(prod_upper) else prod_point,
                        "risk_lower": float(risk_lower) if pd.notna(risk_lower) else risk_point,
                        "risk_upper": float(risk_upper) if pd.notna(risk_upper) else risk_point,
                        "risk_den": risk_den,
                    }
                )

    # Combine segment CIs using variance addition rule (assumes independence).
    # SE_i ≈ (upper - lower) / (2 * z_95); combined SE = sqrt(sum(SE_i²))
    z_95 = 1.96
    for key in aggregated:
        segments = aggregated[key].pop("_ci_segments", [])
        if not segments:
            aggregated[key].setdefault("production_ci_lower", 0)
            aggregated[key].setdefault("production_ci_upper", 0)
            aggregated[key].setdefault("risk_ci_lower", 0)
            aggregated[key].setdefault("risk_ci_upper", 0)
            continue

        # Combine production CIs using variance addition (assumes independence)
        # for consistency with risk CI aggregation.
        agg_prod_point = aggregated[key].get("production", 0)
        prod_var = 0.0
        for s in segments:
            prod_width = max(s["prod_upper"] - s["prod_lower"], 0.0)
            prod_se = prod_width / (2 * z_95) if prod_width else 0.0
            prod_var += prod_se**2
        combined_prod_se = float(np.sqrt(prod_var))
        aggregated[key]["production_ci_lower"] = max(agg_prod_point - z_95 * combined_prod_se, 0.0)
        aggregated[key]["production_ci_upper"] = agg_prod_point + z_95 * combined_prod_se

        agg_num = aggregated[key].get("todu_30ever_h6", 0)
        agg_den = aggregated[key].get("todu_amt_pile_h6", 0)
        agg_risk_point_raw = calculate_b2_ever_h6(
            agg_num,
            agg_den,
            multiplier=multiplier,
            as_percentage=True,
            decimals=6,
        )
        agg_risk_point = float(agg_risk_point_raw) if pd.notna(agg_risk_point_raw) else 0.0
        if agg_den and not (isinstance(agg_den, float) and np.isnan(agg_den)):
            combined_risk_var = 0.0
            for segment_ci in segments:
                risk_width = max(segment_ci["risk_upper"] - segment_ci["risk_lower"], 0.0)
                risk_se = risk_width / (2 * z_95) if risk_width else 0.0
                risk_weight = segment_ci["risk_den"] / agg_den if agg_den else 0.0
                combined_risk_var += (risk_weight * risk_se) ** 2
            combined_risk_se = float(np.sqrt(combined_risk_var))
            aggregated[key]["risk_ci_lower"] = max(agg_risk_point - z_95 * combined_risk_se, 0.0)
            aggregated[key]["risk_ci_upper"] = agg_risk_point + z_95 * combined_risk_se
        else:
            aggregated[key]["risk_ci_lower"] = 0.0
            aggregated[key]["risk_ci_upper"] = 0.0

    aggregated["_total_demand"] = total_demand_sum

    return aggregated


def consolidate_segments(
    output_base: Path,
    segments: dict[str, dict[str, Any]],
    supersegments: dict[str, dict[str, Any]],
    scenarios: list[str] | None = None,
    multiplier: float = float(DEFAULT_RISK_MULTIPLIER),
    multiplier_h3: float = float(DEFAULT_RISK_MULTIPLIER_H3),
) -> pd.DataFrame:
    """
    Consolidate risk production tables across segments, supersegments, and scenarios.

    Args:
        output_base: Base output directory
        segments: Segment configurations
        supersegments: Supersegment configurations
        scenarios: List of scenario suffixes (e.g., ['', '_1.0', '_1.1'])
        multiplier: H6 risk multiplier (from config.toml)
        multiplier_h3: H3 risk multiplier (from config.toml)

    Returns:
        DataFrame with consolidated metrics
    """
    if scenarios is None:
        scenarios = []
        seen_scenarios = set()
        for seg_name in segments:
            seg_dir = output_base / seg_name / "data"
            if seg_dir.exists():
                for f in sorted(seg_dir.glob("risk_production_summary_table*.csv")):
                    # Skip MR files for scenario detection
                    if "_mr" in f.name:
                        continue
                    suffix = find_scenario_suffix(f.name)
                    if suffix not in seen_scenarios:
                        scenarios.append(suffix)
                        seen_scenarios.add(suffix)

        # Ensure we have at least one scenario
        if not scenarios:
            scenarios = [""]

    # Map scenario suffixes to meaningful names (base, pessimistic, optimistic)
    scenario_name_map = map_scenario_names(scenarios)

    # Deduplicate scenarios that map to the same name (e.g., '' and '_base' both map to 'base')
    # Keep the more specific suffix (e.g., '_base' over '')
    seen_names = {}
    for suffix in scenarios:
        name = scenario_name_map.get(suffix, "base")
        if name not in seen_names or (suffix and not seen_names[name]):
            seen_names[name] = suffix

    # Rebuild scenarios list with deduplicated suffixes
    scenarios = list(seen_names.values())
    scenario_name_map = {suffix: name for name, suffix in seen_names.items()}

    logger.info(f"Consolidating data for scenarios: {scenario_name_map}")

    results = []

    # Build supersegment membership map (reporting grouping)
    segment_to_supersegment = {}
    for seg_name, seg_config in segments.items():
        ss = resolve_reporting_supersegment(seg_config)
        if ss:
            segment_to_supersegment[seg_name] = ss

    # Process each scenario
    for scenario_suffix in scenarios:
        scenario_name = scenario_name_map.get(scenario_suffix, "base")

        # Process each period (main and MR)
        for period in ["main", "mr"]:
            # Collect metrics by segment
            segment_metrics = {}
            for seg_name in segments:
                seg_dir = output_base / seg_name
                df = load_risk_production_table(seg_dir, period, scenario_suffix)
                if df is not None:
                    metrics = extract_metrics_from_table(df)
                    segment_metrics[seg_name] = metrics

            if not segment_metrics:
                logger.warning(f"No data found for scenario={scenario_name}, period={period}")
                continue

            # Aggregate by supersegment
            supersegment_data = {}
            for ss_name in supersegments:
                ss_segments = [
                    seg_name
                    for seg_name, ss in segment_to_supersegment.items()
                    if ss == ss_name and seg_name in segment_metrics
                ]
                if ss_segments:
                    ss_metrics_list = [segment_metrics[s] for s in ss_segments]
                    agg = aggregate_metrics(ss_metrics_list, multiplier=multiplier)

                    consolidated = ConsolidatedMetrics(
                        group_name=f"supersegment_{ss_name}",
                        period=period,
                        scenario=scenario_name,
                        segments=ss_segments,
                        multiplier=multiplier,
                        multiplier_h3=multiplier_h3,
                        actual_production=agg["actual"]["production"],
                        actual_todu_30ever_h6=agg["actual"]["todu_30ever_h6"],
                        actual_todu_amt_pile_h6=agg["actual"]["todu_amt_pile_h6"],
                        optimum_production=agg["optimum"]["production"],
                        optimum_todu_30ever_h6=agg["optimum"]["todu_30ever_h6"],
                        optimum_todu_amt_pile_h6=agg["optimum"]["todu_amt_pile_h6"],
                        swap_in_production=agg["swap_in"]["production"],
                        swap_in_todu_30ever_h6=agg["swap_in"]["todu_30ever_h6"],
                        swap_in_todu_amt_pile_h6=agg["swap_in"]["todu_amt_pile_h6"],
                        swap_out_production=agg["swap_out"]["production"],
                        swap_out_todu_30ever_h6=agg["swap_out"]["todu_30ever_h6"],
                        swap_out_todu_amt_pile_h6=agg["swap_out"]["todu_amt_pile_h6"],
                        # H3 complementary metrics
                        actual_todu_30ever_h3=agg["actual"].get("todu_30ever_h3", 0),
                        actual_todu_amt_pile_h3=agg["actual"].get("todu_amt_pile_h3", 0),
                        optimum_todu_30ever_h3=agg["optimum"].get("todu_30ever_h3", 0),
                        optimum_todu_amt_pile_h3=agg["optimum"].get("todu_amt_pile_h3", 0),
                        swap_in_todu_30ever_h3=agg["swap_in"].get("todu_30ever_h3", 0),
                        swap_in_todu_amt_pile_h3=agg["swap_in"].get("todu_amt_pile_h3", 0),
                        swap_out_todu_30ever_h3=agg["swap_out"].get("todu_30ever_h3", 0),
                        swap_out_todu_amt_pile_h3=agg["swap_out"].get("todu_amt_pile_h3", 0),
                        # Total demand for rejection rate
                        total_demand=agg.get("_total_demand", 0),
                        # Pass aggregated CIs
                        optimum_production_ci_lower=agg["optimum"].get("production_ci_lower", 0),
                        optimum_production_ci_upper=agg["optimum"].get("production_ci_upper", 0),
                        optimum_risk_ci_lower=agg["optimum"].get("risk_ci_lower", 0),
                        optimum_risk_ci_upper=agg["optimum"].get("risk_ci_upper", 0),
                    )
                    results.append(consolidated.to_dict())
                    supersegment_data[ss_name] = agg

            # Add all individual segments (including those in supersegments)
            for seg_name, metrics in segment_metrics.items():
                agg = aggregate_metrics([metrics], multiplier=multiplier)
                # Determine group name based on supersegment membership
                if seg_name in segment_to_supersegment:
                    ss_name = segment_to_supersegment[seg_name]
                    group_name = f"{ss_name}/{seg_name}"
                else:
                    group_name = f"segment_{seg_name}"

                consolidated = ConsolidatedMetrics(
                    group_name=group_name,
                    period=period,
                    scenario=scenario_name,
                    segments=[seg_name],
                    multiplier=multiplier,
                    multiplier_h3=multiplier_h3,
                    actual_production=agg["actual"]["production"],
                    actual_todu_30ever_h6=agg["actual"]["todu_30ever_h6"],
                    actual_todu_amt_pile_h6=agg["actual"]["todu_amt_pile_h6"],
                    optimum_production=agg["optimum"]["production"],
                    optimum_todu_30ever_h6=agg["optimum"]["todu_30ever_h6"],
                    optimum_todu_amt_pile_h6=agg["optimum"]["todu_amt_pile_h6"],
                    swap_in_production=agg["swap_in"]["production"],
                    swap_in_todu_30ever_h6=agg["swap_in"]["todu_30ever_h6"],
                    swap_in_todu_amt_pile_h6=agg["swap_in"]["todu_amt_pile_h6"],
                    swap_out_production=agg["swap_out"]["production"],
                    swap_out_todu_30ever_h6=agg["swap_out"]["todu_30ever_h6"],
                    swap_out_todu_amt_pile_h6=agg["swap_out"]["todu_amt_pile_h6"],
                    # H3 complementary metrics
                    actual_todu_30ever_h3=agg["actual"].get("todu_30ever_h3", 0),
                    actual_todu_amt_pile_h3=agg["actual"].get("todu_amt_pile_h3", 0),
                    optimum_todu_30ever_h3=agg["optimum"].get("todu_30ever_h3", 0),
                    optimum_todu_amt_pile_h3=agg["optimum"].get("todu_amt_pile_h3", 0),
                    swap_in_todu_30ever_h3=agg["swap_in"].get("todu_30ever_h3", 0),
                    swap_in_todu_amt_pile_h3=agg["swap_in"].get("todu_amt_pile_h3", 0),
                    swap_out_todu_30ever_h3=agg["swap_out"].get("todu_30ever_h3", 0),
                    swap_out_todu_amt_pile_h3=agg["swap_out"].get("todu_amt_pile_h3", 0),
                    # Total demand for rejection rate
                    total_demand=agg.get("_total_demand", 0),
                    # Pass segment-level CIs (fully available)
                    optimum_production_ci_lower=metrics["optimum"].get("production_ci_lower", 0),
                    optimum_production_ci_upper=metrics["optimum"].get("production_ci_upper", 0),
                    optimum_risk_ci_lower=metrics["optimum"].get("risk_ci_lower", 0),
                    optimum_risk_ci_upper=metrics["optimum"].get("risk_ci_upper", 0),
                )
                results.append(consolidated.to_dict())

            # Aggregate total across all segments
            all_metrics_list = list(segment_metrics.values())
            if all_metrics_list:
                total_agg = aggregate_metrics(all_metrics_list, multiplier=multiplier)
                total_consolidated = ConsolidatedMetrics(
                    group_name="TOTAL",
                    period=period,
                    scenario=scenario_name,
                    segments=list(segment_metrics.keys()),
                    multiplier=multiplier,
                    multiplier_h3=multiplier_h3,
                    actual_production=total_agg["actual"]["production"],
                    actual_todu_30ever_h6=total_agg["actual"]["todu_30ever_h6"],
                    actual_todu_amt_pile_h6=total_agg["actual"]["todu_amt_pile_h6"],
                    optimum_production=total_agg["optimum"]["production"],
                    optimum_todu_30ever_h6=total_agg["optimum"]["todu_30ever_h6"],
                    optimum_todu_amt_pile_h6=total_agg["optimum"]["todu_amt_pile_h6"],
                    swap_in_production=total_agg["swap_in"]["production"],
                    swap_in_todu_30ever_h6=total_agg["swap_in"]["todu_30ever_h6"],
                    swap_in_todu_amt_pile_h6=total_agg["swap_in"]["todu_amt_pile_h6"],
                    swap_out_production=total_agg["swap_out"]["production"],
                    swap_out_todu_30ever_h6=total_agg["swap_out"]["todu_30ever_h6"],
                    swap_out_todu_amt_pile_h6=total_agg["swap_out"]["todu_amt_pile_h6"],
                    # H3 complementary metrics
                    actual_todu_30ever_h3=total_agg["actual"].get("todu_30ever_h3", 0),
                    actual_todu_amt_pile_h3=total_agg["actual"].get("todu_amt_pile_h3", 0),
                    optimum_todu_30ever_h3=total_agg["optimum"].get("todu_30ever_h3", 0),
                    optimum_todu_amt_pile_h3=total_agg["optimum"].get("todu_amt_pile_h3", 0),
                    swap_in_todu_30ever_h3=total_agg["swap_in"].get("todu_30ever_h3", 0),
                    swap_in_todu_amt_pile_h3=total_agg["swap_in"].get("todu_amt_pile_h3", 0),
                    swap_out_todu_30ever_h3=total_agg["swap_out"].get("todu_30ever_h3", 0),
                    swap_out_todu_amt_pile_h3=total_agg["swap_out"].get("todu_amt_pile_h3", 0),
                    # Total demand for rejection rate
                    total_demand=total_agg.get("_total_demand", 0),
                    # Pass aggregated CIs
                    optimum_production_ci_lower=total_agg["optimum"].get("production_ci_lower", 0),
                    optimum_production_ci_upper=total_agg["optimum"].get("production_ci_upper", 0),
                    optimum_risk_ci_lower=total_agg["optimum"].get("risk_ci_lower", 0),
                    optimum_risk_ci_upper=total_agg["optimum"].get("risk_ci_upper", 0),
                )
                results.append(total_consolidated.to_dict())

    df = pd.DataFrame(results)
    df = patch_consolidated_production_from_segment_audits(df, output_base, segments)

    # Reorder columns for clarity
    column_order = [
        "group",
        "period",
        "scenario",
        "n_segments",
        "segments",
        "actual_production",
        "actual_risk_pct",
        "actual_todu_30ever_h6",
        "actual_todu_amt_pile_h6",
        "optimum_production",
        "optimum_risk_pct",
        "optimum_todu_30ever_h6",
        "optimum_todu_amt_pile_h6",
        "production_delta",
        "production_delta_pct",
        "risk_delta_pct",
        "swap_in_production",
        "swap_in_risk_pct",
        "swap_in_todu_30ever_h6",
        "swap_in_todu_amt_pile_h6",
        "swap_out_production",
        "swap_out_risk_pct",
        "swap_out_todu_30ever_h6",
        "swap_out_todu_amt_pile_h6",
        "total_demand",
        "actual_rejection_rate_pct",
        "optimum_rejection_rate_pct",
        "production_ci_lower",
        "production_ci_upper",
        "risk_ci_lower",
        "risk_ci_upper",
    ]
    df = df[[c for c in column_order if c in df.columns]]

    return df


def _display_group_name(group: str) -> str:
    group = str(group)
    if group == "TOTAL":
        return "Total"
    if group.startswith("supersegment_"):
        return f"Supersegment · {group.replace('supersegment_', '', 1)}"
    if "/" in group:
        parent, child = group.split("/", 1)
        return f"{child} · {parent}"
    if group.startswith("segment_"):
        return group.replace("segment_", "", 1)
    return group


def _sort_consolidated_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df.copy()

    ordered = df.copy()
    ordered["_scenario_order"] = ordered["scenario"].map({"pessimistic": 0, "base": 1, "optimistic": 2}).fillna(99)
    ordered["_period_order"] = ordered["period"].map({"main": 0, "mr": 1}).fillna(99)
    ordered["_group_order"] = ordered["group"].map(
        lambda value: 0 if value == "TOTAL" else 1 if str(value).startswith("supersegment_") else 2
    )
    ordered["group_display"] = ordered["group"].map(_display_group_name)
    return ordered.sort_values(
        ["_group_order", "group_display", "_period_order", "_scenario_order"],
        kind="stable",
    )


def create_consolidation_dashboard(df: pd.DataFrame, title: str = "Consolidated Risk Production Report") -> go.Figure:
    """
    Create an interactive dashboard for consolidated metrics.

    Args:
        df: Consolidated DataFrame from consolidate_segments()
        title: Dashboard title

    Returns:
        Plotly Figure
    """
    if df.empty:
        fig = go.Figure()
        fig.add_annotation(text="No data available", x=0.5, y=0.5, showarrow=False)
        return fig

    ordered_df = _sort_consolidated_rows(df)
    total_df = ordered_df[ordered_df["group"] == "TOTAL"].copy()
    if total_df.empty:
        total_df = ordered_df.copy()

    segment_df = ordered_df[
        ~(ordered_df["group"].eq("TOTAL") | ordered_df["group"].astype(str).str.startswith("supersegment_"))
    ].copy()
    focus_scenario = (
        "base" if "base" in ordered_df["scenario"].astype(str).values else str(ordered_df.iloc[0]["scenario"])
    )
    focus_label = focus_scenario.title()

    fig = make_subplots(
        rows=4,
        cols=2,
        specs=[
            [{"type": "indicator"}, {"type": "indicator"}],
            [{"type": "indicator"}, {"type": "indicator"}],
            [{"type": "bar"}, {"type": "scatter"}],
            [{"type": "heatmap"}, {"type": "bar"}],
        ],
        subplot_titles=(
            "Main Base — Optimum Production",
            "Main Base — Optimum Risk",
            "MR Base — Optimum Production",
            "MR Base — Optimum Risk",
            "Total Production by Scenario",
            "Risk vs Production by Scenario",
            f"{focus_label} Scenario — Production Delta Heatmap",
            f"Top Groups — {focus_label} Scenario",
        ),
        vertical_spacing=0.08,
        horizontal_spacing=0.08,
        row_heights=[0.16, 0.16, 0.34, 0.34],
    )

    def _pick_total_row(period: str) -> pd.Series | None:
        rows = total_df[(total_df["period"] == period) & (total_df["scenario"] == "base")]
        if rows.empty:
            rows = total_df[total_df["period"] == period].sort_values(["_scenario_order"]).head(1)
        return rows.iloc[0] if not rows.empty else None

    def _add_indicator(
        row_data: pd.Series | None,
        *,
        row: int,
        col: int,
        title_text: str,
        value_key: str,
        reference_key: str,
        prefix: str = "",
        suffix: str = "",
        relative: bool = False,
        valueformat: str = ",.0f",
        inverse_delta: bool = False,
    ) -> None:
        if row_data is None:
            fig.add_trace(
                go.Indicator(
                    mode="number",
                    value=0,
                    number={"prefix": prefix, "suffix": suffix, "valueformat": valueformat},
                    title={"text": f"<span style='font-size:0.88em;color:#94a3b8'>{title_text}<br>No data</span>"},
                ),
                row=row,
                col=col,
            )
            return

        value = float(row_data.get(value_key, 0) or 0)
        reference_value = float(row_data.get(reference_key, 0) or 0)
        indicator = go.Indicator(
            mode="number+delta",
            value=value,
            number={
                "prefix": prefix,
                "suffix": suffix,
                "valueformat": valueformat,
                "font": {"color": "#0f172a", "size": 30},
            },
            title={
                "text": (
                    f"<span style='font-size:0.88em;color:#0f172a'>{title_text}</span><br>"
                    f"<span style='font-size:0.74em;color:#64748b'>Actual baseline: {prefix}{reference_value:{valueformat}}{suffix}</span>"
                )
            },
        )
        if reference_value != 0:
            indicator.delta = {
                "reference": reference_value,
                "relative": relative,
                "valueformat": ".1%" if relative else ".2f",
                "increasing": {"color": "#dc2626" if inverse_delta else "#2563eb"},
                "decreasing": {"color": "#059669" if inverse_delta else "#dc2626"},
            }
        fig.add_trace(indicator, row=row, col=col)

    _add_indicator(
        _pick_total_row("main"),
        row=1,
        col=1,
        title_text="Main period optimum production",
        value_key="optimum_production",
        reference_key="actual_production",
        prefix="€",
        relative=True,
    )
    _add_indicator(
        _pick_total_row("main"),
        row=1,
        col=2,
        title_text="Main period optimum risk",
        value_key="optimum_risk_pct",
        reference_key="actual_risk_pct",
        suffix="%",
        valueformat=".2f",
        inverse_delta=True,
    )
    _add_indicator(
        _pick_total_row("mr"),
        row=2,
        col=1,
        title_text="MR period optimum production",
        value_key="optimum_production",
        reference_key="actual_production",
        prefix="€",
        relative=True,
    )
    _add_indicator(
        _pick_total_row("mr"),
        row=2,
        col=2,
        title_text="MR period optimum risk",
        value_key="optimum_risk_pct",
        reference_key="actual_risk_pct",
        suffix="%",
        valueformat=".2f",
        inverse_delta=True,
    )

    period_palette = {
        "main": {"actual": "#93c5fd", "optimum": "#2563eb", "line": "#1d4ed8"},
        "mr": {"actual": "#fdba74", "optimum": "#f97316", "line": "#ea580c"},
    }
    scenario_category = ["Pessimistic", "Base", "Optimistic"]
    for period in ["main", "mr"]:
        period_data = total_df[total_df["period"] == period].sort_values(["_scenario_order"])
        if period_data.empty:
            continue

        scenario_labels = [str(value).title() for value in period_data["scenario"]]
        prod_upper = None
        prod_lower = None
        error_visible = False
        if {"production_ci_upper", "production_ci_lower"}.issubset(period_data.columns):
            prod_upper = (period_data["production_ci_upper"] - period_data["optimum_production"]).clip(lower=0)
            prod_lower = (period_data["optimum_production"] - period_data["production_ci_lower"]).clip(lower=0)
            error_visible = bool(np.any((prod_upper.fillna(0) + prod_lower.fillna(0)).to_numpy() > 0))

        fig.add_trace(
            go.Bar(
                name=f"{period.upper()} Actual",
                x=scenario_labels,
                y=period_data["actual_production"],
                marker_color=period_palette[period]["actual"],
                opacity=0.82,
                legendgroup=f"{period}-production",
                offsetgroup=f"{period}-actual",
                hovertemplate=(
                    "<b>%{x}</b><br>Period: " + period.upper() + "<br>Actual production: €%{y:,.0f}<extra></extra>"
                ),
            ),
            row=3,
            col=1,
        )
        fig.add_trace(
            go.Bar(
                name=f"{period.upper()} Optimum",
                x=scenario_labels,
                y=period_data["optimum_production"],
                marker_color=period_palette[period]["optimum"],
                legendgroup=f"{period}-production",
                offsetgroup=f"{period}-optimum",
                text=[f"{value:+.1f}%" for value in period_data["production_delta_pct"]],
                textposition="outside",
                cliponaxis=False,
                customdata=np.column_stack([period_data["risk_delta_pct"].to_numpy()]),
                error_y={
                    "type": "data",
                    "array": prod_upper.tolist() if prod_upper is not None else [0] * len(period_data),
                    "arrayminus": prod_lower.tolist() if prod_lower is not None else [0] * len(period_data),
                    "visible": error_visible,
                },
                hovertemplate=(
                    "<b>%{x}</b><br>Period: "
                    + period.upper()
                    + "<br>Optimum production: €%{y:,.0f}"
                    + "<br>Production delta: %{text}<br>Risk delta: %{customdata[0]:+.2f} pp<extra></extra>"
                ),
            ),
            row=3,
            col=1,
        )
        fig.add_trace(
            go.Scatter(
                name=f"{period.upper()} Optimum path",
                x=period_data["optimum_risk_pct"],
                y=period_data["optimum_production"],
                mode="lines+markers+text",
                text=scenario_labels,
                textposition="top center",
                line={"color": period_palette[period]["line"], "width": 3},
                marker={
                    "size": 12,
                    "color": period_palette[period]["optimum"],
                    "line": {"width": 1.5, "color": "#ffffff"},
                },
                showlegend=False,
                customdata=np.column_stack([period_data["optimum_rejection_rate_pct"].to_numpy()]),
                hovertemplate=(
                    "<b>%{text}</b><br>Period: "
                    + period.upper()
                    + "<br>Optimum risk: %{x:.2f}%"
                    + "<br>Optimum production: €%{y:,.0f}<br>Optimum rejection rate: %{customdata[0]:.1f}%<extra></extra>"
                ),
            ),
            row=3,
            col=2,
        )
        baseline = _pick_total_row(period)
        if baseline is not None:
            fig.add_trace(
                go.Scatter(
                    name=f"{period.upper()} Actual baseline",
                    x=[baseline.get("actual_risk_pct", 0)],
                    y=[baseline.get("actual_production", 0)],
                    mode="markers+text",
                    text=[f"{period.upper()} Actual"],
                    textposition="bottom center",
                    marker={
                        "size": 15,
                        "color": period_palette[period]["actual"],
                        "symbol": "diamond",
                        "line": {"width": 2, "color": period_palette[period]["line"]},
                    },
                    showlegend=False,
                    hovertemplate=(
                        "<b>Actual baseline</b><br>Period: "
                        + period.upper()
                        + "<br>Risk: %{x:.2f}%"
                        + "<br>Production: €%{y:,.0f}<extra></extra>"
                    ),
                ),
                row=3,
                col=2,
            )

    heat_df = ordered_df[ordered_df["scenario"] == focus_scenario].copy()
    if not heat_df.empty:
        if heat_df["group_display"].nunique() > 12:
            ranked_groups = (
                heat_df.assign(_rank_value=heat_df["production_delta"].abs())
                .sort_values(["_group_order", "_rank_value"], ascending=[True, False])["group_display"]
                .drop_duplicates()
                .head(12)
            )
            heat_df = heat_df[heat_df["group_display"].isin(ranked_groups)]

        heat_df = heat_df.sort_values(["_group_order", "group_display", "_period_order"])
        heat_index = heat_df["group_display"].drop_duplicates().tolist()
        heat_cols = [value for value in ["main", "mr"] if value in heat_df["period"].values]
        heat_pivot = heat_df.pivot_table(
            index="group_display", columns="period", values="production_delta_pct", aggfunc="first"
        )
        heat_pivot = heat_pivot.reindex(index=heat_index, columns=heat_cols)
        heat_eur = heat_df.pivot_table(
            index="group_display", columns="period", values="production_delta", aggfunc="first"
        )
        heat_eur = heat_eur.reindex(index=heat_index, columns=heat_cols)
        heat_risk = heat_df.pivot_table(
            index="group_display", columns="period", values="risk_delta_pct", aggfunc="first"
        )
        heat_risk = heat_risk.reindex(index=heat_index, columns=heat_cols)

        heat_text = []
        hovertext = []
        for group_name in heat_pivot.index:
            text_row = []
            hover_row = []
            for period_name in heat_pivot.columns:
                pct_val = heat_pivot.loc[group_name, period_name]
                eur_val = heat_eur.loc[group_name, period_name]
                risk_val = heat_risk.loc[group_name, period_name]
                if pd.isna(pct_val):
                    text_row.append("")
                    hover_row.append(f"<b>{group_name}</b><br>Period: {str(period_name).upper()}<br>No data")
                else:
                    text_row.append(f"{pct_val:+.1f}%")
                    hover_row.append(
                        f"<b>{group_name}</b><br>Period: {str(period_name).upper()}"
                        f"<br>Production delta: €{eur_val:,.0f}"
                        f"<br>Production delta %: {pct_val:+.1f}%"
                        f"<br>Risk delta: {risk_val:+.2f} pp"
                    )
            heat_text.append(text_row)
            hovertext.append(hover_row)

        fig.add_trace(
            go.Heatmap(
                x=[str(value).upper() for value in heat_pivot.columns],
                y=heat_pivot.index.tolist(),
                z=heat_pivot.to_numpy(),
                text=heat_text,
                texttemplate="%{text}",
                textfont={"size": 11},
                hovertext=hovertext,
                hovertemplate="%{hovertext}<extra></extra>",
                zmid=0,
                colorscale=[[0.0, "#b91c1c"], [0.5, "#f8fafc"], [1.0, "#15803d"]],
                colorbar={"title": "Prod Δ %", "ticksuffix": "%"},
            ),
            row=4,
            col=1,
        )

    top_groups = segment_df[segment_df["scenario"] == focus_scenario].copy()
    if top_groups.empty:
        top_groups = ordered_df[(ordered_df["scenario"] == focus_scenario) & (~ordered_df["group"].eq("TOTAL"))].copy()
    if not top_groups.empty:
        ranked = (
            top_groups.groupby("group_display", as_index=False)["production_delta"]
            .max()
            .sort_values("production_delta", ascending=False)
            .head(8)
        )
        ranked_order = ranked.sort_values("production_delta", ascending=True)["group_display"].tolist()
        top_groups = top_groups[top_groups["group_display"].isin(ranked_order)].copy()
        top_groups["group_display"] = pd.Categorical(top_groups["group_display"], categories=ranked_order, ordered=True)
        top_groups = top_groups.sort_values(["group_display", "_period_order"])

        for period in [value for value in ["main", "mr"] if value in top_groups["period"].values]:
            period_groups = top_groups[top_groups["period"] == period].sort_values("group_display")
            fig.add_trace(
                go.Bar(
                    name=f"{period.upper()} Production Δ",
                    y=period_groups["group_display"],
                    x=period_groups["production_delta"],
                    orientation="h",
                    marker_color=period_palette[period]["optimum"],
                    text=[f"{value:+.1f}%" for value in period_groups["production_delta_pct"]],
                    textposition="outside",
                    cliponaxis=False,
                    customdata=np.column_stack(
                        [
                            period_groups["risk_delta_pct"].to_numpy(),
                            period_groups["optimum_risk_pct"].to_numpy(),
                        ]
                    ),
                    hovertemplate=(
                        "<b>%{y}</b><br>Period: "
                        + period.upper()
                        + "<br>Production delta: €%{x:,.0f}"
                        + "<br>Production delta %: %{text}<br>Risk delta: %{customdata[0]:+.2f} pp"
                        + "<br>Optimum risk: %{customdata[1]:.2f}%<extra></extra>"
                    ),
                ),
                row=4,
                col=2,
            )

    fig.update_layout(
        title={
            "text": title
            + "<br><sup>Executive portfolio view across scenarios, periods, and segment opportunities</sup>",
            "x": 0.5,
            "xanchor": "center",
        },
        height=1450,
        barmode="group",
        template="plotly_white",
        paper_bgcolor="#f8fafc",
        plot_bgcolor="#ffffff",
        margin={"t": 120, "r": 40, "b": 70, "l": 70},
        font={"family": "Arial, sans-serif", "size": 12, "color": "#0f172a"},
        legend={
            "orientation": "h",
            "yanchor": "bottom",
            "y": -0.06,
            "xanchor": "center",
            "x": 0.5,
            "bgcolor": "rgba(255,255,255,0.85)",
        },
        hoverlabel={"bgcolor": "#0f172a", "font": {"color": "#ffffff"}},
    )
    fig.update_annotations(font={"size": 13, "color": "#0f172a"})
    fig.update_xaxes(categoryorder="array", categoryarray=scenario_category, title_text="Scenario", row=3, col=1)
    fig.update_yaxes(title_text="Production (€)", tickformat=",.0f", row=3, col=1)
    fig.update_xaxes(title_text="Risk (%)", ticksuffix="%", row=3, col=2)
    fig.update_yaxes(title_text="Production (€)", tickformat=",.0f", row=3, col=2)
    fig.update_xaxes(title_text="Period", row=4, col=1)
    fig.update_yaxes(title_text="Group", automargin=True, row=4, col=1)
    fig.update_xaxes(
        title_text="Production Delta (€)",
        tickformat=",.0f",
        zeroline=True,
        zerolinecolor="#cbd5e1",
        row=4,
        col=2,
    )
    fig.update_yaxes(title_text="Group", automargin=True, row=4, col=2)
    return fig


def generate_consolidation_report(
    output_base: str,
    segments: dict[str, dict[str, Any]],
    supersegments: dict[str, dict[str, Any]],
    scenarios: list[str] | None = None,
    output_path: str | None = None,
    multiplier: float = float(DEFAULT_RISK_MULTIPLIER),
    multiplier_h3: float = float(DEFAULT_RISK_MULTIPLIER_H3),
) -> tuple[pd.DataFrame, go.Figure]:
    """
    Generate complete consolidation report with CSV and HTML dashboard.

    Args:
        output_base: Base output directory
        segments: Segment configurations
        supersegments: Supersegment configurations
        scenarios: List of scenario suffixes
        output_path: Optional output path for files (defaults to output_base)
        multiplier: H6 risk multiplier (from config.toml)
        multiplier_h3: H3 risk multiplier (from config.toml)

    Returns:
        Tuple of (consolidated DataFrame, Plotly figure)
    """
    output_base = Path(output_base)
    output_path = Path(output_path) if output_path else output_base

    logger.info("Generating consolidated risk production report...")

    # Consolidate data
    df = consolidate_segments(
        output_base, segments, supersegments, scenarios, multiplier=multiplier, multiplier_h3=multiplier_h3
    )

    if df.empty:
        logger.warning("No data found to consolidate")
        return df, None

    # Save CSV
    csv_path = output_path / "consolidated_risk_production.csv"
    df.to_csv(csv_path, index=False)
    logger.info(f"Consolidated CSV saved to {csv_path}")

    # Create dashboard
    fig = create_consolidation_dashboard(df)

    # Save HTML
    html_path = output_path / "consolidated_risk_production.html"
    fig.write_html(
        str(html_path),
        config={
            "displaylogo": False,
            "responsive": True,
            "modeBarButtonsToRemove": ["lasso2d", "select2d"],
        },
    )
    logger.info(f"Consolidated dashboard saved to {html_path}")

    # Export Excel workbook
    try:
        xlsx_path = export_consolidated_excel(df, output_base, segments, supersegments)
        logger.info(f"Consolidated Excel saved to {xlsx_path}")
    except Exception as e:
        logger.warning(f"Excel export failed: {e}")

    # Print summary
    print_consolidation_summary(df)

    return df, fig


# =============================================================================
# Excel-export design tokens + column classifications (R2b todo #57)
# =============================================================================
# Extracted from the body of `export_consolidated_excel` so the function
# stops being a 2000-line monolith. These constants are referenced by the
# nested sheet-builder helpers still inside the function; they are not
# intended for external import (all prefixed with `_`). Module-level
# placement makes them (a) easy to inspect in one block and (b) available
# to future module-level helpers extracted from the function body.
from openpyxl.chart import BarChart as _BarChart  # noqa: E402
from openpyxl.chart import Reference as _Reference  # noqa: E402
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side  # noqa: E402
from openpyxl.utils import get_column_letter as _get_column_letter  # noqa: E402

# ----- Core brand colors -----
_CLR_PRIMARY = "1B2A4A"  # Deep navy — titles, header bg
_CLR_PRIMARY_LIGHT = "34495E"  # Lighter navy — secondary headers
_CLR_ACCENT = "2980B9"  # Cerulean blue — KPI accents, links
_CLR_ACCENT_LIGHT = "D6EAF8"  # Pale blue — KPI card bg
_CLR_WHITE = "FFFFFF"

# ----- Semantic colors -----
_CLR_GOOD = "1ABC9C"  # Teal-green — softer than pure green
_CLR_GOOD_LIGHT = "D1F2EB"  # Pale teal — TOTAL row bg
_CLR_GOOD_DARK = "0E6655"  # Dark teal — TOTAL row text
_CLR_BAD = "E74C3C"  # Warm red — risk / negative deltas
_CLR_BAD_LIGHT = "FDEDEC"  # Pale pink — reject cell tint
_CLR_WARN = "F39C12"  # Amber — cutoff tab, warnings
_CLR_MR_BG = "FEF5E7"  # Warm cream — MR KPI tint
_CLR_MR_FG = "CA6F1E"  # Dark amber — MR labels

# ----- Neutral palette -----
_CLR_NEUTRAL_LIGHT = "F8F9FA"  # Near-white stripe
_CLR_NEUTRAL = "DEE2E6"  # Soft grey — table borders
_CLR_NEUTRAL_MID = "AEB6BF"  # Mid grey — subtle text
_CLR_TEXT = "2C3E50"  # Dark grey — body text

# ----- Acceptance grid tones -----
_CLR_GRID_ACCEPT = "58D68D"  # Soft green
_CLR_GRID_REJECT = "EC7063"  # Soft red-coral
_CLR_GRID_NA = "D5DBDB"  # Light warm grey
_CLR_GRID_HDR = "2C3E50"  # Dark header for contrast

# ----- Section styling -----
_CLR_SECTION_BG = "EBF5FB"  # Pale blue — section header bg
_CLR_SECTION_BAR = "2980B9"  # Accent bar

# ----- Row highlights — RP sheets -----
_CLR_OPTIMUM_BG = "D4EFDF"  # Pale green — Optimum selected row
_CLR_OPTIMUM_FG = "1E8449"  # Dark green — Optimum selected text
_CLR_SUMMARY_BG = "D6EAF8"  # Pale blue — Summary / delta row
_CLR_SUMMARY_FG = "1B4F72"  # Dark blue — Summary text

# ----- Sheet-tab colours -----
_CLR_TAB_EXEC = "2980B9"
_CLR_TAB_PORTFOLIO = "1B2A4A"
_CLR_TAB_SEGMENT = "5DADE2"
_CLR_TAB_SEGMENT_MR = "F39C12"
_CLR_TAB_CUTOFF = "F39C12"
_CLR_TAB_GRID = "1ABC9C"

# ----- Fonts -----
_FN = "Calibri"
_FONT_TITLE = Font(bold=True, color=_CLR_PRIMARY, size=20, name=_FN)
_FONT_SUBTITLE = Font(bold=False, color=_CLR_NEUTRAL_MID, size=11, name=_FN)
_FONT_SECTION = Font(bold=True, color=_CLR_PRIMARY, size=12, name=_FN)
_FONT_KPI_VALUE = Font(bold=True, color=_CLR_PRIMARY, size=24, name=_FN)
_FONT_KPI_LABEL = Font(bold=False, color=_CLR_NEUTRAL_MID, size=9, name=_FN)
_FONT_KPI_DELTA_POS = Font(bold=True, color=_CLR_GOOD, size=12, name=_FN)
_FONT_KPI_DELTA_NEG = Font(bold=True, color=_CLR_BAD, size=12, name=_FN)
_FONT_HEADER = Font(bold=True, color=_CLR_WHITE, size=10, name=_FN)
_FONT_DATA = Font(color=_CLR_TEXT, size=10, name=_FN)
_FONT_TOTAL = Font(bold=True, color=_CLR_GOOD_DARK, size=10, name=_FN)
_FONT_MR_LABEL = Font(bold=True, color=_CLR_MR_FG, size=11, name=_FN)
_FONT_GRID_LABEL = Font(bold=True, color=_CLR_PRIMARY, size=9, name=_FN)
_FONT_GRID_CELL = Font(bold=True, color=_CLR_WHITE, size=11, name=_FN)
_FONT_GRID_HEADER = Font(bold=True, color=_CLR_WHITE, size=9, name=_FN)

# ----- Fills -----
_FILL_HEADER = PatternFill(start_color=_CLR_PRIMARY, end_color=_CLR_PRIMARY, fill_type="solid")
_FILL_STRIPE = PatternFill(start_color=_CLR_NEUTRAL_LIGHT, end_color=_CLR_NEUTRAL_LIGHT, fill_type="solid")
_FILL_TOTAL = PatternFill(start_color=_CLR_GOOD_LIGHT, end_color=_CLR_GOOD_LIGHT, fill_type="solid")
_FILL_KPI = PatternFill(start_color=_CLR_ACCENT_LIGHT, end_color=_CLR_ACCENT_LIGHT, fill_type="solid")
_FILL_MR = PatternFill(start_color=_CLR_MR_BG, end_color=_CLR_MR_BG, fill_type="solid")
_FILL_SECTION = PatternFill(start_color=_CLR_SECTION_BG, end_color=_CLR_SECTION_BG, fill_type="solid")
_FILL_ACCEPT = PatternFill(start_color=_CLR_GRID_ACCEPT, end_color=_CLR_GRID_ACCEPT, fill_type="solid")
_FILL_REJECT = PatternFill(start_color=_CLR_GRID_REJECT, end_color=_CLR_GRID_REJECT, fill_type="solid")
_FILL_NA = PatternFill(start_color=_CLR_GRID_NA, end_color=_CLR_GRID_NA, fill_type="solid")
_FILL_GRID_HEADER = PatternFill(start_color=_CLR_GRID_HDR, end_color=_CLR_GRID_HDR, fill_type="solid")
_FILL_OPTIMUM = PatternFill(start_color=_CLR_OPTIMUM_BG, end_color=_CLR_OPTIMUM_BG, fill_type="solid")
_FILL_SUMMARY = PatternFill(start_color=_CLR_SUMMARY_BG, end_color=_CLR_SUMMARY_BG, fill_type="solid")

# Row-highlight fonts (depend on colors above — must appear after them)
_FONT_OPTIMUM = Font(bold=True, color=_CLR_OPTIMUM_FG, size=10, name=_FN)
_FONT_SUMMARY = Font(bold=True, color=_CLR_SUMMARY_FG, size=10, name=_FN)

# ----- Borders -----
_THIN = Side(style="thin", color=_CLR_NEUTRAL)
_HAIR = Side(style="hair", color=_CLR_NEUTRAL)
_BORDER_ALL = Border(top=_HAIR, bottom=_HAIR, left=_HAIR, right=_HAIR)
_BORDER_HEADER = Border(
    top=Side(style="thin", color=_CLR_PRIMARY),
    bottom=Side(style="medium", color=_CLR_ACCENT),
    left=_HAIR,
    right=_HAIR,
)
_BORDER_BOTTOM = Border(bottom=Side(style="medium", color=_CLR_ACCENT))
_ACCENT_LEFT = Border(
    left=Side(style="thick", color=_CLR_ACCENT),
    top=_HAIR,
    bottom=_HAIR,
    right=_HAIR,
)
_SECTION_LEFT = Border(left=Side(style="thick", color=_CLR_SECTION_BAR))
_BORDER_GRID = Border(
    top=Side(style="medium", color=_CLR_WHITE),
    bottom=Side(style="medium", color=_CLR_WHITE),
    left=Side(style="medium", color=_CLR_WHITE),
    right=Side(style="medium", color=_CLR_WHITE),
)

# ----- Alignment -----
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")


# =============================================================================
# Column classifications for Excel formatting
# =============================================================================
_CURRENCY_COLS = {
    "actual_production",
    "optimum_production",
    "swap_in_production",
    "swap_out_production",
    "production_delta",
    "production_ci_lower",
    "production_ci_upper",
    "total_demand",
    "Production (€)",
    "Total Demand (€)",
}
_PCT_COLS = {
    "actual_risk_pct",
    "optimum_risk_pct",
    "swap_in_risk_pct",
    "swap_out_risk_pct",
    "production_delta_pct",
    "risk_delta_pct",
    "risk_ci_lower",
    "risk_ci_upper",
    "actual_rejection_rate_pct",
    "optimum_rejection_rate_pct",
    "actual_risk_h3_pct",
    "optimum_risk_h3_pct",
    "swap_in_risk_h3_pct",
    "swap_out_risk_h3_pct",
    "Risk (%)",
    "Risk H3 (%)",
    "Production (%)",
    "Rejection Rate (%)",
}
_INTEGER_COLS = {
    "n_segments",
    "actual_todu_30ever_h6",
    "actual_todu_amt_pile_h6",
    "optimum_todu_30ever_h6",
    "optimum_todu_amt_pile_h6",
    "swap_in_todu_30ever_h6",
    "swap_in_todu_amt_pile_h6",
    "swap_out_todu_30ever_h6",
    "swap_out_todu_amt_pile_h6",
    "actual_todu_30ever_h3",
    "actual_todu_amt_pile_h3",
    "optimum_todu_30ever_h3",
    "optimum_todu_amt_pile_h3",
    "swap_in_todu_30ever_h3",
    "swap_in_todu_amt_pile_h3",
    "swap_out_todu_30ever_h3",
    "swap_out_todu_amt_pile_h3",
}
_TEXT_COLS = {"group", "period", "scenario", "segments", "Metric", "segment"}
_DELTA_COLS = {"production_delta", "production_delta_pct", "risk_delta_pct"}
_CUTOFF_FIXED_COLS = frozenset(
    {
        "accepted",
        "segment",
        "scenario",
        "risk_pct",
        "production",
        "production_ci_lower",
        "production_ci_upper",
        "risk_ci_lower",
        "risk_ci_upper",
    }
)

_COLUMN_LABELS = {
    "group": "Group",
    "period": "Period",
    "scenario": "Scenario",
    "n_segments": "# Segments",
    "segments": "Segments",
    "actual_production": "Actual Production (€)",
    "actual_risk_pct": "Actual Risk (%)",
    "optimum_production": "Optimum Production (€)",
    "optimum_risk_pct": "Optimum Risk (%)",
    "production_delta": "Production Delta (€)",
    "production_delta_pct": "Production Delta (%)",
    "risk_delta_pct": "Risk Delta (%)",
    "swap_in_production": "Swap-In Production (€)",
    "swap_in_risk_pct": "Swap-In Risk (%)",
    "swap_out_production": "Swap-Out Production (€)",
    "swap_out_risk_pct": "Swap-Out Risk (%)",
    "total_demand": "Total Demand (€)",
    "actual_rejection_rate_pct": "Actual Rejection Rate (%)",
    "optimum_rejection_rate_pct": "Optimum Rejection Rate (%)",
    "actual_risk_h3_pct": "Actual Risk H3 (%)",
    "optimum_risk_h3_pct": "Optimum Risk H3 (%)",
    "swap_in_risk_h3_pct": "Swap-In Risk H3 (%)",
    "swap_out_risk_h3_pct": "Swap-Out Risk H3 (%)",
    "production_ci_lower": "Production CI Lower (€)",
    "production_ci_upper": "Production CI Upper (€)",
    "risk_ci_lower": "Risk CI Lower (%)",
    "risk_ci_upper": "Risk CI Upper (%)",
}


# =============================================================================
# Pure Excel-styling helpers (R2b todo #57 step 2)
# =============================================================================
# Extracted from the body of `export_consolidated_excel`. These take an
# openpyxl worksheet + primitives and have no closure over segment state,
# so lifting them to module level is safe and shrinks the 2000-line function
# by another ~60 lines. They reference only the module-level design tokens
# defined above.


def _set_col_width(ws, col_idx: int, header_text, max_rows: int) -> None:
    """Size column *col_idx* by the longer of its header label and data width (cap 34)."""
    letter = _get_column_letter(col_idx)
    label_len = len(str(header_text))
    data_max = max(
        (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(2, max_rows + 1)),
        default=0,
    )
    ws.column_dimensions[letter].width = min(max(label_len, data_max) + 3, 34)


def _apply_number_format(cell, col_name: str) -> None:
    """Assign a currency / percentage / integer number format based on *col_name*."""
    if col_name in _CURRENCY_COLS:
        cell.number_format = '#,##0 "€"'
    elif col_name in _PCT_COLS:
        cell.number_format = "0.00"
    elif col_name in _INTEGER_COLS:
        cell.number_format = "#,##0"


def _apply_page_setup(ws) -> None:
    """Set landscape orientation, fit-to-width printing, hide on-screen gridlines."""
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _style_table(ws, df_cols, *, header_row: int = 1, highlight_total: bool = True) -> None:
    """Apply full dashboard styling to a data table starting at header_row.

    Pure helper — references only module-level design tokens and
    column-classification sets. Extracted in R2b todo #57 step 3.
    """
    group_col_idx = None
    metric_col_idx = None
    n_cols = len(df_cols)
    for col_idx, col_name in enumerate(df_cols, 1):
        if col_name == "group":
            group_col_idx = col_idx
        if col_name == "Metric":
            metric_col_idx = col_idx
        cell = ws.cell(row=header_row, column=col_idx)
        cell.value = _COLUMN_LABELS.get(col_name, col_name)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
        cell.border = _BORDER_HEADER
        _set_col_width(ws, col_idx, cell.value, ws.max_row)
        for r in range(header_row + 1, ws.max_row + 1):
            data_cell = ws.cell(row=r, column=col_idx)
            data_cell.font = _FONT_DATA
            data_cell.border = _BORDER_ALL
            data_cell.alignment = _ALIGN_LEFT if col_name in _TEXT_COLS else _ALIGN_RIGHT
            _apply_number_format(data_cell, col_name)

    ws.row_dimensions[header_row].height = 28
    delta_col_indices = {ci for ci, cn in enumerate(df_cols, 1) if cn in _DELTA_COLS}
    for r in range(header_row + 1, ws.max_row + 1):
        ws.row_dimensions[r].height = 22
        is_total = False
        is_optimum = False
        is_summary = False
        if highlight_total and group_col_idx:
            is_total = str(ws.cell(row=r, column=group_col_idx).value or "").strip().lower().startswith("total")
        if metric_col_idx:
            metric_val = str(ws.cell(row=r, column=metric_col_idx).value or "").strip().lower()
            is_optimum = metric_val.startswith("optimum")
            is_summary = metric_val == "summary"
        for col_idx in range(1, n_cols + 1):
            data_cell = ws.cell(row=r, column=col_idx)
            if is_total:
                data_cell.fill = _FILL_TOTAL
                data_cell.font = _FONT_TOTAL
            elif is_optimum:
                data_cell.fill = _FILL_OPTIMUM
                data_cell.font = _FONT_OPTIMUM
            elif is_summary:
                data_cell.fill = _FILL_SUMMARY
                data_cell.font = _FONT_SUMMARY
            elif r % 2 == 0:
                data_cell.fill = _FILL_STRIPE
            if col_idx in delta_col_indices:
                val = data_cell.value
                if isinstance(val, (int, float)):
                    col_name = list(df_cols)[col_idx - 1]
                    is_risk = col_name == "risk_delta_pct"
                    good = val <= 0 if is_risk else val >= 0
                    clr = _CLR_GOOD if good else _CLR_BAD
                    bold = is_total or is_optimum or is_summary
                    data_cell.font = Font(bold=bold, color=clr, size=10, name=_FN)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1).coordinate
    ws.auto_filter.ref = (
        f"{ws.cell(row=header_row, column=1).coordinate}:{ws.cell(row=ws.max_row, column=n_cols).coordinate}"
    )


def _write_kpi_card(ws, row, col, label, value_str, delta_str=None, delta_positive: bool = True) -> None:
    """Write a KPI card block: 2 rows × 2 columns. Pure helper."""
    val_cell = ws.cell(row=row, column=col)
    val_cell.value = value_str
    val_cell.font = _FONT_KPI_VALUE
    val_cell.fill = _FILL_KPI
    val_cell.alignment = Alignment(horizontal="center", vertical="bottom")
    val_cell.border = _ACCENT_LEFT
    if delta_str:
        dc = ws.cell(row=row, column=col + 1)
        dc.value = delta_str
        dc.font = _FONT_KPI_DELTA_POS if delta_positive else _FONT_KPI_DELTA_NEG
        dc.fill = _FILL_KPI
        dc.alignment = Alignment(horizontal="center", vertical="bottom")
        dc.border = _BORDER_ALL
    else:
        ws.cell(row=row, column=col + 1).fill = _FILL_KPI
        ws.cell(row=row, column=col + 1).border = _BORDER_ALL
    lc = ws.cell(row=row + 1, column=col)
    lc.value = label
    lc.font = _FONT_KPI_LABEL
    lc.fill = _FILL_KPI
    lc.alignment = Alignment(horizontal="center", vertical="top")
    lc.border = _ACCENT_LEFT
    ws.cell(row=row + 1, column=col + 1).fill = _FILL_KPI
    ws.cell(row=row + 1, column=col + 1).border = _BORDER_ALL


def _write_exec_table(ws, df, start_row: int, section_title: str, *, n_table_cols: int = 8) -> int:
    """Write a section header + styled table on the Executive Summary sheet.

    Returns the next free row below the table. Pure helper — references
    module-level constants and the two helpers above.
    """
    cols_list = list(df.columns)
    span = max(len(cols_list), n_table_cols)

    # Section header with left accent bar
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=span)
    hdr = ws.cell(row=start_row, column=1)
    hdr.value = f"  {section_title}"
    hdr.font = _FONT_SECTION
    hdr.fill = _FILL_SECTION
    hdr.alignment = _ALIGN_LEFT
    hdr.border = Border(left=Side(style="thick", color=_CLR_SECTION_BAR), bottom=_THIN)
    ws.row_dimensions[start_row].height = 28

    table_row = start_row + 1
    for ci, col_name in enumerate(cols_list, 1):
        c = ws.cell(row=table_row, column=ci)
        c.value = _COLUMN_LABELS.get(col_name, col_name)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_HEADER
    ws.row_dimensions[table_row].height = 28

    group_col_idx = cols_list.index("group") + 1 if "group" in cols_list else None
    for ri, (_, data_row) in enumerate(df.iterrows(), table_row + 1):
        ws.row_dimensions[ri].height = 20
        for ci, col_name in enumerate(cols_list, 1):
            cell = ws.cell(row=ri, column=ci)
            cell.value = data_row[col_name]
            cell.font = _FONT_DATA
            cell.border = _BORDER_ALL
            cell.alignment = _ALIGN_LEFT if col_name in _TEXT_COLS else _ALIGN_RIGHT
            _apply_number_format(cell, col_name)

        is_total = group_col_idx and str(ws.cell(row=ri, column=group_col_idx).value or "").strip().lower().startswith(
            "total"
        )
        for ci, col_name in enumerate(cols_list, 1):
            cell = ws.cell(row=ri, column=ci)
            if is_total:
                cell.fill = _FILL_TOTAL
                cell.font = _FONT_TOTAL
            elif ri % 2 == 0:
                cell.fill = _FILL_STRIPE
            if col_name in _DELTA_COLS and isinstance(cell.value, (int, float)):
                good = cell.value <= 0 if col_name == "risk_delta_pct" else cell.value >= 0
                clr = _CLR_GOOD if good else _CLR_BAD
                cell.font = Font(bold=bool(is_total), color=clr, size=10, name=_FN)

    for ci in range(1, len(cols_list) + 1):
        _set_col_width(ws, ci, ws.cell(row=table_row, column=ci).value, ws.max_row)
    return ws.max_row + 2


# =============================================================================
# RP-sheet writer + classification grid (R2b todo #57 step 4)
# =============================================================================
# Extracted from export_consolidated_excel body. Both take an openpyxl worksheet
# plus data primitives; no closure over segment-specific state. Use the
# module-level aliases _BarChart / _Reference / _get_column_letter in place of
# the in-function imports.


def _write_rp_sheet(
    writer,
    df_rp,
    sheet_name,
    seg_name,
    period_label,
    tab_color,
    extra_tables=None,
    classification_grid=None,
    is_mr=False,
):
    """Create a styled RP sheet with title banner, period label, data tables, and classification grid."""
    if extra_tables is None:
        extra_tables = []
    # Write primary table starting at row 4 (leaving room for banner)
    df_rp.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
    ws = writer.sheets[sheet_name]
    ws.sheet_properties.tabColor = tab_color
    ws.sheet_view.showGridLines = False
    n_cols = max(len(df_rp.columns), 6)

    # Title banner
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title = ws.cell(row=1, column=1)
    title.value = f"  {seg_name}"
    title.font = Font(bold=True, color=_CLR_WHITE, size=14, name=_FN)
    title.fill = PatternFill(start_color=_CLR_PRIMARY, end_color=_CLR_PRIMARY, fill_type="solid")
    title.alignment = _ALIGN_LEFT
    ws.row_dimensions[1].height = 32

    # Period subtitle
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    sub = ws.cell(row=2, column=1)
    sub.value = f"  {period_label}"
    if "MR" in period_label.upper():
        sub.font = _FONT_MR_LABEL
        sub.fill = _FILL_MR
    else:
        sub.font = Font(bold=False, color=_CLR_ACCENT, size=11, name=_FN)
        sub.fill = PatternFill(start_color=_CLR_ACCENT_LIGHT, end_color=_CLR_ACCENT_LIGHT, fill_type="solid")
    sub.alignment = _ALIGN_LEFT
    ws.row_dimensions[2].height = 24

    # Thin accent line
    for c in range(1, n_cols + 1):
        ws.cell(row=3, column=c).border = Border(top=Side(style="medium", color=_CLR_ACCENT))
    ws.row_dimensions[3].height = 6

    # Style the primary data table (header at row 4)
    _style_table(ws, df_rp.columns, header_row=4, highlight_total=False)

    # Optional additional tables (e.g., per-income-bin)
    next_row = ws.max_row + 2
    for tbl_title, tbl_df in extra_tables:
        if tbl_df is None or tbl_df.empty:
            continue
        n_tbl_cols = max(len(tbl_df.columns), 6)
        ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=n_tbl_cols)
        t = ws.cell(row=next_row, column=1)
        t.value = f"  {tbl_title}"
        t.font = _FONT_SECTION
        t.fill = _FILL_SECTION
        t.alignment = _ALIGN_LEFT
        t.border = _SECTION_LEFT
        ws.row_dimensions[next_row].height = 24

        # Header is one row below startrow passed to to_excel
        tbl_df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=next_row)
        _style_table(ws, tbl_df.columns, header_row=next_row + 1, highlight_total=False)
        next_row = ws.max_row + 2

    # Classification grid with charts (volume & risk by income bin)
    if classification_grid:
        next_row = _write_classification_grid(ws, classification_grid, next_row, is_mr=is_mr)

    _apply_page_setup(ws)


def _write_classification_grid(ws, grid_data, start_row, is_mr=False):
    """Write the classification-by-income-bin grid table and charts.

    Returns the next free row below all content.
    """
    if not grid_data:
        return start_row

    # Gather income bin labels (ordered)
    seen = {}
    for r in grid_data:
        if r["income_bin"] not in seen:
            seen[r["income_bin"]] = r["income_label"]
    ib_order = list(seen.keys())
    ib_labels = list(seen.values())
    n_ib = len(ib_order)

    categories = ["Keep", "Swap-in", "Swap-out", "Optimum"]
    has_risk = not is_mr

    # Columns per income bin: Volume, Risk (if main), Count
    cols_per_ib = 3 if has_risk else 2
    total_data_cols = n_ib * cols_per_ib + cols_per_ib  # + Total column group
    total_cols = 1 + total_data_cols  # Category column + data

    # ── Section header ──
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=total_cols)
    hdr = ws.cell(row=start_row, column=1)
    hdr.value = "  Volume & Risk by Income Bin" if has_risk else "  Volume by Income Bin"
    hdr.font = _FONT_SECTION
    hdr.fill = _FILL_SECTION
    hdr.alignment = _ALIGN_LEFT
    hdr.border = _SECTION_LEFT
    ws.row_dimensions[start_row].height = 28

    # ── Row 1: Income bin group headers (merged) ──
    r1 = start_row + 1
    ws.cell(row=r1, column=1).fill = _FILL_HEADER
    ws.cell(row=r1, column=1).border = _BORDER_HEADER
    col = 2
    for _i, lbl in enumerate(ib_labels + ["Total"]):
        end_col = col + cols_per_ib - 1
        ws.merge_cells(start_row=r1, start_column=col, end_row=r1, end_column=end_col)
        c = ws.cell(row=r1, column=col)
        c.value = lbl
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER_HEADER
        for cc in range(col, end_col + 1):
            ws.cell(row=r1, column=cc).fill = _FILL_HEADER
            ws.cell(row=r1, column=cc).border = _BORDER_HEADER
        col = end_col + 1
    ws.row_dimensions[r1].height = 24

    # ── Row 2: Sub-headers (Volume / Risk / Count) ──
    r2 = start_row + 2
    cat_hdr = ws.cell(row=r2, column=1)
    cat_hdr.value = "Category"
    cat_hdr.font = _FONT_HEADER
    cat_hdr.fill = PatternFill(start_color=_CLR_PRIMARY_LIGHT, end_color=_CLR_PRIMARY_LIGHT, fill_type="solid")
    cat_hdr.alignment = _ALIGN_CENTER
    cat_hdr.border = _BORDER_HEADER
    col = 2
    sub_headers = ["Volume (€)", "Risk (%)", "# Loans"] if has_risk else ["Volume (€)", "# Loans"]
    for _ in range(n_ib + 1):  # each income bin + Total
        for sh in sub_headers:
            c = ws.cell(row=r2, column=col)
            c.value = sh
            c.font = _FONT_HEADER
            c.fill = PatternFill(start_color=_CLR_PRIMARY_LIGHT, end_color=_CLR_PRIMARY_LIGHT, fill_type="solid")
            c.alignment = _ALIGN_CENTER
            c.border = _BORDER_HEADER
            col += 1
    ws.row_dimensions[r2].height = 22

    # ── Data rows ──
    # Build lookup: (category, income_bin) -> row data
    lookup = {(r["category"], r["income_bin"]): r for r in grid_data}
    data_start_row = start_row + 3
    for ci, cat in enumerate(categories):
        r = data_start_row + ci
        ws.row_dimensions[r].height = 22
        cat_cell = ws.cell(row=r, column=1)
        cat_cell.value = cat
        is_opt = cat == "Optimum"
        cat_cell.font = _FONT_OPTIMUM if is_opt else _FONT_DATA
        cat_cell.fill = _FILL_OPTIMUM if is_opt else (_FILL_STRIPE if ci % 2 == 1 else PatternFill())
        cat_cell.alignment = _ALIGN_LEFT
        cat_cell.border = _BORDER_ALL

        col = 2
        total_vol, total_cnt = 0.0, 0

        for ib in ib_order:
            d = lookup.get((cat, ib), {"volume": 0, "risk": None, "count": 0})
            vol = d["volume"]
            risk = d["risk"]
            cnt = d["count"]
            total_vol += vol
            total_cnt += cnt

            # For totals risk recalculation — accumulate raw numerator/denominator
            # We stored risk as percentage, but to get a proper weighted total we
            # need to re-derive from todu sums.  However we only stored the final
            # risk.  Instead we'll just use the lookup data that was computed with
            # the correct formula (it aggregates at the per-ib level).
            # We'll compute total risk separately below.

            # Volume cell
            vc = ws.cell(row=r, column=col)
            vc.value = vol
            vc.number_format = '#,##0 "€"'
            vc.font = _FONT_OPTIMUM if is_opt else _FONT_DATA
            vc.fill = _FILL_OPTIMUM if is_opt else (_FILL_STRIPE if ci % 2 == 1 else PatternFill())
            vc.alignment = _ALIGN_RIGHT
            vc.border = _BORDER_ALL
            col += 1

            if has_risk:
                rc = ws.cell(row=r, column=col)
                rc.value = risk if risk is not None else ""
                if isinstance(rc.value, float):
                    rc.number_format = "0.00"
                rc.font = _FONT_OPTIMUM if is_opt else _FONT_DATA
                rc.fill = _FILL_OPTIMUM if is_opt else (_FILL_STRIPE if ci % 2 == 1 else PatternFill())
                rc.alignment = _ALIGN_RIGHT
                rc.border = _BORDER_ALL
                col += 1

            cc = ws.cell(row=r, column=col)
            cc.value = cnt
            cc.number_format = "#,##0"
            cc.font = _FONT_OPTIMUM if is_opt else _FONT_DATA
            cc.fill = _FILL_OPTIMUM if is_opt else (_FILL_STRIPE if ci % 2 == 1 else PatternFill())
            cc.alignment = _ALIGN_RIGHT
            cc.border = _BORDER_ALL
            col += 1

        # Total column group
        vc = ws.cell(row=r, column=col)
        vc.value = total_vol
        vc.number_format = '#,##0 "€"'
        vc.font = _FONT_TOTAL
        vc.fill = _FILL_TOTAL
        vc.alignment = _ALIGN_RIGHT
        vc.border = _BORDER_ALL
        col += 1

        if has_risk:
            # Total risk: find all rows for this category across income bins
            # and sum todu values for proper weighted average
            total_risk = None
            cat_rows = [lookup.get((cat, ib), {}) for ib in ib_order]
            # We don't have raw todu here; use weighted average by volume as approximation
            weighted_sum = sum((cr.get("risk", 0) or 0) * (cr.get("volume", 0) or 0) for cr in cat_rows)
            if total_vol > 0 and any(cr.get("risk") is not None for cr in cat_rows):
                total_risk = weighted_sum / total_vol
            rc = ws.cell(row=r, column=col)
            rc.value = total_risk if total_risk is not None else ""
            if isinstance(rc.value, float):
                rc.number_format = "0.00"
            rc.font = _FONT_TOTAL
            rc.fill = _FILL_TOTAL
            rc.alignment = _ALIGN_RIGHT
            rc.border = _BORDER_ALL
            col += 1

        cc = ws.cell(row=r, column=col)
        cc.value = total_cnt
        cc.number_format = "#,##0"
        cc.font = _FONT_TOTAL
        cc.fill = _FILL_TOTAL
        cc.alignment = _ALIGN_RIGHT
        cc.border = _BORDER_ALL

    # Set column widths
    ws.column_dimensions[_get_column_letter(1)].width = 12
    for c in range(2, total_cols + 1):
        ws.column_dimensions[_get_column_letter(c)].width = 14

    chart_anchor_row = data_start_row + len(categories) + 2

    # Chart colours: Keep=teal, Swap-in=cerulean, Swap-out=warm red
    _CHART_COLORS = {"Keep": "1ABC9C", "Swap-in": "2980B9", "Swap-out": "E74C3C"}
    chart_cats = ["Keep", "Swap-in", "Swap-out"]

    # ── Chart data block ──
    # Transposed layout: income bins on X-axis, categories as series (legend)
    #
    #   col 1        | col 2  | col 3    | col 4
    #   Income Bin   | Keep   | Swap-in  | Swap-out      ← header (series titles)
    #   ≤ 2,000€     | vol    | vol      | vol           ← data rows
    #   > 2,000€     | vol    | vol      | vol
    #
    chart_data_row = chart_anchor_row
    n_cats = len(chart_cats)

    # Volume data block
    ws.cell(row=chart_data_row, column=1).value = "Income Bin"
    for ci, cat in enumerate(chart_cats):
        ws.cell(row=chart_data_row, column=2 + ci).value = cat
    for i, ib in enumerate(ib_order):
        r = chart_data_row + 1 + i
        ws.cell(row=r, column=1).value = ib_labels[i]
        for ci, cat in enumerate(chart_cats):
            d = lookup.get((cat, ib), {"volume": 0})
            ws.cell(row=r, column=2 + ci).value = d["volume"]

    # Hide chart data (tiny white-on-white text)
    vol_block_end = chart_data_row + n_ib
    for rr in range(chart_data_row, vol_block_end + 1):
        for cc in range(1, 2 + n_cats):
            ws.cell(row=rr, column=cc).font = Font(size=1, color="F8F9FA", name=_FN)

    # ── Volume bar chart ──
    from openpyxl.chart.label import DataLabelList

    chart = _BarChart()
    chart.type = "col"
    chart.grouping = "clustered"
    chart.title = "Production Volume by Income Bin"
    chart.y_axis.title = "Volume (€)"
    chart.y_axis.numFmt = "#,##0"
    chart.y_axis.delete = False
    chart.x_axis.delete = False
    chart.x_axis.tickLblPos = "low"
    chart.style = 10
    chart.width = 22
    chart.height = 14
    chart.legend.position = "b"
    chart.gapWidth = 80  # tighter bar groups

    data_ref = _Reference(ws, min_col=2, max_col=1 + n_cats, min_row=chart_data_row, max_row=vol_block_end)
    cats_ref = _Reference(ws, min_col=1, min_row=chart_data_row + 1, max_row=vol_block_end)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.shape = 4
    for si, series in enumerate(chart.series):
        cat_name = chart_cats[si] if si < n_cats else ""
        clr = _CHART_COLORS.get(cat_name, "AEB6BF")
        series.graphicalProperties.solidFill = clr
        series.graphicalProperties.line.solidFill = clr
        # Data labels showing series name on each bar
        series.dLbls = DataLabelList()
        series.dLbls.showSerName = True
        series.dLbls.showVal = False
        series.dLbls.showCatName = False

    vol_chart_row = vol_block_end + 2
    ws.add_chart(chart, f"A{vol_chart_row}")

    # ── Risk bar chart (main period only) ──
    if has_risk:
        risk_data_row = vol_block_end + 1
        ws.cell(row=risk_data_row, column=1).value = "Income Bin"
        for ci, cat in enumerate(chart_cats):
            ws.cell(row=risk_data_row, column=2 + ci).value = cat
        for i, ib in enumerate(ib_order):
            r = risk_data_row + 1 + i
            ws.cell(row=r, column=1).value = ib_labels[i]
            for ci, cat in enumerate(chart_cats):
                d = lookup.get((cat, ib), {"risk": None})
                ws.cell(row=r, column=2 + ci).value = d.get("risk") or 0
        risk_block_end = risk_data_row + n_ib
        for rr in range(risk_data_row, risk_block_end + 1):
            for cc in range(1, 2 + n_cats):
                ws.cell(row=rr, column=cc).font = Font(size=1, color="F8F9FA", name=_FN)

        risk_chart = _BarChart()
        risk_chart.type = "col"
        risk_chart.grouping = "clustered"
        risk_chart.title = "Risk (%) by Income Bin"
        risk_chart.y_axis.title = "Risk (%)"
        risk_chart.y_axis.numFmt = "0.00"
        risk_chart.y_axis.delete = False
        risk_chart.x_axis.delete = False
        risk_chart.x_axis.tickLblPos = "low"
        risk_chart.style = 10
        risk_chart.width = 22
        risk_chart.height = 14
        risk_chart.legend.position = "b"
        risk_chart.gapWidth = 80

        risk_data_ref = _Reference(ws, min_col=2, max_col=1 + n_cats, min_row=risk_data_row, max_row=risk_block_end)
        risk_cats_ref = _Reference(ws, min_col=1, min_row=risk_data_row + 1, max_row=risk_block_end)
        risk_chart.add_data(risk_data_ref, titles_from_data=True)
        risk_chart.set_categories(risk_cats_ref)
        risk_chart.shape = 4
        for si, series in enumerate(risk_chart.series):
            cat_name = chart_cats[si] if si < n_cats else ""
            clr = _CHART_COLORS.get(cat_name, "AEB6BF")
            series.graphicalProperties.solidFill = clr
            series.graphicalProperties.line.solidFill = clr
            # Data labels showing series name + value on each bar
            series.dLbls = DataLabelList()
            series.dLbls.showSerName = True
            series.dLbls.showVal = True
            series.dLbls.showCatName = False
            series.dLbls.numFmt = '0.00"%"'

        # Place risk chart to the right of volume chart
        risk_col_letter = _get_column_letter(total_cols + 2)
        ws.add_chart(risk_chart, f"{risk_col_letter}{vol_chart_row}")

    # Return next free row (below charts — each chart ~20 rows tall)
    return vol_chart_row + 20


# =============================================================================
# Small data-transform helpers (R2b todo #57 step 5)
# =============================================================================
# Pure/near-pure helpers extracted from export_consolidated_excel body. They
# reference only module-level utilities (_sort_consolidated_rows,
# _display_group_name) and pandas primitives.


def _append_summary_row_if_missing(df_tbl: pd.DataFrame) -> pd.DataFrame:
    if df_tbl.empty or "Metric" not in df_tbl.columns:
        return df_tbl
    if df_tbl["Metric"].astype(str).str.strip().str.lower().eq("summary").any():
        return df_tbl
    actual = df_tbl[df_tbl["Metric"] == "Actual"]
    optimum = df_tbl[df_tbl["Metric"] == "Optimum selected"]
    if actual.empty or optimum.empty:
        return df_tbl
    summary = dict.fromkeys(df_tbl.columns, np.nan)
    summary["Metric"] = "Summary"
    for c in ("Risk (%)", "Production (€)", "Production (%)"):
        if c in df_tbl.columns:
            a = actual.iloc[0].get(c)
            o = optimum.iloc[0].get(c)
            if pd.notna(a) and pd.notna(o):
                summary[c] = o - a
    for c in ("production_ci_lower", "production_ci_upper", "risk_ci_lower", "risk_ci_upper"):
        if c in df_tbl.columns:
            summary[c] = np.nan
    return pd.concat([df_tbl, pd.DataFrame([summary])], ignore_index=True)


def _prepare_export_df(source_df, cols):
    if source_df.empty:
        return source_df.copy()

    export_df = source_df.copy()
    if {"group", "period", "scenario"}.issubset(export_df.columns):
        export_df = _sort_consolidated_rows(export_df)
    export_df = export_df[[c for c in cols if c in export_df.columns]].copy()
    if "group" in export_df.columns:
        export_df["group"] = export_df["group"].map(_display_group_name)
    if "period" in export_df.columns:
        export_df["period"] = export_df["period"].astype(str).str.upper()
    if "scenario" in export_df.columns:
        export_df["scenario"] = export_df["scenario"].astype(str).str.title()
    return export_df


def _build_top_movers_df(source_df, *, period: str) -> pd.DataFrame:
    movers = source_df.copy()
    if movers.empty:
        return movers

    if "scenario" in movers.columns:
        movers = movers[movers["scenario"] == "base"]
    if "period" in movers.columns:
        movers = movers[movers["period"] == period]
    movers = movers[~(movers["group"].eq("TOTAL") | movers["group"].astype(str).str.startswith("supersegment_"))]
    if movers.empty:
        return movers

    sort_cols = [c for c in ["production_delta", "risk_delta_pct"] if c in movers.columns]
    ascending = [c != "production_delta" for c in sort_cols]
    if sort_cols:
        movers = movers.sort_values(sort_cols, ascending=ascending)
    movers = movers.head(8)
    movers = movers[
        [
            c
            for c in [
                "group",
                "optimum_production",
                "production_delta",
                "production_delta_pct",
                "optimum_risk_pct",
                "risk_delta_pct",
                "optimum_rejection_rate_pct",
            ]
            if c in movers.columns
        ]
    ].copy()
    movers["group"] = movers["group"].map(_display_group_name)
    return movers


# =============================================================================
# Config-loading helpers for the Excel exporter (R2b-ii todo #57 step 6)
# =============================================================================
# Previously nested inside export_consolidated_excel / _build_income_bin_tables,
# these read segment/global config files to derive multipliers, inv_vars,
# baseline_mode, and income-bin thresholds. Lifted to module level with
# explicit output_base / segments / supersegments parameters so the
# function body shrinks further and these are independently testable.


def _load_segment_settings(output_base: Path, seg_name_local: str) -> dict[str, Any]:
    """Load multiplier, multiplier_h3, inv_vars, baseline_mode from the segment's saved config."""
    defaults: dict[str, Any] = {
        "multiplier": float(DEFAULT_RISK_MULTIPLIER),
        "multiplier_h3": None,
        "inv_vars": None,
        "baseline_mode": False,
    }
    try:
        import tomllib

        seg_cfg_path = output_base / seg_name_local / "config_segment.toml"
        if seg_cfg_path.exists():
            cfg = tomllib.loads(seg_cfg_path.read_text(encoding="utf-8"))
            prep = cfg.get("preprocessing", cfg)
            if "multiplier" in prep:
                defaults["multiplier"] = float(prep["multiplier"])
            if "multiplier_h3" in prep:
                defaults["multiplier_h3"] = float(prep["multiplier_h3"])
            if "inv_vars" in prep:
                defaults["inv_vars"] = list(prep["inv_vars"])
            if "baseline_mode" in prep:
                defaults["baseline_mode"] = bool(prep["baseline_mode"])
    except Exception:
        logger.warning(
            f"_get_segment_defaults: failed to read segment config for '{seg_name_local}'; using global defaults",
            exc_info=True,
        )
    return defaults


def _extract_binary_income_threshold(bin_edges: Any) -> float | None:
    if not isinstance(bin_edges, list) or len(bin_edges) < 3:
        return None
    finite_edges = []
    for e in bin_edges:
        try:
            ef = float(e)
        except (TypeError, ValueError):
            continue
        if np.isfinite(ef):
            finite_edges.append(ef)
    if len(finite_edges) != 1:
        return None
    return float(finite_edges[0])


def _resolve_income_threshold(
    output_base: Path,
    segments: dict[str, dict[str, Any]],
    supersegments: dict[str, dict[str, Any]],
    seg_name_local: str,
) -> float | None:
    # 0) Segment run config snapshot (contains learned/injected edges in batch mode)
    try:
        import tomllib

        seg_cfg_path = output_base / seg_name_local / "config_segment.toml"
        if seg_cfg_path.exists():
            cfg = tomllib.loads(seg_cfg_path.read_text(encoding="utf-8"))
            prep = cfg.get("preprocessing", cfg)
            seg_file_edges = prep.get("bins", {}).get("income_bin", {}).get("bin_edges")
            th = _extract_binary_income_threshold(seg_file_edges)
            if th is not None:
                return th
    except Exception:
        logger.warning(
            f"_resolve_income_threshold: failed to read segment-file edges for "
            f"'{seg_name_local}'; falling back to segments.toml",
            exc_info=True,
        )

    seg_cfg = segments.get(seg_name_local, {})
    # 1) Segment-level bins override
    seg_edges = seg_cfg.get("bins", {}).get("income_bin", {}).get("bin_edges")
    th = _extract_binary_income_threshold(seg_edges)
    if th is not None:
        return th

    # 2) Reporting supersegment-level fixed bin edges
    reporting_ss = seg_cfg.get("reporting_supersegment") or seg_cfg.get("supersegment")
    if reporting_ss and reporting_ss in supersegments:
        ss_edges = supersegments[reporting_ss].get("bin_edges", {}).get("income_bin")
        th = _extract_binary_income_threshold(ss_edges)
        if th is not None:
            return th

    # 3) Global config fallback (config.toml)
    try:
        cfg_path = Path("config.toml")
        if cfg_path.exists():
            cfg = tomllib.loads(cfg_path.read_text(encoding="utf-8"))
            global_edges = cfg.get("preprocessing", {}).get("bins", {}).get("income_bin", {}).get("bin_edges")
            th = _extract_binary_income_threshold(global_edges)
            if th is not None:
                return th
    except Exception:
        logger.warning(
            f"_resolve_income_threshold: failed to read global config.toml edges for "
            f"'{seg_name_local}'; no threshold will be resolved",
            exc_info=True,
        )
    return None


# =============================================================================
# Small helpers extracted from income-bin tables (R2b-ii todo #57 step 7)
# =============================================================================


def _sort_pivot(pivot):
    """Sort pivot index and columns numerically where possible."""
    try:
        pivot = pivot.sort_index(key=lambda x: pd.to_numeric(x, errors="coerce"))
    except (TypeError, ValueError):
        pass
    try:
        pivot = pivot[
            sorted(
                pivot.columns,
                key=lambda x: float(x) if str(x).replace(".", "").replace("-", "").isdigit() else x,
            )
        ]
    except (TypeError, ValueError):
        pass
    return pivot


def _audit_slice_for_income_bin(audit_df: pd.DataFrame, income_val: Any) -> pd.DataFrame:
    if "income_bin" not in audit_df.columns:
        return pd.DataFrame()
    ib = pd.to_numeric(audit_df["income_bin"], errors="coerce")
    try:
        target = float(income_val)
    except (TypeError, ValueError):
        return pd.DataFrame()
    return audit_df.loc[ib == target].copy()


def _empty_rp_skeleton(template_cols: list[str]) -> pd.DataFrame:
    base = dict.fromkeys(template_cols, np.nan)
    rows = []
    for m in ["Actual", "Swap-in", "Swap-out", "Optimum selected", "Summary"]:
        r = dict(base)
        r["Metric"] = m
        rows.append(r)
    return pd.DataFrame(rows)


def export_consolidated_excel(
    consolidated_df: pd.DataFrame,
    output_base: str | Path,
    segments: dict[str, dict[str, Any]],
    supersegments: dict[str, dict[str, Any]] | None = None,
) -> Path:
    """Export consolidated report data to a management-ready Excel dashboard.

    Sheets:
        - Executive Summary: KPI cards (main + MR), curated tables, acceptance grids
        - Portfolio Summary: TOTAL + supersegment rows (all scenarios/periods)
        - Segment Detail: per-segment rows
        - Cutoff Comparison: concatenated cutoff data
        - Per-segment acceptance grids (one sheet each)
        - Per-segment risk production summaries

    Args:
        consolidated_df: DataFrame from consolidate_segments()
        output_base: Base output directory (where segment subdirectories live)
        segments: Segment configurations dict
        supersegments: Reporting supersegment configurations (optional)

    Returns:
        Path to the written .xlsx file.
    """
    from datetime import date

    from openpyxl.utils import get_column_letter

    # openpyxl.styles (Alignment, Border, Font, PatternFill, Side) are
    # imported at module level — used by the design-token constants above.

    output_base = Path(output_base)
    supersegments = supersegments or {}
    xlsx_path = output_base / "consolidated_risk_production.xlsx"

    # Design tokens and column classifications are module-level constants
    # (R2b todo #57 — extracted to slim down this 2000-line function).
    # Nested helpers below reference them via lexical scoping.

    # Helpers
    # =====================================================================

    # _set_col_width, _apply_number_format, _apply_page_setup are now
    # module-level helpers (R2b todo #57 step 2).

    # _style_table, _write_kpi_card, _write_exec_table are now module-level
    # helpers (R2b todo #57 step 3).

    # _write_rp_sheet and _write_classification_grid are now module-level
    # helpers (R2b todo #57 step 4).

    # _load_segment_settings is now a module-level helper (R2b-ii todo #57 step 6).

    def _build_income_bin_tables(
        seg_name: str, period: str, template_cols: list[str]
    ) -> list[tuple[str, pd.DataFrame]]:
        data_dir = output_base / seg_name / "data"
        accepted_cells_path = data_dir / "accepted_cells_base.csv"
        if not accepted_cells_path.exists():
            return []
        try:
            variables = list(pd.read_csv(accepted_cells_path, nrows=1).columns)
        except (pd.errors.ParserError, OSError, ValueError):
            return []
        if not variables:
            return []

        if period == "mr":
            ds_path = data_dir / "data_summary_desagregado_mr_base.csv"
            if not ds_path.exists():
                ds_path = data_dir / "data_summary_desagregado_mr.csv"
            # Use MR-specific optimal solution (has re-optimized mask after recalibration);
            # fall back to main-period if MR-specific doesn't exist (pre-fix runs).
            opt_path = data_dir / "optimal_solution_mr_base.csv"
            if not opt_path.exists():
                opt_path = data_dir / "optimal_solution_mr.csv"
            if not opt_path.exists():
                opt_path = data_dir / "optimal_solution_base.csv"
        else:
            ds_path = data_dir / "data_summary_desagregado_base.csv"
            opt_path = data_dir / "optimal_solution_base.csv"
        if not ds_path.exists() or not opt_path.exists():
            return []
        try:
            df_sum = pd.read_csv(ds_path)
            df_opt = pd.read_csv(opt_path)
        except (pd.errors.ParserError, OSError, ValueError):
            return []
        if df_sum.empty or df_opt.empty or "income_bin" not in df_sum.columns:
            return []

        mask = None
        grid = None
        if "acceptance_mask" in df_opt.columns and pd.notna(df_opt.iloc[0].get("acceptance_mask")):
            try:
                from src.optimization_utils import CellGrid, decode_mask

                grid = CellGrid.from_summary(df_sum, variables)
                mask = decode_mask(str(df_opt.iloc[0]["acceptance_mask"]))
                if len(mask) != len(grid.cell_data):
                    logger.warning(
                        f"[{seg_name}] acceptance_mask length ({len(mask)}) does not match grid cells ({len(grid.cell_data)})"
                    )
                    mask = None
                    grid = None
            except Exception as e:
                logger.warning(f"[{seg_name}] Could not decode acceptance_mask for income-bin RP tables: {e}")
                mask = None
                grid = None

        from src.mr_pipeline import calculate_metrics_from_cuts

        seg_settings = _load_segment_settings(output_base, seg_name)

        # _extract_binary_income_threshold and _resolve_income_threshold are
        # now module-level helpers (R2b-ii todo #57 step 6).
        income_threshold = _resolve_income_threshold(output_base, segments, supersegments, seg_name)

        def _income_bin_title(income_val: Any) -> str:
            """Portfolio-owner labels based on configured income_bin threshold."""
            try:
                iv = int(float(income_val))
            except (TypeError, ValueError):
                return f"Income Bin {income_val}"
            if iv == 1:
                if income_threshold is not None:
                    return f"income_bin <= {income_threshold:,.0f}€"
                return "Income Bin 1"
            if iv == 2:
                if income_threshold is not None:
                    return f"income_bin > {income_threshold:,.0f}€"
                return "Income Bin 2"
            return f"Income Bin {income_val}"

        audit_path = data_dir / ("audit_base_mr.csv" if period == "mr" else "audit_base.csv")
        audit_full: pd.DataFrame | None = None
        if audit_path.exists():
            try:
                audit_full = pd.read_csv(audit_path, skipinitialspace=True)
                audit_full.columns = audit_full.columns.str.strip()
            except (pd.errors.ParserError, OSError, ValueError):
                audit_full = None

        # _audit_slice_for_income_bin and _empty_rp_skeleton are now
        # module-level helpers (R2b-ii todo #57 step 7).

        out_tables: list[tuple[str, pd.DataFrame]] = []
        # Keep deterministic order with 1 first, then 2.
        income_values = sorted(pd.Series(df_sum["income_bin"]).dropna().unique().tolist(), key=lambda x: float(x))
        ib_grid = pd.to_numeric(df_sum["income_bin"], errors="coerce")
        for income_val in income_values:
            try:
                iv = float(income_val)
            except (TypeError, ValueError):
                continue
            df_bin = df_sum.loc[ib_grid == iv].copy()
            if df_bin.empty:
                continue
            tbl = calculate_metrics_from_cuts(
                data_summary_desagregado=df_bin,
                optimal_solution_df=df_opt,
                variables=variables,
                inv_vars=seg_settings["inv_vars"],
                mask=mask,
                grid=grid,
                multiplier=seg_settings["multiplier"],
                multiplier_h3=seg_settings["multiplier_h3"],
            )
            if tbl is None or tbl.empty:
                continue
            tbl = _append_summary_row_if_missing(tbl)
            # Match main RP sheet: production € from loan-level audit for this income_bin.
            # Skip in baseline mode — the accept-all mask makes the audit classify all
            # rejected applicants as swap-in, which would overwrite the correct zero values.
            if audit_full is not None and not audit_full.empty and not seg_settings["baseline_mode"]:
                audit_bin = _audit_slice_for_income_bin(audit_full, income_val)
                if not audit_bin.empty:
                    tbl = reconcile_risk_production_summary_with_audit(tbl, audit_bin, silent=True)
            # Match the same shape/columns as the total RP table in Excel.
            for c in template_cols:
                if c not in tbl.columns:
                    tbl[c] = np.nan
            tbl = tbl[template_cols]
            out_tables.append((_income_bin_title(income_val), tbl))

        # Rows with missing or out-of-grid income_bin (audit-only; closes sum vs main RP)
        if audit_full is not None and "income_bin" in audit_full.columns and income_values:
            iba = pd.to_numeric(audit_full["income_bin"], errors="coerce")
            targets = {float(x) for x in income_values}
            uncovered = iba.isna() | (~iba.isin(list(targets)))
            if uncovered.any():
                au_rest = audit_full.loc[uncovered].copy()
                if not au_rest.empty:
                    tbl_u = _empty_rp_skeleton(template_cols)
                    tbl_u = reconcile_risk_production_summary_with_audit(tbl_u, au_rest, silent=True)
                    for c in template_cols:
                        if c not in tbl_u.columns:
                            tbl_u[c] = np.nan
                    tbl_u = tbl_u[template_cols]
                    out_tables.append(("Income bin — unassigned / outside grid bins", tbl_u))

        return out_tables

    def _build_classification_grid(
        seg_name: str,
        period: str,
        variables: list[str],
        multiplier: float,
        multiplier_h3: float | None,
        inv_vars: list[str] | None,
        income_threshold: float | None,
    ) -> list[dict] | None:
        """Build volume/risk breakdown by income_bin and classification (keep/swap-in/swap-out).

        Uses audit data for volumes and counts (matching the RP summary tables which are
        reconciled against the audit), and desagregado data for risk calculations.

        Returns a list of dicts, one per row, with keys:
            category, income_bin, income_label, volume, risk, count
        or None if data is unavailable.
        """
        data_dir = output_base / seg_name / "data"
        is_mr = period == "mr"

        # ── Load audit file (source of truth for volumes/counts) ──
        audit_path = data_dir / ("audit_base_mr.csv" if is_mr else "audit_base.csv")
        if not audit_path.exists():
            return None
        try:
            audit = pd.read_csv(audit_path, skipinitialspace=True)
            audit.columns = audit.columns.str.strip()
        except (pd.errors.ParserError, OSError, ValueError):
            return None
        if audit.empty or "classification" not in audit.columns or "income_bin" not in audit.columns:
            return None

        amt_col = "oa_amt_adjusted" if "oa_amt_adjusted" in audit.columns else "oa_amt_h0"
        if amt_col not in audit.columns:
            return None

        # ── Load desagregado for risk (grid-level todu values) ──
        if is_mr:
            ds_path = data_dir / "data_summary_desagregado_mr_base.csv"
            if not ds_path.exists():
                ds_path = data_dir / "data_summary_desagregado_mr.csv"
            opt_path = data_dir / "optimal_solution_mr_base.csv"
            if not opt_path.exists():
                opt_path = data_dir / "optimal_solution_mr.csv"
            if not opt_path.exists():
                opt_path = data_dir / "optimal_solution_base.csv"
        else:
            ds_path = data_dir / "data_summary_desagregado_base.csv"
            opt_path = data_dir / "optimal_solution_base.csv"

        # Risk from desagregado (optional — volumes work without it)
        df_sum = None
        if not is_mr and ds_path.exists() and opt_path.exists():
            try:
                df_sum = pd.read_csv(ds_path)
                df_opt = pd.read_csv(opt_path)
                if df_sum.empty or df_opt.empty or "income_bin" not in df_sum.columns:
                    df_sum = None
                else:
                    # Determine passes_cut
                    _mask = None
                    _grid = None
                    if "acceptance_mask" in df_opt.columns and pd.notna(df_opt.iloc[0].get("acceptance_mask")):
                        try:
                            from src.optimization_utils import CellGrid, decode_mask

                            _grid = CellGrid.from_summary(df_sum, variables)
                            _mask = decode_mask(str(df_opt.iloc[0]["acceptance_mask"]))
                            if len(_mask) != len(_grid.cell_data):
                                _mask = None
                                _grid = None
                        except Exception:
                            logger.warning(
                                "Failed to decode acceptance_mask or construct CellGrid; "
                                "falling back to legacy 2-var cut_map path. Downstream audit "
                                "tables may use a different cutoff interpretation.",
                                exc_info=True,
                            )
                            _mask = None
                            _grid = None
                    if _mask is not None and _grid is not None:
                        from src.optimization_utils import classify_by_mask

                        df_sum["passes_cut"] = classify_by_mask(df_sum, _mask, _grid)
                    else:
                        var0 = variables[0]
                        var1 = variables[1] if len(variables) > 1 else None
                        if var1 is not None:
                            opt_row = df_opt.iloc[0]
                            cut_map = {}
                            for bv in sorted(df_sum[var0].unique()):
                                for key in [bv, str(bv), str(float(bv))]:
                                    if key in df_opt.columns:
                                        cut_map[bv] = opt_row[key]
                                        break
                                else:
                                    cut_map[bv] = np.inf if (inv_vars and var1 in inv_vars) else -np.inf
                            df_sum["cut_limit"] = df_sum[var0].map(cut_map)
                            if inv_vars and var1 in inv_vars:
                                df_sum["passes_cut"] = df_sum[var1] >= df_sum["cut_limit"]
                            else:
                                df_sum["passes_cut"] = df_sum[var1] <= df_sum["cut_limit"]
                        else:
                            df_sum = None
            except Exception:
                logger.warning(
                    "Failed to construct audit summary (passes_cut / cut_map); "
                    "audit tab for this segment/scenario will be skipped.",
                    exc_info=True,
                )
                df_sum = None

        # ── Income bin labels ──
        ib_vals = sorted(pd.to_numeric(audit["income_bin"], errors="coerce").dropna().unique())
        if not ib_vals:
            return None

        def _ib_label(iv: float) -> str:
            try:
                iv_int = int(iv)
            except (TypeError, ValueError):
                return f"Bin {iv}"
            if income_threshold is not None:
                if iv_int == 1:
                    return f"≤ {income_threshold:,.0f}€"
                if iv_int == 2:
                    return f"> {income_threshold:,.0f}€"
            return f"Bin {iv_int}"

        # ── Compute risk per (income_bin, category) from desagregado ──
        risk_lookup: dict[tuple[float, str], float | None] = {}
        if df_sum is not None and "passes_cut" in df_sum.columns:
            ib_col_ds = pd.to_numeric(df_sum["income_bin"], errors="coerce")
            for iv in ib_vals:
                sub = df_sum.loc[ib_col_ds == iv]
                for cat, filt, suffix in [
                    ("Keep", sub[sub["passes_cut"]], "_boo"),
                    ("Swap-out", sub[~sub["passes_cut"]], "_boo"),
                    ("Swap-in", sub[sub["passes_cut"]], "_rep"),
                ]:
                    rn_col = f"todu_30ever_h6{suffix}"
                    rd_col = f"todu_amt_pile_h6{suffix}"
                    rn = filt[rn_col].sum() if rn_col in filt.columns else 0
                    rd = filt[rd_col].sum() if rd_col in filt.columns else 0
                    risk = None
                    if rd > 0:
                        r_raw = calculate_b2_ever_h6(rn, rd, multiplier=multiplier, as_percentage=True, decimals=4)
                        risk = float(r_raw) if pd.notna(r_raw) else None
                    risk_lookup[(iv, cat)] = risk
                # Optimum risk = combined kept_boo + swap_in_rep
                sub_pass = sub[sub["passes_cut"]]
                o_rn = (sub_pass["todu_30ever_h6_boo"].sum() if "todu_30ever_h6_boo" in sub_pass.columns else 0) + (
                    sub_pass["todu_30ever_h6_rep"].sum() if "todu_30ever_h6_rep" in sub_pass.columns else 0
                )
                o_rd = (sub_pass["todu_amt_pile_h6_boo"].sum() if "todu_amt_pile_h6_boo" in sub_pass.columns else 0) + (
                    sub_pass["todu_amt_pile_h6_rep"].sum() if "todu_amt_pile_h6_rep" in sub_pass.columns else 0
                )
                opt_risk = None
                if o_rd > 0:
                    or_raw = calculate_b2_ever_h6(o_rn, o_rd, multiplier=multiplier, as_percentage=True, decimals=4)
                    opt_risk = float(or_raw) if pd.notna(or_raw) else None
                risk_lookup[(iv, "Optimum")] = opt_risk

        # ── Build grid rows from audit (volumes/counts) + risk lookup ──
        cls_col = audit["classification"].astype(str).str.strip()
        ib_col_a = pd.to_numeric(audit["income_bin"], errors="coerce")
        # Map audit classification names to grid categories
        cat_map = {"keep": "Keep", "swap_in": "Swap-in", "swap_out": "Swap-out"}

        rows: list[dict] = []
        for iv in ib_vals:
            label = _ib_label(iv)
            ib_mask = ib_col_a == iv
            for audit_cls, grid_cat in cat_map.items():
                cls_mask = cls_col == audit_cls
                subset = audit.loc[ib_mask & cls_mask]
                vol = float(subset[amt_col].sum()) if not subset.empty else 0.0
                cnt = len(subset)
                risk = risk_lookup.get((iv, grid_cat))
                rows.append(
                    {
                        "category": grid_cat,
                        "income_bin": iv,
                        "income_label": label,
                        "volume": vol,
                        "risk": risk,
                        "count": cnt,
                    }
                )
            # Optimum = Keep + Swap-in
            keep_r = next(r for r in rows if r["category"] == "Keep" and r["income_bin"] == iv)
            si_r = next(r for r in rows if r["category"] == "Swap-in" and r["income_bin"] == iv)
            rows.append(
                {
                    "category": "Optimum",
                    "income_bin": iv,
                    "income_label": label,
                    "volume": keep_r["volume"] + si_r["volume"],
                    "risk": risk_lookup.get((iv, "Optimum")),
                    "count": keep_r["count"] + si_r["count"],
                }
            )
        return rows if rows else None

    def _write_single_pivot_grid(ws, pivot, col_var, row_var, start_row, col_offset=0):
        """Draw one pivot grid at (start_row, col_offset+1). Returns bottom row used."""
        c0 = col_offset + 1

        # Corner label
        corner = ws.cell(row=start_row, column=c0)
        corner.value = f"{row_var} \\ {col_var}"
        corner.font = _FONT_GRID_LABEL
        corner.fill = _FILL_GRID_HEADER
        corner.alignment = _ALIGN_CENTER
        corner.border = _BORDER_GRID
        ws.column_dimensions[get_column_letter(c0)].width = max(
            ws.column_dimensions[get_column_letter(c0)].width or 8, len(str(corner.value)) + 4
        )
        ws.row_dimensions[start_row].height = 26

        # Column headers
        for ci, col_val in enumerate(pivot.columns, c0 + 1):
            c = ws.cell(row=start_row, column=ci)
            c.value = col_val
            c.font = _FONT_GRID_HEADER
            c.fill = _FILL_GRID_HEADER
            c.alignment = _ALIGN_CENTER
            c.border = _BORDER_GRID
            ws.column_dimensions[get_column_letter(ci)].width = max(
                ws.column_dimensions[get_column_letter(ci)].width or 0, 7
            )

        # Data rows
        for ri, idx_val in enumerate(pivot.index, start_row + 1):
            rh = ws.cell(row=ri, column=c0)
            rh.value = idx_val
            rh.font = _FONT_GRID_HEADER
            rh.fill = _FILL_GRID_HEADER
            rh.alignment = _ALIGN_CENTER
            rh.border = _BORDER_GRID
            ws.row_dimensions[ri].height = 26

            for ci, col_val in enumerate(pivot.columns, c0 + 1):
                cell = ws.cell(row=ri, column=ci)
                val = pivot.loc[idx_val, col_val]
                if pd.isna(val):
                    cell.fill = _FILL_NA
                    cell.value = "—"
                    cell.font = Font(bold=False, color=_CLR_NEUTRAL_MID, size=10, name=_FN)
                elif val == 1:
                    cell.fill = _FILL_ACCEPT
                    cell.value = "A"
                    cell.font = _FONT_GRID_CELL
                else:
                    cell.fill = _FILL_REJECT
                    cell.value = "R"
                    cell.font = _FONT_GRID_CELL
                cell.alignment = _ALIGN_CENTER
                cell.border = _BORDER_GRID

        return start_row + len(pivot.index)

    def _write_acceptance_strip_1d(ws, cutoff_df, seg_name, start_row, scenario="base"):
        """Draw a horizontal 1D acceptance strip on *ws* starting at *start_row*.

        For single-variable optimization, shows each bin as a colored cell (green=A, red=R)
        in a horizontal row, with a summary header and legend.
        Returns next free row.
        """
        variable_cols = [c for c in cutoff_df.columns if c not in _CUTOFF_FIXED_COLS]
        if not variable_cols:
            return start_row
        var_col = variable_cols[0]

        scen_df = cutoff_df
        if "scenario" in cutoff_df.columns:
            scen_df = cutoff_df[cutoff_df["scenario"] == scenario]
        if scen_df.empty:
            return start_row

        scen_sorted = scen_df.sort_values(var_col)
        bins = scen_sorted[var_col].values
        accepted = scen_sorted["accepted"].values
        n_accepted = int(accepted.sum())
        n_total = len(accepted)

        # Section header
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max(len(bins) + 1, 12))
        lbl = ws.cell(row=start_row, column=1)
        lbl.value = f"  {seg_name}  |  {scenario.title()}  —  {n_accepted}/{n_total} bins accepted ({100 * n_accepted / n_total:.0f}%)"
        lbl.font = _FONT_SECTION
        lbl.fill = _FILL_SECTION
        lbl.alignment = _ALIGN_LEFT
        lbl.border = Border(left=Side(style="thick", color=_CLR_SECTION_BAR), bottom=_HAIR)
        ws.row_dimensions[start_row].height = 28
        start_row += 1

        # Variable name label
        label_cell = ws.cell(row=start_row, column=1)
        label_cell.value = var_col
        label_cell.font = _FONT_GRID_LABEL
        label_cell.fill = _FILL_GRID_HEADER
        label_cell.alignment = _ALIGN_CENTER
        label_cell.border = _BORDER_GRID
        ws.column_dimensions[get_column_letter(1)].width = max(
            ws.column_dimensions[get_column_letter(1)].width or 8, len(var_col) + 4
        )

        # Bin header row
        for ci, bv in enumerate(bins, 2):
            c = ws.cell(row=start_row, column=ci)
            c.value = int(bv) if isinstance(bv, float) and bv == int(bv) else bv
            c.font = _FONT_GRID_HEADER
            c.fill = _FILL_GRID_HEADER
            c.alignment = _ALIGN_CENTER
            c.border = _BORDER_GRID
            ws.column_dimensions[get_column_letter(ci)].width = max(
                ws.column_dimensions[get_column_letter(ci)].width or 0, 7
            )
        ws.row_dimensions[start_row].height = 26
        start_row += 1

        # Status label
        status_cell = ws.cell(row=start_row, column=1)
        status_cell.value = "Status"
        status_cell.font = _FONT_GRID_LABEL
        status_cell.fill = _FILL_GRID_HEADER
        status_cell.alignment = _ALIGN_CENTER
        status_cell.border = _BORDER_GRID

        # Acceptance cells
        for ci, acc in enumerate(accepted, 2):
            cell = ws.cell(row=start_row, column=ci)
            if pd.isna(acc):
                cell.fill = _FILL_NA
                cell.value = "—"
                cell.font = Font(bold=False, color=_CLR_NEUTRAL_MID, size=11, name=_FN)
            elif acc == 1:
                cell.fill = _FILL_ACCEPT
                cell.value = "A"
                cell.font = _FONT_GRID_CELL
            else:
                cell.fill = _FILL_REJECT
                cell.value = "R"
                cell.font = _FONT_GRID_CELL
            cell.alignment = _ALIGN_CENTER
            cell.border = _BORDER_GRID
        ws.row_dimensions[start_row].height = 30
        start_row += 1

        # Legend
        start_row += 1
        legend_items = [
            ("  A  Accept  ", _FILL_ACCEPT, _CLR_WHITE),
            ("  R  Reject  ", _FILL_REJECT, _CLR_WHITE),
            ("  —  N/A  ", _FILL_NA, _CLR_TEXT),
        ]
        for ci, (label, fill, fg) in enumerate(legend_items, 1):
            c = ws.cell(row=start_row, column=ci)
            c.value = label
            c.font = Font(bold=True, color=fg, size=9, name=_FN)
            c.fill = fill
            c.alignment = _ALIGN_CENTER
            c.border = _BORDER_GRID

        return start_row + 2

    def _write_acceptance_grid(ws, cutoff_df, seg_name, start_row, scenario="base"):
        """Draw coloured acceptance/rejection grids on *ws* starting at *start_row*.

        For N>2 variables (e.g. octroi_bin x efx_bin x income_bin), creates one
        sub-grid per slice of the 3rd+ variables, laid out side by side.
        For 1D, draws a horizontal acceptance strip.
        Returns next free row.
        """
        variable_cols = [c for c in cutoff_df.columns if c not in _CUTOFF_FIXED_COLS]
        if "accepted" not in cutoff_df.columns or len(variable_cols) < 1:
            return start_row

        # 1D: delegate to strip renderer
        if len(variable_cols) == 1:
            return _write_acceptance_strip_1d(ws, cutoff_df, seg_name, start_row, scenario)

        scen_df = cutoff_df
        if "scenario" in cutoff_df.columns:
            scen_df = cutoff_df[cutoff_df["scenario"] == scenario]
        if scen_df.empty:
            return start_row

        col_var = variable_cols[0]  # columns of each pivot
        row_var = variable_cols[1]  # rows of each pivot
        slice_vars = variable_cols[2:]  # additional dimensions (e.g. income_bin)

        # Section label with left accent bar
        slice_info = f"  —  sliced by {', '.join(slice_vars)}" if slice_vars else ""
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=12)
        lbl = ws.cell(row=start_row, column=1)
        lbl.value = f"  {seg_name}  |  {scenario.title()}{slice_info}"
        lbl.font = _FONT_SECTION
        lbl.fill = _FILL_SECTION
        lbl.alignment = _ALIGN_LEFT
        lbl.border = Border(left=Side(style="thick", color=_CLR_SECTION_BAR), bottom=_HAIR)
        ws.row_dimensions[start_row].height = 28
        start_row += 1

        # Build list of (label, slice_df) pairs
        if slice_vars:
            groups = scen_df.groupby(slice_vars, sort=True)
            group_items = list(groups)[:12]  # cap at 12 sub-grids
        else:
            group_items = [(None, scen_df)]

        # Lay out sub-grids side by side (horizontally)
        col_offset = 0
        grid_bottom = start_row
        for slice_key, slice_df in group_items:
            try:
                pivot = slice_df.pivot_table(index=row_var, columns=col_var, values="accepted", aggfunc="first")
                pivot = _sort_pivot(pivot)
            except (TypeError, ValueError, KeyError):
                continue
            if pivot.empty:
                continue

            # Slice label above the sub-grid
            if slice_key is not None:
                if isinstance(slice_key, tuple):
                    label = ", ".join(f"{v}={k}" for v, k in zip(slice_vars, slice_key))
                else:
                    label = f"{slice_vars[0]}={slice_key}"
                lbl_cell = ws.cell(row=start_row, column=col_offset + 1)
                lbl_cell.value = label
                lbl_cell.font = Font(bold=True, color=_CLR_ACCENT, size=9, name=_FN)
                lbl_cell.alignment = _ALIGN_LEFT
                grid_start = start_row + 1
            else:
                grid_start = start_row

            bottom = _write_single_pivot_grid(ws, pivot, col_var, row_var, grid_start, col_offset)
            grid_bottom = max(grid_bottom, bottom)

            # Advance column offset for next sub-grid (grid width + 1 gap column)
            col_offset += len(pivot.columns) + 2  # +1 for row header, +1 gap

        # Legend row
        legend_row = grid_bottom + 2
        legend_items = [
            ("  A  Accept  ", _FILL_ACCEPT, _CLR_WHITE),
            ("  R  Reject  ", _FILL_REJECT, _CLR_WHITE),
            ("  —  N/A  ", _FILL_NA, _CLR_TEXT),
        ]
        for ci, (label, fill, fg) in enumerate(legend_items, 1):
            c = ws.cell(row=legend_row, column=ci)
            c.value = label
            c.font = Font(bold=True, color=fg, size=9, name=_FN)
            c.fill = fill
            c.alignment = _ALIGN_CENTER
            c.border = _BORDER_GRID

        return legend_row + 2

    def _get_total_row(period: str, scenario: str = "base"):
        mask = (
            (consolidated_df["group"] == "TOTAL")
            & (consolidated_df["period"] == period)
            & (consolidated_df["scenario"] == scenario)
        )
        rows = consolidated_df[mask]
        if not rows.empty:
            return rows.iloc[0]
        fallback = consolidated_df[(consolidated_df["group"] == "TOTAL") & (consolidated_df["period"] == period)]
        return fallback.iloc[0] if not fallback.empty else None

    # =====================================================================
    # Build workbook
    # =====================================================================
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        if consolidated_df.empty:
            consolidated_df.to_excel(writer, sheet_name="Executive Summary", index=False)
            logger.debug("Empty consolidated DataFrame — wrote empty workbook")
            return xlsx_path

        wb = writer.book

        # =============================================================
        # Sheet 1: Executive Summary
        # =============================================================
        ws_exec = wb.create_sheet("Executive Summary", 0)
        ws_exec.sheet_properties.tabColor = _CLR_TAB_EXEC
        ws_exec.sheet_view.showGridLines = False

        # --- Title banner with coloured bar ---
        title_fill = PatternFill(start_color=_CLR_PRIMARY, end_color=_CLR_PRIMARY, fill_type="solid")
        ws_exec.merge_cells("A1:J1")
        title_cell = ws_exec["A1"]
        title_cell.value = "  Consolidated Risk & Production Dashboard"
        title_cell.font = Font(bold=True, color=_CLR_WHITE, size=20, name=_FN)
        title_cell.fill = title_fill
        title_cell.alignment = _ALIGN_LEFT
        ws_exec.row_dimensions[1].height = 44

        ws_exec.merge_cells("A2:J2")
        ws_exec["A2"].value = (
            f"  Generated {date.today().strftime('%d %b %Y')}  |  {len(segments)} segment(s)  |  "
            "Consolidated portfolio view"
        )
        ws_exec["A2"].font = Font(bold=False, color=_CLR_ACCENT_LIGHT, size=11, name=_FN)
        ws_exec["A2"].fill = title_fill
        ws_exec["A2"].alignment = _ALIGN_LEFT
        ws_exec.row_dimensions[2].height = 22

        # Thin accent line under title bar
        for c in range(1, 11):
            ws_exec.cell(row=3, column=c).border = Border(top=Side(style="medium", color=_CLR_ACCENT))
        ws_exec.row_dimensions[3].height = 6

        # --- MAIN PERIOD KPI cards (rows 4-5) ---
        tr_main = _get_total_row("main")
        kpi_row = 4
        ws_exec.row_dimensions[kpi_row].height = 38
        ws_exec.row_dimensions[kpi_row + 1].height = 22

        if tr_main is not None:
            _write_kpi_card(ws_exec, kpi_row, 1, "Optimum Production", f"€{tr_main.get('optimum_production', 0):,.0f}")
            pd_val = tr_main.get("production_delta", 0)
            _write_kpi_card(
                ws_exec,
                kpi_row,
                3,
                "Production Delta",
                f"€{pd_val:+,.0f}",
                delta_str=f"{tr_main.get('production_delta_pct', 0):+.1f}%",
                delta_positive=pd_val >= 0,
            )
            rd = tr_main.get("risk_delta_pct", 0)
            _write_kpi_card(
                ws_exec,
                kpi_row,
                5,
                "Optimum Risk",
                f"{tr_main.get('optimum_risk_pct', 0):.2f}%",
                delta_str=f"{rd:+.2f} pp",
                delta_positive=rd <= 0,
            )
            _write_kpi_card(
                ws_exec, kpi_row, 7, "Rejection Rate", f"{tr_main.get('optimum_rejection_rate_pct', 0):.1f}%"
            )
            # Label
            ws_exec.cell(row=kpi_row, column=9).value = "MAIN PERIOD"
            ws_exec.cell(row=kpi_row, column=9).font = Font(bold=True, color=_CLR_ACCENT, size=9, name=_FN)
            ws_exec.cell(row=kpi_row, column=9).alignment = _ALIGN_LEFT
        else:
            ws_exec.cell(row=kpi_row, column=1).value = "No TOTAL/base/main data"
            ws_exec.cell(row=kpi_row, column=1).font = _FONT_SUBTITLE

        # --- MR PERIOD KPI cards (rows 6-7) ---
        tr_mr = _get_total_row("mr")
        mr_row = 6
        ws_exec.row_dimensions[mr_row].height = 38
        ws_exec.row_dimensions[mr_row + 1].height = 22

        if tr_mr is not None:
            _write_kpi_card(ws_exec, mr_row, 1, "MR Optimum Prod.", f"€{tr_mr.get('optimum_production', 0):,.0f}")
            mr_pd = tr_mr.get("production_delta", 0)
            _write_kpi_card(
                ws_exec,
                mr_row,
                3,
                "MR Prod. Delta",
                f"€{mr_pd:+,.0f}",
                delta_str=f"{tr_mr.get('production_delta_pct', 0):+.1f}%",
                delta_positive=mr_pd >= 0,
            )
            mr_rd = tr_mr.get("risk_delta_pct", 0)
            _write_kpi_card(
                ws_exec,
                mr_row,
                5,
                "MR Optimum Risk",
                f"{tr_mr.get('optimum_risk_pct', 0):.2f}%",
                delta_str=f"{mr_rd:+.2f} pp",
                delta_positive=mr_rd <= 0,
            )
            _write_kpi_card(
                ws_exec, mr_row, 7, "MR Rejection Rate", f"{tr_mr.get('optimum_rejection_rate_pct', 0):.1f}%"
            )
            ws_exec.cell(row=mr_row, column=9).value = "MR PERIOD"
            ws_exec.cell(row=mr_row, column=9).font = Font(bold=True, color=_CLR_MR_FG, size=9, name=_FN)
            ws_exec.cell(row=mr_row, column=9).alignment = _ALIGN_LEFT
            # Tint MR KPI row backgrounds
            for c in range(1, 9):
                for r in (mr_row, mr_row + 1):
                    cell = ws_exec.cell(row=r, column=c)
                    cell.fill = _FILL_MR
        else:
            ws_exec.cell(row=mr_row, column=1).value = "MR period data not available"
            ws_exec.cell(row=mr_row, column=1).font = Font(italic=True, color=_CLR_NEUTRAL_MID, size=10, name=_FN)

        # --- Spacer ---
        ws_exec.row_dimensions[8].height = 10
        next_row = 9

        # --- Main-period summary table ---
        exec_cols = [
            "group",
            "actual_production",
            "optimum_production",
            "production_delta",
            "production_delta_pct",
            "actual_risk_pct",
            "optimum_risk_pct",
            "risk_delta_pct",
        ]
        main_base_mask = (consolidated_df["period"] == "main") & (consolidated_df["scenario"] == "base")
        exec_main = _prepare_export_df(
            consolidated_df.loc[main_base_mask],
            exec_cols,
        )
        if not exec_main.empty:
            next_row = _write_exec_table(ws_exec, exec_main, next_row, "Main Period — Base Scenario")

        # --- MR-period summary table ---
        mr_base_mask = (consolidated_df["period"] == "mr") & (consolidated_df["scenario"] == "base")
        exec_mr = _prepare_export_df(
            consolidated_df.loc[mr_base_mask],
            exec_cols,
        )
        if not exec_mr.empty:
            next_row = _write_exec_table(ws_exec, exec_mr, next_row, "MR Period — Base Scenario")

        total_overview_cols = [
            "group",
            "period",
            "scenario",
            "actual_production",
            "optimum_production",
            "production_delta",
            "production_delta_pct",
            "actual_risk_pct",
            "optimum_risk_pct",
            "risk_delta_pct",
            "actual_rejection_rate_pct",
            "optimum_rejection_rate_pct",
        ]
        total_overview_df = _prepare_export_df(
            consolidated_df[consolidated_df["group"] == "TOTAL"],
            total_overview_cols,
        )
        if not total_overview_df.empty:
            next_row = _write_exec_table(
                ws_exec, total_overview_df, next_row, "Scenario Overview — Total Portfolio", n_table_cols=12
            )

        main_top_movers = _build_top_movers_df(consolidated_df, period="main")
        if not main_top_movers.empty:
            next_row = _write_exec_table(
                ws_exec, main_top_movers, next_row, "Top Segment Opportunities — Main Base Scenario"
            )

        mr_top_movers = _build_top_movers_df(consolidated_df, period="mr")
        if not mr_top_movers.empty:
            next_row = _write_exec_table(
                ws_exec, mr_top_movers, next_row, "Top Segment Opportunities — MR Base Scenario"
            )

        # --- Acceptance grids per segment on Executive Summary ---
        cutoff_data: dict[str, pd.DataFrame] = {}
        for seg_name in segments:
            csv_path = output_base / seg_name / "data" / "cutoff_summary_wide.csv"
            if not csv_path.exists():
                continue
            try:
                df_cut = pd.read_csv(csv_path)
            except (pd.errors.ParserError, OSError, ValueError):
                continue
            if df_cut.empty:
                continue
            if "segment" not in df_cut.columns:
                df_cut["segment"] = seg_name
            cutoff_data[seg_name] = df_cut

        if cutoff_data:
            for seg_name, df_cut in cutoff_data.items():
                next_row = _write_acceptance_grid(ws_exec, df_cut, seg_name, next_row)

        # Ensure KPI columns are wide enough
        for c in range(1, 11):
            cur = ws_exec.column_dimensions[get_column_letter(c)].width or 12
            ws_exec.column_dimensions[get_column_letter(c)].width = max(cur, 18)
        _apply_page_setup(ws_exec)

        # =============================================================
        # Sheet 2: Portfolio Summary
        # =============================================================
        portfolio_mask = consolidated_df["group"].str.match(r"^(TOTAL|supersegment_)")
        portfolio_cols = [
            "group",
            "period",
            "scenario",
            "n_segments",
            "actual_production",
            "optimum_production",
            "production_delta",
            "production_delta_pct",
            "actual_risk_pct",
            "optimum_risk_pct",
            "risk_delta_pct",
            "actual_rejection_rate_pct",
            "optimum_rejection_rate_pct",
            "total_demand",
            "production_ci_lower",
            "production_ci_upper",
            "risk_ci_lower",
            "risk_ci_upper",
        ]
        portfolio_df = _prepare_export_df(consolidated_df[portfolio_mask], portfolio_cols)
        if not portfolio_df.empty:
            portfolio_df.to_excel(writer, sheet_name="Portfolio Summary", index=False)
            _style_table(writer.sheets["Portfolio Summary"], portfolio_df.columns)
            writer.sheets["Portfolio Summary"].sheet_properties.tabColor = _CLR_TAB_PORTFOLIO
            _apply_page_setup(writer.sheets["Portfolio Summary"])

        # =============================================================
        # Sheet 3: Segment Detail
        # =============================================================
        segment_mask = ~consolidated_df["group"].str.match(r"^(TOTAL|supersegment_)")
        segment_cols = [
            "group",
            "period",
            "scenario",
            "segments",
            "actual_production",
            "optimum_production",
            "production_delta",
            "production_delta_pct",
            "actual_risk_pct",
            "optimum_risk_pct",
            "risk_delta_pct",
            "actual_risk_h3_pct",
            "optimum_risk_h3_pct",
            "actual_rejection_rate_pct",
            "optimum_rejection_rate_pct",
            "swap_in_production",
            "swap_out_production",
        ]
        segment_df = _prepare_export_df(consolidated_df[segment_mask], segment_cols)
        if not segment_df.empty:
            segment_df.to_excel(writer, sheet_name="Segment Detail", index=False)
            _style_table(writer.sheets["Segment Detail"], segment_df.columns)
            writer.sheets["Segment Detail"].sheet_properties.tabColor = _CLR_TAB_SEGMENT
            _apply_page_setup(writer.sheets["Segment Detail"])

        # =============================================================
        # Sheet 4: Cutoff Comparison (raw data table)
        # =============================================================
        if cutoff_data:
            all_cutoffs = pd.concat(cutoff_data.values(), ignore_index=True)
            all_cutoffs.to_excel(writer, sheet_name="Cutoff Comparison", index=False)
            _style_table(writer.sheets["Cutoff Comparison"], all_cutoffs.columns, highlight_total=False)
            writer.sheets["Cutoff Comparison"].sheet_properties.tabColor = _CLR_TAB_CUTOFF
            _apply_page_setup(writer.sheets["Cutoff Comparison"])

        # =============================================================
        # Per-segment acceptance grid sheets
        # =============================================================
        for seg_name, df_cut in cutoff_data.items():
            sheet_name = f"Grid {seg_name}"[:31]
            ws_grid = wb.create_sheet(sheet_name)
            ws_grid.sheet_properties.tabColor = _CLR_TAB_GRID
            ws_grid.sheet_view.showGridLines = False

            # Title bar
            title_fill = PatternFill(start_color=_CLR_PRIMARY, end_color=_CLR_PRIMARY, fill_type="solid")
            ws_grid.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
            t = ws_grid.cell(row=1, column=1)
            t.value = f"  Acceptance Grid — {seg_name}"
            t.font = Font(bold=True, color=_CLR_WHITE, size=16, name=_FN)
            t.fill = title_fill
            t.alignment = _ALIGN_LEFT
            ws_grid.row_dimensions[1].height = 36
            for gc in range(1, 15):
                ws_grid.cell(row=2, column=gc).border = Border(top=Side(style="medium", color=_CLR_ACCENT))
            ws_grid.row_dimensions[2].height = 4

            # Draw grids per scenario
            cur_row = 4
            for scen in ["pessimistic", "base", "optimistic"]:
                if "scenario" in df_cut.columns and scen in df_cut["scenario"].values:
                    cur_row = _write_acceptance_grid(ws_grid, df_cut, f"{scen.title()}", cur_row, scenario=scen)
            _apply_page_setup(ws_grid)

        # =============================================================
        # Per-segment RP summary sheets (Main + MR)
        # =============================================================
        _rp_exclude_cols = {"todu_30ever_h6", "todu_amt_pile_h6", "Total Demand (€)"}
        _rp_exclude_cols_mr = _rp_exclude_cols | {"todu_30ever_h3", "todu_amt_pile_h3"}

        for seg_name in segments:
            seg_settings = _load_segment_settings(output_base, seg_name)

            # Resolve variables list — fall back to global config
            try:
                import tomllib as _tomllib

                _seg_cfg_path = output_base / seg_name / "config_segment.toml"
                if _seg_cfg_path.exists():
                    _scfg = _tomllib.loads(_seg_cfg_path.read_text(encoding="utf-8"))
                    _prep = _scfg.get("preprocessing", _scfg)
                    seg_vars_full = _prep.get("variables") or _prep.get("inference_variables")
                    if seg_vars_full:
                        seg_vars_full = list(seg_vars_full)
                    else:
                        seg_vars_full = None
                else:
                    seg_vars_full = None
            except Exception:
                logger.warning(
                    f"Failed to read variables from segment config for '{seg_name}'; "
                    f"will fall back to segments.toml / config.toml",
                    exc_info=True,
                )
                seg_vars_full = None
            # Fall back to global variables from segments.toml / config.toml
            if not seg_vars_full:
                seg_cfg = segments.get(seg_name, {})
                seg_vars_full = seg_cfg.get("variables")
            if not seg_vars_full:
                try:
                    import tomllib as _tomllib

                    _gcfg_path = Path("config.toml")
                    if _gcfg_path.exists():
                        _gcfg = _tomllib.loads(_gcfg_path.read_text(encoding="utf-8"))
                        seg_vars_full = _gcfg.get("preprocessing", {}).get("variables")
                except Exception:
                    logger.warning(
                        f"Failed to read variables from global config.toml for '{seg_name}'; "
                        f"will use hardcoded default list",
                        exc_info=True,
                    )
            if not seg_vars_full:
                # All config sources unreadable; use the documented fallback.
                seg_vars_full = list(_FALLBACK_REPORTING_VARIABLES)

            # Resolve income threshold for labels
            _income_th = None
            try:
                import tomllib as _tomllib

                _seg_cfg_path2 = output_base / seg_name / "config_segment.toml"
                if _seg_cfg_path2.exists():
                    _scfg2 = _tomllib.loads(_seg_cfg_path2.read_text(encoding="utf-8"))
                    _edges = _scfg2.get("preprocessing", _scfg2).get("bins", {}).get("income_bin", {}).get("bin_edges")
                    if _edges and isinstance(_edges, list) and len(_edges) >= 3:
                        finite = [float(e) for e in _edges if np.isfinite(float(e))]
                        if len(finite) == 1:
                            _income_th = finite[0]
            except Exception:
                logger.warning(
                    f"Failed to resolve income threshold from segment config for '{seg_name}'; "
                    f"will try reporting supersegment edges next",
                    exc_info=True,
                )
            if _income_th is None:
                # Try reporting supersegment edges
                _rs = segments.get(seg_name, {}).get("reporting_supersegment") or segments.get(seg_name, {}).get(
                    "supersegment"
                )
                if _rs and _rs in supersegments:
                    _edges = supersegments[_rs].get("bin_edges", {}).get("income_bin")
                    if _edges and isinstance(_edges, list) and len(_edges) >= 3:
                        finite = [float(e) for e in _edges if np.isfinite(float(e))]
                        if len(finite) == 1:
                            _income_th = finite[0]

            # --- Main period ---
            csv_path = output_base / seg_name / "data" / "risk_production_summary_table_base.csv"
            if not csv_path.exists():
                continue
            try:
                df_rp = pd.read_csv(csv_path)
            except (pd.errors.ParserError, OSError, ValueError):
                continue
            if df_rp.empty:
                continue
            df_rp = df_rp.drop(columns=[c for c in _rp_exclude_cols if c in df_rp.columns])
            rp_extra_tables = _build_income_bin_tables(seg_name, period="main", template_cols=list(df_rp.columns))

            # Build classification grid for main period
            main_grid = _build_classification_grid(
                seg_name,
                period="main",
                variables=seg_vars_full,
                multiplier=seg_settings["multiplier"],
                multiplier_h3=seg_settings.get("multiplier_h3"),
                inv_vars=seg_settings.get("inv_vars"),
                income_threshold=_income_th,
            )

            sheet_name = f"RP {seg_name}"[:31]
            _write_rp_sheet(
                writer,
                df_rp,
                sheet_name,
                seg_name,
                "Main Period — Base Scenario",
                _CLR_TAB_SEGMENT,
                extra_tables=rp_extra_tables,
                classification_grid=main_grid,
                is_mr=False,
            )

            # --- MR period ---
            mr_csv_path = output_base / seg_name / "data" / "risk_production_summary_table_mr_base.csv"
            if not mr_csv_path.exists():
                mr_csv_path = output_base / seg_name / "data" / "risk_production_summary_table_mr.csv"
            if not mr_csv_path.exists():
                continue
            try:
                df_mr = pd.read_csv(mr_csv_path)
            except (pd.errors.ParserError, OSError, ValueError):
                continue
            if df_mr.empty:
                continue
            df_mr = df_mr.drop(columns=[c for c in _rp_exclude_cols_mr if c in df_mr.columns])
            mr_extra_tables = _build_income_bin_tables(seg_name, period="mr", template_cols=list(df_mr.columns))

            # Build classification grid for MR period (volumes only)
            mr_grid = _build_classification_grid(
                seg_name,
                period="mr",
                variables=seg_vars_full,
                multiplier=seg_settings["multiplier"],
                multiplier_h3=seg_settings.get("multiplier_h3"),
                inv_vars=seg_settings.get("inv_vars"),
                income_threshold=_income_th,
            )

            mr_sheet_name = f"RP MR {seg_name}"[:31]
            _write_rp_sheet(
                writer,
                df_mr,
                mr_sheet_name,
                seg_name,
                "MR Period — Base Scenario",
                _CLR_TAB_SEGMENT_MR,
                extra_tables=mr_extra_tables,
                classification_grid=mr_grid,
                is_mr=True,
            )

        # Remove default "Sheet" if auto-created
        if "Sheet" in wb.sheetnames and len(wb.sheetnames) > 1:
            del wb["Sheet"]

    logger.debug(f"Excel workbook written to {xlsx_path}")
    return xlsx_path


def print_consolidation_summary(df: pd.DataFrame) -> None:
    """Log a formatted summary of consolidated metrics via loguru.

    (Name retained for backward compatibility; output is now routed through
    the logger so it respects --log-file and log-level filtering. todo #61.)
    """
    logger.info("\n" + "=" * 80)
    logger.info("CONSOLIDATED RISK PRODUCTION SUMMARY")
    logger.info("=" * 80)

    # Get unique scenarios and periods
    scenarios = df["scenario"].unique()

    for scenario in scenarios:
        logger.info(f"\n{'─' * 40}")
        logger.info(f"SCENARIO: {scenario}")
        logger.info(f"{'─' * 40}")

        scenario_df = df[df["scenario"] == scenario]

        # Show TOTAL rows
        total_df = scenario_df[scenario_df["group"] == "TOTAL"]

        for _, row in total_df.iterrows():
            period = row["period"].upper()
            logger.info(f"\n  {period} Period:")
            logger.info(f"    Actual Production:  €{row['actual_production']:,.0f}")
            logger.info(f"    Optimum Production: €{row['optimum_production']:,.0f}")
            logger.info(f"    Delta:              €{row['production_delta']:,.0f} ({row['production_delta_pct']:.1f}%)")
            logger.info(f"    Risk:               {row['actual_risk_pct']:.2f}% → {row['optimum_risk_pct']:.2f}%")

        # Show supersegment breakdown
        ss_df = scenario_df[scenario_df["group"].str.startswith("supersegment_")]
        if not ss_df.empty:
            logger.info("\n  By Supersegment (Main Period):")
            main_ss = ss_df[ss_df["period"] == "main"]
            for _, row in main_ss.iterrows():
                ss_name = row["group"].replace("supersegment_", "")
                logger.info(
                    f"    {ss_name}: €{row['actual_production']:,.0f} → €{row['optimum_production']:,.0f} "
                    f"({row['production_delta_pct']:+.1f}%), Risk: {row['actual_risk_pct']:.2f}% → {row['optimum_risk_pct']:.2f}%"
                )

    logger.info("\n" + "=" * 80)
